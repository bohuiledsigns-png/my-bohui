"""GLOWFORGE CRM — Flask 主程序"""
import os
import sys
import json
import uuid
import threading
import time
import random
import signal
import atexit
import socket
import subprocess
import requests as http_requests
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
from flask import Flask, request, jsonify, render_template, send_from_directory, session
from datetime import datetime
import io
import openpyxl

from database import init_db, init_ai_evolution, get_stats, get_revenue_trend, get_chart_data, get_workbench, get_customers, get_customer, get_customer_detail, add_customer, update_customer, delete_customer, bulk_add_customers, add_activity_log, get_activity_logs, get_activity_logs_count
from database import get_messages, add_message, get_media, add_media, delete_media
from database import get_ai_generations, delete_ai_generation, add_ai_generation, get_ai_generation_stats
from database import get_email_settings, save_email_settings, add_email_log, get_email_log, get_all_email_log
from database import get_products, get_product, add_product, update_product, delete_product, get_product_categories
from database import get_cases, get_case, add_case, update_case, delete_case, get_case_categories
from database import get_quotes, get_quote, add_quote, update_quote, delete_quote
from database import get_users, get_user_by_username, add_user, update_user
from database import get_orders, get_order, add_order, update_order, delete_order
from database import add_timeline_entry, get_payment_dashboard, get_order_profit_stats, get_order_stats, get_commission_stats, get_production_schedule, get_production_tasks, save_production_tasks, update_production_task_status, get_production_task_defaults, get_shipments, get_shipment, add_shipment, update_shipment, delete_shipment
from database import get_monthly_sales_detail, get_customer_acquisition_stats, get_production_stats
from database import get_qc_templates, get_qc_template, save_qc_template, delete_qc_template, get_order_qc_inspections, get_qc_inspection, add_qc_inspection, update_qc_inspection, delete_qc_inspection
from database import add_payment, get_ar_summary, get_ar_by_customer, get_payment_history, get_aging_analysis, migrate_payments_from_orders
from database import get_leads, get_lead_summary, get_lead_funnel, assign_lead, update_lead_status, update_lead_source
from database import get_leads_due_followup, get_today_followup_summary, update_last_contacted
from database import get_users, unassign_lead, claim_lead, reclaim_expired_leads, get_leads_with_pool_info
from database import add_notification, get_notifications, get_unread_count, mark_notification_read, mark_all_notifications_read, get_user_by_id, get_active_user_ids
from database import get_media_tags, add_media_tag, delete_media_tag, get_media_tags_for, update_media_tags, get_media_by_tag
from database import get_inventory_items, get_inventory_item, add_inventory_item, update_inventory_item, delete_inventory_item
from database import add_stock_movement, get_stock_movements, get_inventory_summary
from database import get_partners, get_partner, add_partner, update_partner, delete_partner
from database import get_purchase_orders, get_purchase_order, add_purchase_order, update_purchase_order, delete_purchase_order, add_po_timeline_entry
from ai_engine import translate, generate_image, generate_video, VIDEO_PRESETS, analyze_customer_message, summarize_chat, analyze_viral, rewrite_copy
from ai_engine import get_country_context, generate_customized_script, search_industry_knowledge, analyze_chat_history, parse_whatsapp_export, analyze_product_image
from ai_engine import ask_ali, get_ai_greeting, get_ai_followup_message, clear_knowledge_base_cache
from ai_engine import generate_sign_prompt, generate_sign_prompt_and_image, generate_sign_prompt_ali, generate_storefront_promo_video
from catalog_generator import generate_catalog
from scripts.generate_product_catalog import generate_product_catalog
from whatsapp_engine import send_text, send_image_clipboard, send_media_file, read_messages, start_monitor, stop_monitor, get_monitor_status, refresh_whatsapp_page, set_remote_server, get_qr_base64, is_logged_in
from lead_state_engine import update_lead_state, get_lead_state, init_customer_state
from decision_engine import decide_action
from action_router import execute_action, register_action, log_action

# Publishing Manager（多平台视频发布管理系统）
from publishing_manager import init_publishing

app = Flask(__name__)
app.secret_key = os.environ.get("FLASK_SECRET_KEY", "glowforge-crm-2026-secret-key-change-in-production")
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
UPLOAD_DIR = os.path.join(BASE_DIR, "uploads")
os.makedirs(UPLOAD_DIR, exist_ok=True)

# 自动连接独立 WhatsApp 服务（gunicorn 模式也适用）
def _auto_connect_wa():
    import urllib.request, json as _j
    try:
        with urllib.request.urlopen("http://127.0.0.1:15789/health", timeout=2) as r:
            if _j.loads(r.read().decode()).get("ok"):
                set_remote_server("http://127.0.0.1:15789")
    except:
        pass
threading.Thread(target=_auto_connect_wa, daemon=True).start()

# ========= AI 风格学习 =========
# 你手动发的消息会被保存下来，AI 学习你的语气
STYLE_FILE = os.path.join(BASE_DIR, ".communication_style.json")

def _save_style_sample(cn, en):
    import json
    samples = []
    if os.path.exists(STYLE_FILE):
        with open(STYLE_FILE, "r", encoding="utf-8") as f:
            samples = json.load(f)
    samples.append({"cn": cn, "en": en, "time": time.strftime("%Y-%m-%d %H:%M:%S")})
    if len(samples) > 30:
        samples = samples[-30:]
    with open(STYLE_FILE, "w", encoding="utf-8") as f:
        json.dump(samples, f, ensure_ascii=False, indent=2)

def _get_style_samples(limit=5):
    import json
    if os.path.exists(STYLE_FILE):
        with open(STYLE_FILE, "r", encoding="utf-8") as f:
            samples = json.load(f)
        return samples[-limit:]
    return []

# ========= 销售个人资料 =========
PROFILE_FILE = os.path.join(BASE_DIR, ".sales_profile.json")
_DEFAULT_PROFILE = {"name": "Philip", "title": "外贸销售", "company": "Bohui GLOWFORGE"}

def _get_profile():
    import json
    if os.path.exists(PROFILE_FILE):
        with open(PROFILE_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return dict(_DEFAULT_PROFILE)

def _save_profile(data):
    import json
    profile = dict(_DEFAULT_PROFILE)
    profile.update(data)
    with open(PROFILE_FILE, "w", encoding="utf-8") as f:
        json.dump(profile, f, ensure_ascii=False, indent=2)
    return profile

# ========= 知识库 CRUD =========
KNOWLEDGE_DIR = os.path.join(BASE_DIR, 'knowledge')
SCRIPTS_DIR = os.path.join(BASE_DIR, 'scripts')
COUNTRIES_DIR = os.path.join(BASE_DIR, 'countries')

def _read_content_dir(directory):
    """列出目录中的txt文件"""
    files = []
    os.makedirs(directory, exist_ok=True)
    for f in sorted(os.listdir(directory)):
        if f.endswith(".txt"):
            path = os.path.join(directory, f)
            with open(path, "r", encoding="utf-8") as fp:
                content = fp.read()
            lines = content.strip().split("\n")
            title = lines[0] if lines else f
            category = ""
            for line in lines:
                if "类别:" in line:
                    category = line.split("类别:")[-1].strip()
                    break
            files.append({"name": f, "title": title, "category": category, "size": os.path.getsize(path)})
    return files

def _read_file_or_404(directory, filename):
    path = os.path.join(directory, filename)
    if not os.path.exists(path):
        return None
    with open(path, "r", encoding="utf-8") as f:
        return f.read()

def _write_content_file(directory, name, content):
    if not name.endswith(".txt"):
        name += ".txt"
    path = os.path.join(directory, name)
    if os.path.exists(path):
        return None  # conflict
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        f.write(content)
    return True

def _update_content_file(directory, filename, content):
    path = os.path.join(directory, filename)
    if not os.path.exists(path):
        return False
    with open(path, "w", encoding="utf-8") as f:
        f.write(content)
    return True

def _delete_content_file(directory, filename):
    path = os.path.join(directory, filename)
    if os.path.exists(path):
        os.remove(path)
    return True

# 知识库
@app.route("/api/knowledge", endpoint="knowledge_list")
def api_knowledge_list():
    return jsonify(_read_content_dir(KNOWLEDGE_DIR))

@app.route("/api/knowledge/<path:filename>", endpoint="knowledge_get")
def api_knowledge_get(filename):
    content = _read_file_or_404(KNOWLEDGE_DIR, filename)
    if content is None:
        return jsonify({"error": "not found"}), 404
    return jsonify({"name": filename, "content": content})

@app.route("/api/knowledge", methods=["POST"], endpoint="knowledge_create")
def api_knowledge_create():
    data = request.json
    name = data.get("name", "").strip()
    content = data.get("content", "").strip()
    if not name or not content:
        return jsonify({"error": "name and content required"}), 400
    result = _write_content_file(KNOWLEDGE_DIR, name, content)
    if result is None:
        return jsonify({"error": "已存在同名文件"}), 409
    return jsonify({"ok": True, "name": name})

@app.route("/api/knowledge/<path:filename>", methods=["PUT"], endpoint="knowledge_update")
def api_knowledge_update(filename):
    data = request.json
    if not _update_content_file(KNOWLEDGE_DIR, filename, data.get("content", "")):
        return jsonify({"error": "not found"}), 404
    return jsonify({"ok": True})

@app.route("/api/knowledge/<path:filename>", methods=["DELETE"], endpoint="knowledge_delete")
def api_knowledge_delete(filename):
    _delete_content_file(KNOWLEDGE_DIR, filename)
    return jsonify({"ok": True})

# 话术库
@app.route("/api/scripts", endpoint="scripts_list")
def api_scripts_list():
    return jsonify(_read_content_dir(SCRIPTS_DIR))

@app.route("/api/scripts/<path:filename>", endpoint="scripts_get")
def api_scripts_get(filename):
    content = _read_file_or_404(SCRIPTS_DIR, filename)
    if content is None:
        return jsonify({"error": "not found"}), 404
    return jsonify({"name": filename, "content": content})

@app.route("/api/scripts", methods=["POST"], endpoint="scripts_create")
def api_scripts_create():
    data = request.json
    name = data.get("name", "").strip()
    content = data.get("content", "").strip()
    if not name or not content:
        return jsonify({"error": "name and content required"}), 400
    result = _write_content_file(SCRIPTS_DIR, name, content)
    if result is None:
        return jsonify({"error": "已存在同名文件"}), 409
    return jsonify({"ok": True, "name": name})

@app.route("/api/scripts/<path:filename>", methods=["PUT"], endpoint="scripts_update")
def api_scripts_update(filename):
    data = request.json
    if not _update_content_file(SCRIPTS_DIR, filename, data.get("content", "")):
        return jsonify({"error": "not found"}), 404
    return jsonify({"ok": True})

@app.route("/api/scripts/<path:filename>", methods=["DELETE"], endpoint="scripts_delete")
def api_scripts_delete(filename):
    _delete_content_file(SCRIPTS_DIR, filename)
    return jsonify({"ok": True})

# 国家档案
@app.route("/api/countries", endpoint="countries_list")
def api_countries_list():
    return jsonify(_read_content_dir(COUNTRIES_DIR))

@app.route("/api/countries/<path:filename>", endpoint="countries_get")
def api_countries_get(filename):
    content = _read_file_or_404(COUNTRIES_DIR, filename)
    if content is None:
        return jsonify({"error": "not found"}), 404
    return jsonify({"name": filename, "content": content})

@app.route("/api/countries", methods=["POST"], endpoint="countries_create")
def api_countries_create():
    data = request.json
    name = data.get("name", "").strip()
    content = data.get("content", "").strip()
    if not name or not content:
        return jsonify({"error": "name and content required"}), 400
    result = _write_content_file(COUNTRIES_DIR, name, content)
    if result is None:
        return jsonify({"error": "已存在同名文件"}), 409
    return jsonify({"ok": True, "name": name})

@app.route("/api/countries/<path:filename>", methods=["PUT"], endpoint="countries_update")
def api_countries_update(filename):
    data = request.json
    if not _update_content_file(COUNTRIES_DIR, filename, data.get("content", "")):
        return jsonify({"error": "not found"}), 404
    return jsonify({"ok": True})

@app.route("/api/countries/<path:filename>", methods=["DELETE"], endpoint="countries_delete")
def api_countries_delete(filename):
    _delete_content_file(COUNTRIES_DIR, filename)
    return jsonify({"ok": True})

# 初始化数据库
init_db()
init_ai_evolution()

# V3 自优化销售系统：加载优化后的权重
try:
    from revenue_engine import load_v3_optimizations
    load_v3_optimizations()
except Exception:
    pass

# V4 自动销售系统：加载动态定价覆盖
try:
    from ai_engine.dynamic_pricing import load_v4_dynamic_anchors
    load_v4_dynamic_anchors()
except Exception:
    pass

# V4 自动销售系统：启动后台调度器
try:
    from ai_engine.revenue_scheduler import start_v4_scheduler_background
    start_v4_scheduler_background()
except Exception as e:
    print(f"[v4] Scheduler start failed: {e}")

# 发布管理系统：初始化数据库 + Blueprint + 后台调度线程
try:
    DB_PATH = os.path.join(BASE_DIR, "crm_data.db")
    init_publishing(app, DB_PATH)
    print("[Publishing] Manager initialized")
except Exception as e:
    print(f"[Publishing] Init failed: {e}")

# V5 Agent Competition系统：初始化数据库表
try:
    from database import init_v5_tables
    init_v5_tables()
except Exception as e:
    print(f"[v5] DB init failed: {e}")


# ========= 页面 =========
@app.route("/")
def index():
    return render_template("index.html")


# ========= API: 提示词库 =========
PROMPT_DIR = os.path.join(BASE_DIR, "prompts")

def _list_prompt_files(ptype="image"):
    """列出指定类型的提示词文件"""
    d = os.path.join(PROMPT_DIR, ptype)
    os.makedirs(d, exist_ok=True)
    files = []
    for f in sorted(os.listdir(d)):
        if f.endswith(".txt"):
            path = os.path.join(d, f)
            with open(path, "r", encoding="utf-8") as fp:
                content = fp.read()
            # 提取标题（第一行）
            lines = content.strip().split("\n")
            title = lines[0] if lines else f
            # 提取类别标签
            category = ""
            for line in lines:
                if "类别:" in line:
                    category = line.split("类别:")[-1].strip()
                    break
            files.append({
                "name": f,
                "title": title,
                "category": category,
                "size": os.path.getsize(path)
            })
    return files

@app.route("/api/prompts")
def api_prompts():
    ptype = request.args.get("type", "image")
    return jsonify(_list_prompt_files(ptype))

@app.route("/api/prompts/<path:filename>")
def api_prompt_content(filename):
    ptype = request.args.get("type", "image")
    path = os.path.join(PROMPT_DIR, ptype, filename)
    if not os.path.exists(path):
        return jsonify({"error": "not found"}), 404
    with open(path, "r", encoding="utf-8") as f:
        content = f.read()
    return jsonify({"name": filename, "content": content})

@app.route("/api/prompts", methods=["POST"])
def api_create_prompt():
    data = request.json
    name = data.get("name", "").strip()
    ptype = data.get("type", "image")
    content = data.get("content", "").strip()
    if not name or not content:
        return jsonify({"error": "name and content required"}), 400
    if not name.endswith(".txt"):
        name += ".txt"
    path = os.path.join(PROMPT_DIR, ptype, name)
    if os.path.exists(path):
        return jsonify({"error": "已存在同名提示词"}), 409
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        f.write(content)
    return jsonify({"ok": True, "name": name})

@app.route("/api/prompts/<path:filename>", methods=["PUT"])
def api_update_prompt(filename):
    ptype = request.args.get("type", "image")
    path = os.path.join(PROMPT_DIR, ptype, filename)
    if not os.path.exists(path):
        return jsonify({"error": "not found"}), 404
    data = request.json
    content = data.get("content", "")
    with open(path, "w", encoding="utf-8") as f:
        f.write(content)
    return jsonify({"ok": True})

@app.route("/api/prompts/<path:filename>", methods=["DELETE"])
def api_delete_prompt(filename):
    ptype = request.args.get("type", "image")
    path = os.path.join(PROMPT_DIR, ptype, filename)
    if os.path.exists(path):
        os.remove(path)
    return jsonify({"ok": True})


@app.route("/api/prompts/generate-from-requirements", methods=["POST"])
def api_generate_from_requirements():
    """AI根据客户需求生成招牌可视化提示词"""
    data = request.json
    industry = data.get("industry", "")
    sign_text = data.get("sign_text", "")
    reference = data.get("reference", "")
    material = data.get("material", "")
    scene = data.get("scene", "")
    image_data = data.get("image_data", "")

    if not sign_text and not reference and not image_data:
        return jsonify({"error": "请至少提供招牌文字、参考描述或上传图片"}), 400

    do_preview = data.get("preview", False)
    image_engine = data.get("image_engine", "volc")

    if do_preview:
        if image_engine == "ali":
            result, image_url, err = generate_sign_prompt_ali(industry, sign_text, reference, material, scene, image_data)
        else:
            result, image_url, err = generate_sign_prompt_and_image(industry, sign_text, reference, material, scene, image_data)
        if err:
            return jsonify({"error": err}), 500
        return jsonify({"result": result, "image_url": image_url})
    else:
        result, err = generate_sign_prompt(industry, sign_text, reference, material, scene, image_data)
        if err:
            return jsonify({"error": err}), 500
        return jsonify({
            "result": result
        })


# ========= API: 统计 =========
@app.route("/api/stats")
def api_stats():
    stats = get_stats()
    # 追加沉睡客户数量（最后一条消息是我们发的，且超过3天没动静）
    from database import get_db
    conn = get_db()
    dormant = conn.execute("""
        SELECT COUNT(*) FROM customers c
        WHERE (
            SELECT direction FROM messages m
            WHERE m.customer_id=c.id ORDER BY m.created_at DESC LIMIT 1
        ) = 'sent'
        AND CAST(julianday('now') - julianday(c.updated_at) AS INTEGER) >= 3
    """).fetchone()[0]
    conn.close()
    stats["dormant"] = dormant
    return jsonify(stats)

# ========= 系统存活监控 =========
@app.route("/api/health")
def api_health():
    """系统健康检查"""
    wa_ok = False
    try:
        import urllib.request
        with urllib.request.urlopen("http://127.0.0.1:15789/health", timeout=5) as r:
            wa_ok = r.status == 200
    except Exception:
        wa_ok = False
    return jsonify({
        "status": "ok",
        "timestamp": datetime.now().isoformat(),
        "whatsapp_server": "connected" if wa_ok else "offline",
        "db": "ok",
        "version": "GLOWFORGE CRM V5",
    })

@app.route("/api/monitor")
def api_monitor():
    """返回最近20条监控日志"""
    try:
        log_path = os.path.join(BASE_DIR, ".whatsapp_session", "monitor.log")
        if not os.path.exists(log_path):
            return jsonify({"lines": []})
        with open(log_path, "r", encoding="utf-8") as f:
            lines = f.readlines()
        return jsonify({"lines": lines[-20:]})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/stale-customers")
def api_stale_customers():
    """沉睡客户：我们报价/发了消息后，客户超过3天没回复"""
    from database import get_db
    conn = get_db()
    rows = conn.execute("""
        SELECT c.id, c.name, c.company, c.status,
               CAST(julianday('now') - julianday(COALESCE(c.last_contacted_at, c.updated_at)) AS INTEGER) as days_since,
               (SELECT content_en FROM messages m WHERE m.customer_id=c.id ORDER BY m.created_at DESC LIMIT 1) as last_sent,
               (SELECT created_at FROM messages m WHERE m.customer_id=c.id ORDER BY m.created_at DESC LIMIT 1) as last_time
        FROM customers c
        WHERE (
            SELECT direction FROM messages m
            WHERE m.customer_id=c.id ORDER BY m.created_at DESC LIMIT 1
        ) = 'sent'
        AND CAST(julianday('now') - julianday(c.updated_at) AS INTEGER) >= 3
        ORDER BY days_since DESC
    """).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route("/api/follow-up/<int:cid>", methods=["POST"])
def api_follow_up(cid):
    """给沉睡客户发一条跟进消息"""
    from database import get_db
    conn = get_db()
    c = conn.execute("SELECT name, language FROM customers WHERE id=?", (cid,)).fetchone()
    conn.close()
    if not c:
        return jsonify({"error": "客户不存在"}), 404

    contact_name = c["name"]
    lang = c["language"] if c["language"] else "English"

    # 根据天数选择不同的话术
    from database import get_messages
    msgs = get_messages(cid, limit=5)
    days = 0
    if msgs:
        from datetime import datetime
        last = msgs[-1].get("created_at", "")
        if last:
            try:
                dt = datetime.strptime(last, "%Y-%m-%d %H:%M:%S")
                days = (datetime.now() - dt).days
            except:
                pass
    # AI生成跟进话术
    from ai_engine import ask_ali
    follow_prompt = f"""你叫Philip，是博汇Bohui GLOWFORGE工厂的销售。
你在跟进一个{lang}市场的客户「{contact_name}」，上次报价后对方没有回复。
目前已经{max(days, 3)}天没有联系了。

请用{lang}写一条简短的跟进消息，要求：
1. 语气友好自然，不要pushy
2. 简短2-3句，不要写小作文
3. 像是真人销售随手发的，不是模板
4. 不要用"I hope this message finds you well"等正式腔
5. 可以顺带提一下上次聊的内容（如果有）

直接输出跟进消息，不要解释："""
    try:
        reply = ask_ali(follow_prompt, f"最近聊天：{[(m.get('content_cn','') or m.get('content_en',''))[:80] for m in msgs[-3:]]}", max_tokens=500, timeout=30)
    except Exception as e:
        print(f"[FollowUp] AI调用失败: {e}")
        reply = None
    if not reply:
        reply = f"Hi {contact_name}, just checking in — have you had a chance to review the quote? Happy to answer any questions."

    # 发送
    try:
        from whatsapp_engine import send_text
        send_text(reply, contact_name=contact_name)
        from database import add_message
        add_message(cid, "sent", f"[跟进] {reply[:60]}...", reply)
        _add_wa_activity("replied", cid, contact_name, f"已发送跟进消息")
        return jsonify({"ok": True, "reply": reply})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

@app.route("/api/workbench")
def api_workbench():
    return jsonify(get_workbench())


@app.route("/api/stats/revenue-trend")
def api_revenue_trend():
    months = request.args.get("months", 12, type=int)
    return jsonify(get_revenue_trend(months))


@app.route("/api/stats/chart-data")
def api_chart_data():
    return jsonify(get_chart_data())


# ========= API: 操作日志 =========
@app.route("/api/activity-logs")
def api_activity_logs():
    limit = request.args.get("limit", 50, type=int)
    offset = request.args.get("offset", 0, type=int)
    logs = get_activity_logs(limit, offset)
    total = get_activity_logs_count()
    return jsonify({"logs": logs, "total": total})


# ========= API: 客户 =========
@app.route("/api/customers")
def api_customers():
    return jsonify(get_customers())

@app.route("/api/customers/<int:cid>")
def api_customer(cid):
    c = get_customer(cid)
    return jsonify(c) if c else (jsonify({"error": "not found"}), 404)


@app.route("/api/customers/<int:cid>/detail")
def api_customer_detail(cid):
    """聚合客户详情"""
    detail = get_customer_detail(cid)
    if not detail:
        return jsonify({"error": "not found"}), 404
    return jsonify(detail)


@app.route("/api/customers", methods=["POST"])
def api_add_customer():
    data = request.json
    result = add_customer(
        name=data.get("name", ""),
        company=data.get("company", ""),
        whatsapp=data.get("whatsapp", ""),
        country=data.get("country", ""),
        language=data.get("language", "English"),
        status=data.get("status", "warm"),
        notes=data.get("notes", "")
    )
    uid = session.get("user_id")
    if uid and result.get("id"):
        add_activity_log(uid, "create", "customer", result["id"],
            f"创建了客户 {data.get('name','')} (ID:{result['id']})")
    return jsonify(result)

@app.route("/api/customers/<int:cid>", methods=["PUT"])
def api_update_customer(cid):
    old = get_customer(cid)
    update_customer(cid, **request.json)
    uid = session.get("user_id")
    if uid and old:
        add_activity_log(uid, "update", "customer", cid,
            f"修改了客户 {old.get('name','')} (ID:{cid})")
    return jsonify({"ok": True})

@app.route("/api/customers/<int:cid>", methods=["DELETE"])
def api_delete_customer(cid):
    old = get_customer(cid)
    delete_customer(cid)
    uid = session.get("user_id")
    if uid and old:
        add_activity_log(uid, "delete", "customer", cid,
            f"删除了客户 {old.get('name','')} (ID:{cid})")
    return jsonify({"ok": True})


# ========= API: 消息 =========
@app.route("/api/customers/<int:cid>/messages")
def api_messages(cid):
    limit = request.args.get("limit", 50, type=int)
    return jsonify(get_messages(cid, limit))

@app.route("/api/customers/<int:cid>/messages", methods=["POST"])
def api_send_message(cid):
    data = request.json
    content_cn = data.get("content_cn", "")
    content_en = data.get("content_en", "")
    direction = data.get("direction", "sent")

    add_message(cid, direction, content_cn, content_en)
    return jsonify({"ok": True})

# ==================== 订单 WhatsApp 通知 ====================
_ORDER_MSG_TEMPLATES = {
    "confirmed": {
        "en": "Dear {name}, your order {order_no} has been confirmed. We will start production soon. Thank you for your trust!",
        "zh": "亲爱的{name}，您的订单{order_no}已确认。我们将尽快安排生产。感谢您的信任！",
    },
    "in_production": {
        "en": "Dear {name}, your order {order_no} is now in production. We will keep you updated on progress.",
        "zh": "亲爱的{name}，您的订单{order_no}已进入生产阶段。我们会持续更新进度。",
    },
    "shipped": {
        "en": "Dear {name}, great news! Your order {order_no} has been shipped. We will share tracking information soon.",
        "zh": "亲爱的{name}，好消息！您的订单{order_no}已发货。物流信息将稍后提供。",
    },
    "delivered": {
        "en": "Dear {name}, your order {order_no} has been delivered. We hope you love it! Please let us know if you need anything.",
        "zh": "亲爱的{name}，您的订单{order_no}已完成交付。希望您满意！如有任何需要请随时联系我们。",
    },
    "payment_deposit": {
        "en": "Dear {name}, thank you! We have received your deposit of {currency}{amount} for order {order_no}.",
        "zh": "亲爱的{name}，感谢您！我们已收到您订单{order_no}的定金{currency}{amount}。",
    },
    "payment_balance": {
        "en": "Dear {name}, great news! We have received the final payment of {currency}{amount} for order {order_no}. Your order is now paid in full. Thank you!",
        "zh": "亲爱的{name}，好消息！我们已收到您订单{order_no}的尾款{currency}{amount}。您的订单已付清。感谢您的支持！",
    },
}

def _notify_order_status_change(order, customer, old_status, new_status):
    """Send WhatsApp notification to customer on order status change (run in bg thread)."""
    if not customer or not customer.get("whatsapp") or not customer.get("name"):
        return
    lang = (customer.get("language") or "English").lower()
    locale = "zh" if "zh" in lang or "chinese" in lang else "en"
    template = _ORDER_MSG_TEMPLATES.get(new_status, {}).get(locale)
    if not template:
        return
    text = template.format(name=customer["name"], order_no=order.get("order_no", ""), currency=order.get("currency", "USD"))
    try:
        send_text(text, contact_name=customer["name"])
        add_timeline_entry(order["id"], new_status, 0, f"[WhatsApp通知] 已通知客户订单状态变更: {old_status} → {new_status}")
    except Exception as e:
        print(f"[OrderNotify] WA send failed (oid={order['id']}, status={new_status}): {e}")

def _notify_payment_received(order, customer, ptype, amount):
    """Send WhatsApp notification to customer on payment received (run in bg thread)."""
    if not customer or not customer.get("whatsapp") or not customer.get("name"):
        return
    lang = (customer.get("language") or "English").lower()
    locale = "zh" if "zh" in lang or "chinese" in lang else "en"
    tkey = f"payment_{ptype}"
    template = _ORDER_MSG_TEMPLATES.get(tkey, {}).get(locale)
    if not template:
        return
    text = template.format(name=customer["name"], order_no=order.get("order_no", ""), amount=amount, currency=order.get("currency", "USD"))
    try:
        send_text(text, contact_name=customer["name"])
        label = "定金" if ptype == "deposit" else "尾款"
        add_timeline_entry(order["id"], "payment", 0, f"[WhatsApp通知] 已通知客户收到{label} {amount}")
    except Exception as e:
        print(f"[OrderNotify] WA send failed (oid={order['id']}, payment={ptype}): {e}")


@app.route("/api/customers/<int:cid>/send-whatsapp", methods=["POST"])
def api_send_whatsapp(cid):
    data = request.json
    text = data.get("text", "")
    content_cn = data.get("content_cn", "")
    content_en = data.get("content_en", "")

    # 查客户名用于切换聊天
    c = get_customer(cid)
    contact_name = c.get("name", "") if c else ""

    def job():
        try:
            send_text(text, contact_name=contact_name)
            add_message(cid, "sent", content_cn, content_en)
            # 手动发送的消息作为风格样本保存，AI会学习你的语气
            if content_cn and content_en:
                _save_style_sample(content_cn, content_en)
        except Exception as e:
            print(f"Send error: {e}")

    threading.Thread(target=job, daemon=True).start()
    return jsonify({"ok": True, "msg": "发送中"})


# ========= API: 翻译 =========
@app.route("/api/translate", methods=["POST"])
def api_translate():
    data = request.json
    text = data.get("text", "")
    lang = data.get("language", "English")
    country = data.get("country", "")
    result = translate(text, target_language=lang, country=country)
    return jsonify({"result": result})


# ========= API: 话术生成 =========
@app.route("/api/generate-script", methods=["POST"])
def api_generate_script():
    """根据话术模板+国家上下文+客户上下文，AI生成个性化回复"""
    data = request.json
    scenario = data.get("scenario", "")
    country = data.get("country", "")
    customer_context = data.get("context", "")
    if not scenario:
        return jsonify({"error": "请选择话术场景"}), 400
    # 如果有客户ID，从数据库加载最近消息
    customer_id = data.get("customer_id")
    if customer_id and not customer_context:
        from database import get_messages
        msgs = get_messages(customer_id, limit=5)
        customer_context = "\n".join(
            f"{'【我】' if m['direction']=='sent' else '【客户】'}{m.get('content_en','') or m.get('content_cn','')}"
            for m in msgs[-5:]
        )
    result = generate_customized_script(scenario, country, customer_context)
    return jsonify(result)


# ========= API: 知识库搜索 =========
@app.route("/api/knowledge-search", methods=["POST"])
def api_knowledge_search():
    """搜索行业知识库，AI回答问题"""
    data = request.json
    query = data.get("query", "").strip()
    if not query:
        return jsonify({"error": "请输入问题"}), 400
    result = search_industry_knowledge(query)
    return jsonify({"result": result})


# ========= API: 生图 =========
@app.route("/api/generate-image", methods=["POST"])
def api_generate_image():
    data = request.json
    prompt = data.get("prompt", "")
    customer_id = data.get("customer_id")
    image_data = data.get("image_data", "")  # 图生图参考图(base64)

    # 图生图：提示词本身就是英文描述（客户发的图+修改要求）
    # 文生图：把中文翻译成英文
    if image_data:
        en_prompt = prompt
    else:
        try:
            en_prompt = translate(prompt) or prompt
        except Exception as e:
            print(f"[ImageGen] 翻译失败，使用原文: {e}")
            en_prompt = prompt

    url, error = generate_image(en_prompt, image_data=image_data or None)
    if error:
        return jsonify({"error": error}), 500

    # 下载并保存
    try:
        import requests as req
        r = req.get(url, timeout=30)
        ext = ".png"
        filename = f"ai_{uuid.uuid4().hex[:8]}{ext}"
        filepath = os.path.join(UPLOAD_DIR, filename)
        with open(filepath, "wb") as f:
            f.write(r.content)
        add_media(filename, filepath, "image", len(r.content), prompt, customer_id)
        add_ai_generation("image", prompt, url, filename, customer_id=customer_id,
                          metadata={"mode": "text_to_image" if not image_data else "image_to_image"})
        return jsonify({"url": url, "saved": filename, "filepath": filepath})
    except Exception as e:
        return jsonify({"url": url, "saved": None, "error": str(e)})


# ========= API: 视频生成 =========
_video_tasks = {}  # task_id -> {"status": "pending|running|done|error", "url": ..., "error": ...}
_video_task_lock = threading.Lock()

@app.route("/api/video-presets")
def api_video_presets():
    return jsonify({k: {
        "name": v["name"],
        "duration": v["duration"],
        "shot_type": v["shot_type"],
        "size": v["size"]
    } for k, v in VIDEO_PRESETS.items()})

@app.route("/api/generate-video", methods=["POST"])
def api_generate_video():
    data = request.json
    prompt = data.get("prompt", "").strip()
    if not prompt:
        return jsonify({"error": "请输入视频描述"}), 400

    size = data.get("size", "1280*720")
    duration = data.get("duration", 15)
    shot_type = data.get("shot_type", "multi")
    task_id = uuid.uuid4().hex[:12]

    with _video_task_lock:
        _video_tasks[task_id] = {"status": "pending", "url": None, "error": None}

    def _run():
        with _video_task_lock:
            _video_tasks[task_id]["status"] = "running"
        print(f"[Video] 任务{task_id}启动: {prompt[:60]}...")
        url, error = generate_video(prompt, size, duration, shot_type)
        with _video_task_lock:
            if error:
                _video_tasks[task_id]["status"] = "error"
                _video_tasks[task_id]["error"] = error
                print(f"[Video] 任务{task_id}失败: {error}")
            else:
                # 下载到本地
                try:
                    import requests as req
                    r = req.get(url, timeout=120)
                    filename = f"video_{uuid.uuid4().hex[:8]}.mp4"
                    filepath = os.path.join(UPLOAD_DIR, filename)
                    with open(filepath, "wb") as f:
                        f.write(r.content)
                    add_media(filename, filepath, "video", len(r.content), prompt, data.get("customer_id"))
                    add_ai_generation("video", prompt, url, filename,
                                      customer_id=data.get("customer_id"),
                                      metadata={"size": size, "duration": duration, "shot_type": shot_type})
                    _video_tasks[task_id]["status"] = "done"
                    _video_tasks[task_id]["url"] = url
                    _video_tasks[task_id]["saved"] = filename
                    print(f"[Video] 任务{task_id}完成: {filename}")
                except Exception as e:
                    _video_tasks[task_id]["status"] = "error"
                    _video_tasks[task_id]["error"] = str(e)
                    _video_tasks[task_id]["url"] = url

    threading.Thread(target=_run, daemon=True).start()
    return jsonify({"ok": True, "task_id": task_id, "msg": "视频生成中（约1-5分钟）"})


# ========= API: 通义万相 图生视频 =========
@app.route("/api/generate-video-from-image", methods=["POST"])
def api_generate_video_from_image():
    """上传产品图片 → 通义万相图生视频"""
    image_file = request.files.get("image")
    image_path = request.form.get("image_path", "")
    prompt = request.form.get("prompt", "")
    duration = int(request.form.get("duration", 5))
    customer_id = request.form.get("customer_id")

    # 保存上传的图片
    saved_path = None
    if image_file:
        import uuid
        ext = image_file.filename.rsplit(".", 1)[-1].lower() if "." in image_file.filename else "jpg"
        filename = f"wan_{uuid.uuid4().hex[:8]}.{ext}"
        saved_path = os.path.join(UPLOAD_DIR, filename)
        image_file.save(saved_path)
    elif image_path:
        if os.path.exists(image_path):
            saved_path = image_path
        else:
            return jsonify({"error": f"图片路径不存在: {image_path}"}), 400
    else:
        return jsonify({"error": "请上传图片或提供图片路径"}), 400

    # 调用通义万相（阿里云），图生视频
    url, error = generate_video(prompt, quality="720p", duration=duration,
                                 aspect_ratio="16:9", image_path=saved_path)
    if error:
        return jsonify({"error": error}), 500

    # 下载视频到本地
    try:
        import requests as req
        r = req.get(url, timeout=120)
        vid = f"wan_{uuid.uuid4().hex[:8]}.mp4"
        vpath = os.path.join(UPLOAD_DIR, vid)
        with open(vpath, "wb") as f:
            f.write(r.content)
        add_media(vid, vpath, "video", len(r.content),
                  f"通义万相图生视频: {prompt[:60] if prompt else '产品展示'}",
                  customer_id)
        add_ai_generation("video", f"[通义万相] {prompt}", url, vid,
                          customer_id=customer_id,
                          metadata={"engine": "wan2.7", "source": os.path.basename(saved_path)})
        return jsonify({"ok": True, "url": url, "saved": vid, "filepath": vpath})
    except Exception as e:
        return jsonify({"ok": True, "url": url, "saved": None, "error": str(e)})


# ========= API: 通义万相 文生图 =========
@app.route("/api/generate-image-wan", methods=["POST"])
def api_generate_image_wan():
    """通义万相文生图（备用端点）"""
    data = request.json
    prompt = data.get("prompt", "")
    customer_id = data.get("customer_id")
    if not prompt:
        return jsonify({"error": "请输入描述"}), 400

    url, error = generate_image(prompt)
    if error:
        return jsonify({"error": error}), 500

    # 保存
    try:
        import requests as req
        r = req.get(url, timeout=60)
        fn = f"wan_image_{uuid.uuid4().hex[:8]}.png"
        fp = os.path.join(UPLOAD_DIR, fn)
        with open(fp, "wb") as f:
            f.write(r.content)
        add_media(fn, fp, "image", len(r.content), prompt, customer_id)
        add_ai_generation("image", f"[通义万相] {prompt}", url, fn,
                          customer_id=customer_id,
                          metadata={"engine": "wan2.7-image-pro"})
        return jsonify({"url": url, "saved": fn, "filepath": fp})
    except Exception as e:
        return jsonify({"url": url, "saved": None, "error": str(e)})

@app.route("/api/video-status/<task_id>")
def api_video_status(task_id):
    with _video_task_lock:
        task = _video_tasks.get(task_id)
    if not task:
        return jsonify({"error": "task not found"}), 404
    return jsonify(task)


# ========= API: 导入聊天记录 =========
@app.route("/api/customers/<int:cid>/import-chat", methods=["POST"])
def api_import_chat_history(cid):
    """上传WhatsApp导出的.txt聊天记录，AI分析后存入系统"""
    if "file" not in request.files:
        return jsonify({"error": "请上传文件"}), 400
    file = request.files["file"]
    if not file.filename.endswith(".txt"):
        return jsonify({"error": "仅支持.txt文件（WhatsApp导出的聊天记录）"}), 400
    try:
        raw_text = file.read().decode("utf-8", errors="replace")
    except:
        return jsonify({"error": "文件编码错误，请确认是UTF-8编码的txt文件"}), 400

    # 1. 解析消息
    parsed = parse_whatsapp_export(raw_text)
    if not parsed:
        return jsonify({"error": "无法解析聊天记录，请确认是WhatsApp导出的txt格式"}), 400

    # 2. AI分析
    analysis = analyze_chat_history(raw_text)

    return jsonify({
        "ok": True,
        "message_count": len(parsed),
        "parsed_messages": parsed[:20],  # 只返回前20条预览
        "total_messages": len(parsed),
        "analysis": analysis,
    })


@app.route("/api/customers/<int:cid>/import-chat/save", methods=["POST"])
def api_save_chat_history(cid):
    """确认保存解析后的聊天记录和AI分析到客户备注"""
    data = request.json
    analysis = data.get("analysis", {})
    parsed_messages = data.get("messages", [])

    # 保存到消息表
    from database import add_message
    saved = 0
    for msg in parsed_messages:
        content = msg.get("content", "")
        sender = msg.get("sender", "")
        # 判断方向：对方发的=received，我方发的=sent
        is_from_us = any(kw in sender.lower() for kw in ["philip", "bohui", "glowforge", "杨", "chen"])
        direction = "sent" if is_from_us else "received"
        add_message(cid, direction, content_cn=content, content_en="")
        saved += 1

    # 保存分析结果到客户备注
    from database import get_customer, update_customer
    c = get_customer(cid)
    existing_notes = (c.get("notes") or "") if c else ""
    summary_parts = []
    profile = analysis.get("customer_profile", {})
    if profile.get("interest_products"):
        summary_parts.append(f"【意向产品】{', '.join(profile['interest_products'])}")
    if analysis.get("key_needs"):
        summary_parts.append(f"【关键需求】{', '.join(analysis['key_needs'])}")
    if analysis.get("chat_summary"):
        summary_parts.append(f"【历史小结】{analysis['chat_summary']}")
    if analysis.get("decision_stage"):
        summary_parts.append(f"【阶段】{analysis['decision_stage']}")
    if analysis.get("price_sensitivity") and analysis['price_sensitivity'] != '未提及':
        summary_parts.append(f"【价格敏感度】{analysis['price_sensitivity']}")
    if analysis.get("next_steps"):
        summary_parts.append(f"【下一步】{analysis['next_steps']}")
    if analysis.get("tags"):
        summary_parts.append(f"【标签】{', '.join(analysis['tags'])}")
    if analysis.get("special_requirements"):
        summary_parts.append(f"【特殊要求】{', '.join(analysis['special_requirements'])}")
    if analysis.get("price_quoted", {}).get("has_quote"):
        summary_parts.append(f"【报价】{analysis['price_quoted'].get('quote_details', '')}")

    new_notes = (existing_notes + "\n\n" + "\n".join(summary_parts)).strip()
    update_customer(cid, notes=new_notes)

    # 更新客户语言
    lang = analysis.get("customer_profile", {}).get("language", "")
    if lang and lang in ("English", "中文"):
        update_customer(cid, language=lang)

    return jsonify({"ok": True, "saved": saved, "notes_updated": True})


# ========= API: 阅读消息 =========
@app.route("/api/read-whatsapp")
def api_read_whatsapp():
    result = read_messages()
    return jsonify({"result": result})


# ========= API: WhatsApp监控状态 =========
@app.route("/api/whatsapp-status")
def api_whatsapp_status():
    return jsonify(get_monitor_status())


@app.route("/api/whatsapp-refresh", methods=["POST"])
def api_whatsapp_refresh():
    """手动刷新 WhatsApp Web 页面（二维码过期时用）"""
    ok = refresh_whatsapp_page()
    return jsonify({"ok": ok, "msg": "刷新成功" if ok else "刷新失败"})


@app.route("/api/whatsapp-qr")
def api_whatsapp_qr():
    """获取 WhatsApp 登录二维码 + 登录状态"""
    qr = get_qr_base64()
    logged_in = is_logged_in()
    return jsonify({"qr": qr, "logged_in": logged_in})


@app.route("/whatsapp-login")
def whatsapp_login_page():
    """WhatsApp 二维码扫码页面（云端用）"""
    return """<!DOCTYPE html><html lang=zh><head>
<meta charset=utf-8><meta name=viewport content="width=device-width,initial-scale=1">
<title>WhatsApp 登录</title>
<style>
*{margin:0;padding:0;box-sizing:border-box}
body{background:#111;color:#fff;font-family:sans-serif;display:flex;flex-direction:column;align-items:center;justify-content:center;min-height:100vh;gap:20px;padding:20px}
h2{color:#00e676}
#qrBox{background:#fff;border-radius:16px;padding:16px;display:flex;align-items:center;justify-content:center;max-width:90vw;min-height:300px}
#qrBox img{max-width:100%;max-height:70vh;height:auto;display:block;object-fit:contain;border-radius:8px}
#qrBox .loading{color:#666;font-size:14px}
.status{font-size:14px;color:#aaa}
.btn{background:#00e676;color:#000;border:none;padding:10px 24px;border-radius:8px;font-size:14px;cursor:pointer}
.btn:hover{background:#00c853}
.hint{font-size:13px;color:#888;text-align:center;max-width:400px;line-height:1.6}
</style></head><body>
<h2>🔐 WhatsApp Web 登录</h2>
<div id=qrBox><div class=loading>正在加载二维码...</div></div>
<div class=status id=status>等待扫码...</div>
<button class=btn onclick=refreshQR()>刷新二维码</button>
<div class=hint>请用手机 WhatsApp 扫描上方二维码<br>二维码每30秒自动刷新</div>
<script>
let lastQR=""
async function refreshQR(){
  try{
    const r=await fetch('/api/whatsapp-qr')
    const d=await r.json()
    if(d.logged_in){
      document.getElementById('qrBox').innerHTML='<div class=loading>✅ 已登录</div>'
      document.getElementById('status').textContent='✅ 已登录'
      document.getElementById('status').style.color='#00e676'
      lastQR=""
      return
    }
    if(d.qr){
      document.getElementById('qrBox').innerHTML='<img src="data:image/png;base64,'+d.qr+'">'
      document.getElementById('status').textContent='📱 请用手机扫描二维码'
      lastQR=d.qr
    }else if(!lastQR){
      document.getElementById('qrBox').innerHTML='<div class=loading>等待 WhatsApp 生成二维码...</div>'
    }
  }catch(e){
    document.getElementById('status').textContent='❌ 连接失败: '+e.message
  }
}
refreshQR()
setInterval(refreshQR,5000)
</script></body></html>"""


# ========= WhatsApp 24×7自动监控 =========
# AI自动回复总开关（True=自动回复开启，False=只录入不回复）
_auto_reply_enabled = True

# 最近自动回复活动记录（用于前端通知）
_recent_wa_activity = []
_wa_activity_lock = threading.Lock()

def _add_wa_activity(typ, customer_id, customer_name, summary):
    """添加一条活动记录"""
    entry = {
        "type": typ,  # "new_customer", "received", "replied", "error"
        "customer_id": customer_id,
        "customer_name": customer_name,
        "summary": summary,
        "time": time.strftime("%H:%M:%S")
    }
    with _wa_activity_lock:
        _recent_wa_activity.insert(0, entry)
        _recent_wa_activity[:] = _recent_wa_activity[:50]  # 最多保留50条


def _do_auto_reply(cid, chat_name, reply_en, reply_cn):
    """在独立线程中执行自动回复，不阻塞事件循环
    - 等待1-5分钟（不定时，模拟真人）
    - 发前检查是否已被手动回复
    """
    try:
        # 1-5分钟不定时，模拟真人回复节奏
        wait = random.uniform(60, 300)
        print(f"[Auto] 等待{wait:.0f}秒后回复 {chat_name}...")
        time.sleep(wait)

        # ===== 防冲突检查：等待期间是否有人工回复 =====
        latest = get_messages(cid, limit=3)
        if latest:
            # get_messages返回按时间升序，最后一条是最新的
            last = latest[-1]
            if last["direction"] == "sent":
                print(f"[Auto] ⏭ {chat_name} 等待期间已被回复，跳过AI自动回复")
                _add_wa_activity("skipped", cid, chat_name, "AI等待期间已被手动回复，跳过")
                return

        # 真正发送
        send_text(reply_en, contact_name=chat_name)
        add_message(cid, "sent", reply_cn, reply_en)
        _add_wa_activity("replied", cid, chat_name, f"已自动回复: {reply_en[:50]}...")
        print(f"[Auto] ✅ 已回复 {chat_name}")
    except Exception as e:
        _add_wa_activity("error", cid, chat_name, f"自动回复失败: {e}")
        print(f"[Auto] 回复出错: {e}")


def _whatsapp_message_handler(chat_name, messages):
    """收到新消息时自动分析+记录+回复（不会阻塞事件循环）"""
    try:
        text = messages[-1]["text"] if messages else ""
        if not text:
            return

        print(f"[Auto] 收到 {chat_name}: {text[:60]}...")

        # 查找CRM中对应的客户，找不到就自动添加
        customers = get_customers()
        profile = _get_profile()
        c = next((c for c in customers if c["name"] == chat_name), None)
        is_new = False
        if not c:
            print(f"[Auto] 新客户「{chat_name}」，自动添加到CRM")
            # 先用AI greeting分析客户（基于知识库做背调+分级）
            greeting = get_ai_greeting(customer_name=chat_name, country="", sales_name=profile.get("name", "Philip"))
            wa_country = greeting.get("country_detected", "") if greeting and not greeting.get("error") else ""
            result = add_customer(name=chat_name, status="new", source="whatsapp", lead_status="new", country=wa_country, notes="WhatsApp自动添加")
            if result.get("duplicate"):
                c = get_customer(result["id"])
            elif result.get("id"):
                c = get_customer(result["id"])
            if not c:
                print(f"[Auto] 添加客户失败")
                return
            is_new = True

            # 更新AI评分和评级
            if greeting and not greeting.get("error"):
                grade = greeting.get("customer_grade", "")
                estimated_status = "qualified" if grade in ("A", "B") else "new"
                update_lead_status(c["id"], estimated_status)
                update_lead_source(c["id"], "whatsapp")
                print(f"[Auto] AI评级={grade}, lead_status={estimated_status}, 国家={wa_country}")

        # AI分析客户消息（注入国家上下文 + 聊天历史 + 知识库）
        customer_country = c.get("country", "") if c else ""
        chat_history = get_messages(c["id"], limit=10)
        history_for_ai = []
        for m in chat_history:
            history_for_ai.append({
                "role": m["direction"],  # "received" 或 "sent"
                "content_cn": m.get("content_cn", ""),
                "content_en": m.get("content_en", "")
            })
        style_samples = _get_style_samples(5)
        try:
            analysis = analyze_customer_message(text, country=customer_country, history=history_for_ai,
                                               style_samples=style_samples, sales_name=profile.get("name", "Philip"),
                                               customer_id=c.get("id", 0))
        except Exception as e:
            print(f"[Auto] AI分析异常: {e}")
            _add_wa_activity("error", c["id"], chat_name, f"AI分析异常: {e}")
            return
        if not analysis or "error" in analysis:
            print(f"[Auto] AI分析失败: {analysis}")
            _add_wa_activity("error", c["id"], chat_name, f"AI分析失败")
            return

        translation = analysis.get("translation", text)
        reply_en = analysis.get("suggested_reply_en", "")
        intent = analysis.get("intent", "其他")

        # ====== Lead State Engine: 更新客户状态 ======
        if is_new:
            init_customer_state(c["id"])
        state_result = update_lead_state(c["id"], intent, trigger_detail=text[:80])
        if state_result["transitioned"]:
            print(f"[LeadState] 客户#{c['id']} {chat_name}: {state_result['from_state']} → {state_result['to_state']} ({state_result['reason']})")
            _add_wa_activity("state_change", c["id"], chat_name,
                             f"{state_result['from_state']}→{state_result['to_state']}: {state_result['reason']}")

        # ====== Decision Engine: 决定下一步动作 ======
        decision = decide_action(c["id"], intent, extra_context={
            "urgency": analysis.get("urgency", "中"),
        })
        if decision["action"] != "ROUTINE_REPLY":
            print(f"[Decision] 客户#{c['id']} {chat_name}: 动作={decision['action']} ({decision['reason']})")
            # ====== Action Router: 执行动作 ======
            action_result = execute_action(decision["action"], c["id"], context={
                "chat_name": chat_name,
                "intent": intent,
                "reply_text": reply_en,
                "analysis": analysis,
            })
            if not action_result.get("ok"):
                print(f"[Action] 执行失败: {decision['action']} — {action_result.get('error', '')}")

        # 保存收到的消息到数据库
        add_message(c["id"], "received", translation, text)

        # 记录活动
        if is_new:
            _add_wa_activity("new_customer", c["id"], chat_name, f"新客户自动添加: {text[:40]}...")
        _add_wa_activity("received", c["id"], chat_name, f"收到消息: {translation[:40]}...")

        # 自动回复（检查总开关后在独立线程执行）
        if reply_en and _auto_reply_enabled:
            reply_cn = analysis.get("suggested_reply_cn", "")
            t = threading.Thread(
                target=_do_auto_reply,
                args=(c["id"], chat_name, reply_en, reply_cn),
                daemon=True
            )
            t.start()
        else:
            print(f"[Auto] 无需回复 {chat_name}")

    except Exception as e:
        print(f"[Auto] 处理出错: {e}")


# ========= API: 聊天活动通知 =========
@app.route("/api/whatsapp-activity")
def api_wa_activity():
    """获取最近的WhatsApp自动活动（新客户、收到消息、自动回复）"""
    since = request.args.get("since", "0")
    with _wa_activity_lock:
        if since == "0":
            return jsonify(_recent_wa_activity[:10])
        try:
            idx = int(since)
            return jsonify(_recent_wa_activity[:idx])
        except:
            return jsonify(_recent_wa_activity[:10])


# ========= API: AI风格学习状态 =========
@app.route("/api/style-samples")
def api_style_samples():
    """查看AI学到了多少你的回复风格"""
    samples = _get_style_samples(30)
    return jsonify({
        "total": len(samples),
        "samples": samples
    })


# ========= API: 销售个人资料 =========
@app.route("/api/profile", methods=["GET", "POST"])
def api_profile():
    """读取/修改销售个人资料（名字、称呼等）"""
    if request.method == "POST":
        data = request.json or {}
        profile = _save_profile(data)
        return jsonify(profile)
    return jsonify(_get_profile())


# ========= API: AI自动回复开关 =========
@app.route("/api/whatsapp-auto-reply", methods=["GET", "POST"])
def api_wa_auto_reply():
    """GET获取当前状态，POST切换自动回复开关"""
    global _auto_reply_enabled
    if request.method == "POST":
        data = request.json or {}
        if "enabled" in data:
            _auto_reply_enabled = bool(data["enabled"])
        else:
            _auto_reply_enabled = not _auto_reply_enabled
        status = "开启" if _auto_reply_enabled else "关闭"
        print(f"[Auto] AI自动回复已{status}")
        _add_wa_activity("toggle", 0, "系统", f"AI自动回复已{status}")
    return jsonify({"enabled": _auto_reply_enabled})


# ========= API: 客户简报 =========
@app.route("/api/customers/<int:cid>/brief")
def api_customer_brief(cid):
    """根据历史聊天记录生成客户简报"""
    msgs = get_messages(cid, limit=50)
    summary = summarize_chat(msgs)
    c = get_customer(cid)
    return jsonify({
        "summary": summary,
        "customer": c
    })


# ========= API: 导入WhatsApp聊天历史 =========
@app.route("/api/customers/<int:cid>/import-chat", methods=["POST"])
def api_import_chat(cid):
    """从WhatsApp读取当前聊天记录，存入数据库并分析"""
    try:
        raw = read_messages()
        if not raw or raw == "读取失败":
            return jsonify({"error": "WhatsApp读取失败，请确认已打开该客户的聊天窗口"}), 400

        # 尝试解析结构化消息
        parsed = parse_whatsapp_export(raw)
        imported = 0
        if parsed and len(parsed) > 1:
            # 获取客户信息以判断哪一方是"我们"
            customer = get_customer(cid)
            customer_name = customer["name"] if customer else "客户"
            for pm in parsed:
                sender = pm.get("sender", "")
                is_me = sender and sender.lower() not in customer_name.lower()
                direction = "sent" if is_me else "received"
                content_en = pm.get("text", "")
                # 去重检查：避免重复导入
                existing = get_messages(cid, limit=3)
                if any(m.get("content_en", "") == content_en for m in existing):
                    continue
                add_message(cid, direction, content_en=content_en)
                imported += 1

        if imported == 0:
            # 未能结构化解析，整段存为收到消息
            add_message(cid, "received", content_en=raw[:2000])
            imported = 1

        # 分析导入的内容
        msgs = get_messages(cid, limit=50)
        summary = summarize_chat(msgs)
        return jsonify({
            "imported": imported,
            "summary": summary,
            "total": len(msgs)
        })
    except Exception as e:
        return jsonify({"error": f"导入失败: {str(e)}"}), 500


# ========= API: 客户意图分析 =========
@app.route("/api/analyze-message", methods=["POST"])
def api_analyze_message():
    """分析客户消息意图，返回翻译+意图+建议回复"""
    data = request.json
    text = data.get("text", "").strip()
    if not text:
        return jsonify({"error": "请输入客户消息"}), 400
    try:
        result = analyze_customer_message(text)
    except Exception as e:
        print(f"[Analyze] AI分析异常: {e}")
        return jsonify({"error": f"AI分析异常: {e}"}), 500
    return jsonify(result or {"error": "分析失败"})


# ========= API: 多语言语音合成（Qwen-TTS） =========
@app.route("/api/tts", methods=["POST"])
def api_tts():
    """文字转语音，支持多语言

    POST JSON:
        text: str           — 要合成的文字
        lang: str           — English / Chinese / Auto / Arabic / etc（默认 Auto）
        voice: str          — Cherry / Chelsie / Stella（默认 Cherry）
        response_format: str — wav / mp3（默认 wav）
    Returns:
        {ok: true, url: "...", path: "...", duration: 1.5}
    """
    data = request.json
    text = (data or {}).get("text", "").strip()
    if not text:
        return jsonify({"error": "请输入文字"}), 400
    lang = data.get("lang", "Auto")
    voice = data.get("voice", "Cherry")
    fmt = data.get("response_format", "wav")
    try:
        from voice_engine import synthesize, get_audio_duration
        url, path = synthesize(text, lang=lang, voice=voice, response_format=fmt)
        if not path:
            return jsonify({"error": "语音合成失败"}), 500
        duration = get_audio_duration(path)
        # 返回相对路径
        rel_path = os.path.relpath(path, BASE_DIR) if os.path.isabs(path) else path
        return jsonify({"ok": True, "url": url, "path": rel_path.replace("\\", "/"), "duration": duration})
    except Exception as e:
        return jsonify({"error": f"语音合成异常: {e}"}), 500


@app.route("/api/tts/voices", methods=["GET"])
def api_tts_voices():
    """列出可用音色"""
    try:
        from voice_engine import list_voices
        return jsonify({"voices": list_voices()})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/tts/languages", methods=["GET"])
def api_tts_languages():
    """列出支持的语言"""
    try:
        from voice_engine import LANGUAGES
        return jsonify({"languages": LANGUAGES})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ========= API: 爆款分析 =========
@app.route("/api/analyze-viral", methods=["POST"])
def api_analyze_viral():
    """分析爆款内容"""
    data = request.json
    text = data.get("text", "").strip()
    if not text:
        return jsonify({"error": "请输入爆款文案"}), 400
    try:
        result = analyze_viral(text)
    except Exception as e:
        print(f"[Viral] AI分析异常: {e}")
        return jsonify({"error": f"AI分析异常: {e}"}), 500
    return jsonify(result or {"error": "分析失败"})

@app.route("/api/rewrite-copy", methods=["POST"])
def api_rewrite_copy():
    """仿写爆款文案"""
    data = request.json
    analysis_text = data.get("analysis_text", "").strip()
    industry = data.get("industry", "发光字/亚克力")
    if not analysis_text:
        return jsonify({"error": "缺少分析内容"}), 400
    try:
        result = rewrite_copy(analysis_text, industry)
    except Exception as e:
        print(f"[Rewrite] AI仿写异常: {e}")
        return jsonify({"error": f"AI仿写异常: {e}"}), 500
    return jsonify({"result": result or "生成失败"})

@app.route("/api/oembed")
def api_oembed():
    """通过oEmbed获取视频页面信息（YouTube/TikTok，无需API Key）"""
    url = request.args.get("url", "").strip()
    if not url:
        return jsonify({"error": "缺少URL"}), 400

    oembed_url = None
    if "youtube.com" in url or "youtu.be" in url:
        oembed_url = f"https://www.youtube.com/oembed?url={url}&format=json"
    elif "tiktok.com" in url:
        oembed_url = f"https://www.tiktok.com/oembed?url={url}"
    else:
        return jsonify({"error": "仅支持YouTube和TikTok链接"}), 400

    try:
        r = http_requests.get(oembed_url, timeout=10)
        if r.status_code == 200:
            data = r.json()
            return jsonify({
                "title": data.get("title", ""),
                "author_name": data.get("author_name", ""),
                "description": data.get("description", "") or data.get("title", ""),
                "thumbnail_url": data.get("thumbnail_url", ""),
            })
        return jsonify({"error": f"oEmbed请求失败: {r.status_code}"}), 502
    except Exception as e:
        return jsonify({"error": f"请求异常: {str(e)}"}), 502


@app.route("/api/read-and-analyze")
def api_read_and_analyze():
    """读取WhatsApp消息并分析客户意图"""
    raw = read_messages()
    if not raw:
        return jsonify({"error": "读取WhatsApp失败，请确认WhatsApp窗口已打开"})
    try:
        analysis = analyze_customer_message(raw)
    except Exception as e:
        print(f"[ReadAnalyze] AI分析异常: {e}")
        return jsonify({"error": f"AI分析异常: {e}", "raw": raw}), 500
    return jsonify({"raw": raw, "analysis": analysis or {}})


# ========= API: 文件管理 =========
@app.route("/api/media")
def api_media():
    ft = request.args.get("type")
    tag = request.args.get("tag")
    if tag:
        return jsonify(get_media_by_tag(tag_id=int(tag), filetype=ft))
    return jsonify(get_media(ft))

@app.route("/api/media/upload", methods=["POST"])
def api_upload_media():
    f = request.files.get("file")
    if not f:
        return jsonify({"error": "no file"}), 400
    filename = f"{uuid.uuid4().hex[:8]}_{f.filename}"
    filepath = os.path.join(UPLOAD_DIR, filename)
    f.save(filepath)
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    ft = "image" if ext in ("jpg","jpeg","png","gif","bmp","webp") else "video" if ext in ("mp4","mov","avi","wmv") else "document"
    mid = add_media(filename, filepath, ft, os.path.getsize(filepath), request.form.get("description", ""))
    return jsonify({"id": mid, "filename": filename, "filepath": filepath, "filetype": ft})

@app.route("/api/media/<int:mid>", methods=["DELETE"])
def api_delete_media(mid):
    delete_media(mid)
    return jsonify({"ok": True})

@app.route("/api/media/<int:mid>/tags")
def api_media_tags_for(mid):
    return jsonify(get_media_tags_for(mid))


@app.route("/api/media/<int:mid>/tags", methods=["PUT"])
def api_update_media_tags(mid):
    data = request.json or {}
    tag_ids = data.get("tag_ids", [])
    update_media_tags(mid, tag_ids)
    return jsonify({"ok": True})


@app.route("/api/media/tags")
def api_media_tags_list():
    return jsonify(get_media_tags())


@app.route("/api/media/tags", methods=["POST"])
def api_add_media_tag():
    data = request.json or {}
    name = data.get("name", "").strip()
    if not name:
        return jsonify({"error": "缺少标签名"}), 400
    tag_type = data.get("type", "general")
    color = data.get("color", "#00f2ff")
    result = add_media_tag(name, tag_type, color)
    if "error" in result:
        return jsonify(result), 400
    return jsonify(result)


@app.route("/api/media/tags/<int:tid>", methods=["DELETE"])
def api_delete_media_tag(tid):
    delete_media_tag(tid)
    return jsonify({"ok": True})


# ========= API: 产品视频/图片库（G盘视频库） =========
@app.route("/api/media-library")
def api_media_library():
    """列出产品视频/图片库，按分类组织"""
    from database import get_db as _get_db
    conn = _get_db()

    mtype = request.args.get("type", "video")  # video / image
    ft = "video_library" if mtype == "video" else "image_library"

    rows = conn.execute(
        "SELECT * FROM media_files WHERE filetype=? ORDER BY description ASC",
        (ft,)
    ).fetchall()
    conn.close()

    items = [dict(r) for r in rows]
    # 按分类分组
    categories = {}
    for item in items:
        cat = "未分类"
        desc = item.get("description", "")
        if "分类:" in desc:
            cat = desc.split("分类:")[1].split("|")[0].strip()
        if cat not in categories:
            categories[cat] = []
        # 提取产品名
        prod_name = ""
        if "产品:" in desc:
            prod_name = desc.split("产品:")[1].strip()
        categories[cat].append({
            "id": item["id"],
            "filename": item["filename"],
            "filepath": item["filepath"],
            "filesize": item["filesize"],
            "product_name": prod_name,
            "description": desc,
        })

    return jsonify({
        "total": len(items),
        "categories": [{"name": k, "items": v} for k, v in categories.items()]
    })


@app.route("/api/media-library/categories")
def api_media_library_categories():
    """获取视频/图片库分类列表"""
    from database import get_db as _get_db
    conn = _get_db()

    mtype = request.args.get("type", "video")
    ft = "video_library" if mtype == "video" else "image_library"

    rows = conn.execute(
        "SELECT description FROM media_files WHERE filetype=?", (ft,)
    ).fetchall()
    conn.close()

    cat_counts = {}
    for r in rows:
        desc = r["description"]
        cat = "未分类"
        if "分类:" in desc:
            cat = desc.split("分类:")[1].split("|")[0].strip()
        cat_counts[cat] = cat_counts.get(cat, 0) + 1

    return jsonify([
        {"name": k, "count": v} for k, v in sorted(cat_counts.items())
    ])


@app.route("/api/media-library/search")
def api_media_library_search():
    """搜索产品视频/图片"""
    from database import get_db as _get_db
    conn = _get_db()

    q = request.args.get("q", "").strip().lower()
    mtype = request.args.get("type", "video")
    ft = "video_library" if mtype == "video" else "image_library"

    if q:
        rows = conn.execute(
            "SELECT * FROM media_files WHERE filetype=? AND (filename LIKE ? OR description LIKE ?) ORDER BY description LIMIT 50",
            (ft, f"%{q}%", f"%{q}%")
        ).fetchall()
    else:
        rows = conn.execute(
            "SELECT * FROM media_files WHERE filetype=? ORDER BY description LIMIT 50",
            (ft,)
        ).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@app.route("/api/media-library/send-whatsapp", methods=["POST"])
def api_media_library_send():
    """从视频/图片库发送文件到WhatsApp客户"""
    from database import get_db as _get_db
    data = request.json
    media_id = data.get("media_id")
    customer_id = data.get("customer_id")

    if not media_id or not customer_id:
        return jsonify({"error": "缺少media_id或customer_id"}), 400

    conn = _get_db()
    media = conn.execute("SELECT * FROM media_files WHERE id=?", (media_id,)).fetchone()
    c = conn.execute("SELECT * FROM customers WHERE id=?", (customer_id,)).fetchone()
    conn.close()

    if not media:
        return jsonify({"error": "文件不存在"}), 404
    if not c:
        return jsonify({"error": "客户不存在"}), 404

    filepath = media["filepath"]
    contact_name = c["name"]

    if not os.path.exists(filepath):
        return jsonify({"error": f"文件不存在: {filepath}"}), 404

    def job():
        try:
            send_media_file(filepath, contact_name=contact_name)
            desc = media.get("description", "")
            prod = ""
            if "产品:" in desc:
                prod = desc.split("产品:")[1].strip()
            is_video = media["filetype"] == "video_library"
            media_type_str = "视频" if is_video else "图片"
            media_type_en = "video" if is_video else "image"
            msg_cn = f"[发送产品{media_type_str}] {prod}"
            msg_en = f"Sent product {media_type_en}: {prod}"
            add_message(customer_id, "sent", msg_cn, msg_en)
            _add_wa_activity("replied", customer_id, contact_name, f"已发送产品{prod}")
        except Exception as e:
            print(f"[MediaSend] 发送失败: {e}")

    threading.Thread(target=job, daemon=True).start()
    return jsonify({"ok": True, "msg": "发送中"})


# ========= AI生成记录 =========
@app.route("/api/ai-generations")
def api_ai_generations():
    gtype = request.args.get("type")
    return jsonify(get_ai_generations(gtype))

@app.route("/api/ai-generations/stats")
def api_ai_generation_stats():
    return jsonify(get_ai_generation_stats())

@app.route("/api/ai-generations/<int:gid>", methods=["DELETE"])
def api_delete_ai_generation(gid):
    delete_ai_generation(gid)
    return jsonify({"ok": True})

@app.route("/uploads/<path:filename>")
def serve_upload(filename):
    return send_from_directory(UPLOAD_DIR, filename)


# ========= 邮件 =========
def send_email_smtp(settings, to_email, subject, body, attachments=None):
    """底层SMTP发送，返回 (ok, error_msg)"""
    msg = MIMEMultipart()
    msg["From"] = f"{settings['from_name']} <{settings['from_email']}>"
    msg["To"] = to_email
    msg["Subject"] = subject
    msg.attach(MIMEText(body, "html" if "<html" in body else "plain", "utf-8"))

    for att in (attachments or []):
        fp = att.get("filepath") or att.get("path", "")
        if not os.path.exists(fp):
            continue
        with open(fp, "rb") as f:
            part = MIMEBase("application", "octet-stream")
            part.set_payload(f.read())
        encoders.encode_base64(part)
        part.add_header("Content-Disposition", f'attachment; filename="{att.get("filename", os.path.basename(fp))}"')
        msg.attach(part)

    try:
        port = int(settings.get("smtp_port", 587))
        use_tls = settings.get("use_tls", True)
        if use_tls:
            if port == 465:
                srv = smtplib.SMTP_SSL(settings["smtp_host"], port, timeout=30)
            else:
                srv = smtplib.SMTP(settings["smtp_host"], port, timeout=30)
                srv.starttls()
        else:
            srv = smtplib.SMTP(settings["smtp_host"], port, timeout=30)
        if settings.get("smtp_user"):
            srv.login(settings["smtp_user"], settings.get("smtp_pass", ""))
        srv.sendmail(settings["from_email"], [to_email], msg.as_string())
        srv.quit()
        return True, ""
    except Exception as e:
        return False, str(e)

@app.route("/api/email/settings", methods=["GET", "PUT"])
def api_email_settings():
    if request.method == "PUT":
        data = request.json
        save_email_settings(data)
        return jsonify({"ok": True})
    s = get_email_settings()
    # 密码脱敏
    pw = s.get("smtp_pass", "")
    if pw and len(pw) > 4:
        s["smtp_pass"] = pw[:4] + "****"
    elif pw:
        s["smtp_pass"] = "****"
    return jsonify(s)

@app.route("/api/email/test", methods=["POST"])
def api_email_test():
    s = get_email_settings()
    ok, err = send_email_smtp(s, s["from_email"], "GLOWFORGE CRM - 测试邮件", "如果收到此邮件，SMTP配置正确 ✅")
    if ok:
        return jsonify({"ok": True, "msg": "✅ 测试邮件发送成功！请检查收件箱"})
    return jsonify({"ok": False, "msg": f"❌ 连接失败: {err}"})

@app.route("/api/customers/<int:cid>/send-email", methods=["POST"])
def api_send_email(cid):
    data = request.json
    to_email = data.get("to_email", "").strip()
    subject = data.get("subject", "").strip()
    body = data.get("body", "").strip()
    attachments = data.get("attachments", [])
    if not to_email:
        return jsonify({"error": "请输入收件人邮箱"}), 400
    if not subject:
        return jsonify({"error": "请输入邮件主题"}), 400
    if not body:
        return jsonify({"error": "请输入邮件正文"}), 400

    settings = get_email_settings()

    def job():
        ok, err = send_email_smtp(settings, to_email, subject, body, attachments)
        if ok:
            add_email_log(cid, to_email, subject, body, attachments, "sent", "")
            add_message(cid, "sent", f"[邮件] {subject}", f"[Email] {subject}")
        else:
            add_email_log(cid, to_email, subject, body, attachments, "failed", err)
            add_message(cid, "sent", f"[邮件失败] {subject}: {err}", f"[Email Failed] {subject}: {err}")

    threading.Thread(target=job, daemon=True).start()
    return jsonify({"ok": True, "msg": "邮件发送中..."})

@app.route("/api/customers/<int:cid>/email-log")
def api_email_log(cid):
    return jsonify(get_email_log(cid))

@app.route("/api/email/log")
def api_email_log_all():
    return jsonify(get_all_email_log())


# ========= 产品目录 =========
@app.route("/api/products")
def api_products():
    cat = request.args.get("category")
    if cat == "__all__":
        cat = None
    return jsonify(get_products(cat))

@app.route("/api/products/categories")
def api_product_categories():
    return jsonify(get_product_categories())

@app.route("/api/products/<int:pid>")
def api_product(pid):
    p = get_product(pid)
    if not p:
        return jsonify({"error": "not found"}), 404
    return jsonify(p)

@app.route("/api/products", methods=["POST"])
def api_add_product():
    data = request.json
    pid = add_product(data)
    return jsonify({"ok": True, "id": pid})

@app.route("/api/products/<int:pid>", methods=["PUT"])
def api_update_product(pid):
    data = request.json
    update_product(pid, data)
    return jsonify({"ok": True})

@app.route("/api/products/<int:pid>", methods=["DELETE"])
def api_delete_product(pid):
    delete_product(pid)
    return jsonify({"ok": True})

# ========= 工程案例 =========
@app.route("/api/cases")
def api_cases():
    cat = request.args.get("category")
    if cat == "__all__":
        cat = None
    return jsonify(get_cases(cat))

@app.route("/api/cases/categories")
def api_case_categories():
    return jsonify(get_case_categories())

@app.route("/api/cases/<int:cid>")
def api_case(cid):
    c = get_case(cid)
    if not c:
        return jsonify({"error": "not found"}), 404
    return jsonify(c)

@app.route("/api/cases", methods=["POST"])
def api_add_case():
    data = request.json
    cid = add_case(data)
    return jsonify({"ok": True, "id": cid})

@app.route("/api/cases/<int:cid>", methods=["PUT"])
def api_update_case(cid):
    data = request.json
    update_case(cid, data)
    return jsonify({"ok": True})

@app.route("/api/cases/<int:cid>", methods=["DELETE"])
def api_delete_case(cid):
    delete_case(cid)
    return jsonify({"ok": True})

@app.route("/api/products/analyze-image", methods=["POST"])
def api_analyze_product_image():
    """上传产品图片 → AI分析 → 返回结构化产品信息 + 生图提示词"""
    f = request.files.get("file")
    if not f:
        return jsonify({"error": "no file"}), 400
    import base64
    image_data = base64.b64encode(f.read()).decode("utf-8")
    image_b64 = f"data:{f.content_type or 'image/jpeg'};base64,{image_data}"
    result = analyze_product_image(image_b64)
    if "error" in result:
        return jsonify(result), 500
    return jsonify(result)

@app.route("/api/products/generate-promo-video", methods=["POST"])
def api_generate_promo_video():
    """从产品库中的图片生成客户店面宣传视频（即梦图生视频）"""
    data = request.json
    product_id = data.get("product_id")
    biz_name = data.get("biz_name", "")
    sign_text = data.get("sign_text", "")

    if not product_id:
        return jsonify({"error": "缺少 product_id"}), 400

    p = get_product(product_id)
    if not p:
        return jsonify({"error": "产品不存在"}), 404

    images = p.get("images", [])
    if isinstance(images, str):
        images = json.loads(images)
    if isinstance(images, dict):
        images = [images]
    if not images:
        return jsonify({"error": "产品没有图片，请先上传"}), 400
    if isinstance(images[0], dict) and "url" in images[0]:
        pass
    elif isinstance(images[0], str):
        images = [{"url": images[0]}]
    if not images or not images[0].get("url"):
        return jsonify({"error": "产品没有图片，请先上传"}), 400

    image_url = images[0]["url"]
    video_url, err = generate_storefront_promo_video(image_url, biz_name, sign_text)

    if err:
        return jsonify({"error": err}), 500
    return jsonify({"video_url": video_url})

@app.route("/api/products/<int:pid>/quote", methods=["POST"])
def api_generate_quote(pid):
    """生成 PDF 报价单（博汇标准格式）"""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable)
    from datetime import datetime

    p = get_product(pid)
    if not p:
        return jsonify({"error": "产品不存在"}), 404

    data = request.json or {}
    customer_name = data.get("customer_name", "")
    notes = data.get("notes", "")

    # 报价编号 BH-YYYYMMDD-CUSTOMER-XXX
    date_str = time.strftime("%Y%m%d")
    seq = random.randint(100, 999)
    cust_code = (customer_name[:6] if customer_name else p.get("category", "PROD")).upper().replace(" ", "_")
    if not cust_code:
        cust_code = "PROD"
    quote_no = f"BH-{date_str}-{cust_code}-{seq}"

    # 目录
    quotes_dir = os.path.join(UPLOAD_DIR, "quotes")
    os.makedirs(quotes_dir, exist_ok=True)
    pdf_path = os.path.join(quotes_dir, f"{quote_no}.pdf")

    # 注册 Arial 字体（匹配博汇现有报价单）
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    arial_ok = False
    arial_paths = [r"C:\Windows\Fonts\Arial.ttf", r"C:\Windows\Fonts\arial.ttf",
                   "/usr/share/fonts/crm/arial.ttf"]
    arial_bold_paths = [r"C:\Windows\Fonts\Arialbd.ttf", r"C:\Windows\Fonts\arialbd.ttf",
                        "/usr/share/fonts/crm/arialbd.ttf"]
    for ap in arial_paths:
        if os.path.exists(ap):
            try:
                pdfmetrics.registerFont(TTFont("Arial", ap))
                arial_ok = True
                break
            except:
                pass
    for ap in arial_bold_paths:
        if os.path.exists(ap):
            try:
                pdfmetrics.registerFont(TTFont("Arial-Bold", ap))
                break
            except:
                pass

    FONT = "Arial" if arial_ok else "Helvetica"
    FONT_BOLD = "Arial-Bold" if arial_ok else "Helvetica-Bold"

    # 构建 PDF
    doc = SimpleDocTemplate(pdf_path, pagesize=A4,
                            topMargin=18*mm, bottomMargin=18*mm,
                            leftMargin=22*mm, rightMargin=22*mm)

    styles = getSampleStyleSheet()

    s_title = ParagraphStyle("CompanyTitle", fontSize=16, leading=18,
                              fontName=FONT_BOLD, alignment=0, spaceAfter=1*mm)
    s_subtitle = ParagraphStyle("CompanySub", fontSize=10, leading=13,
                                fontName=FONT, textColor=colors.HexColor("#333"), alignment=0, spaceAfter=0.5*mm)
    s_ref = ParagraphStyle("RefLine", fontSize=9, leading=12,
                           fontName=FONT, spaceAfter=1*mm, spaceBefore=3*mm)
    s_section = ParagraphStyle("SectionHead", fontSize=12, leading=15,
                                fontName=FONT_BOLD, spaceBefore=5*mm, spaceAfter=3*mm,
                                textColor=colors.HexColor("#222"))
    s_body = ParagraphStyle("Body", fontSize=9, leading=13,
                            fontName=FONT, spaceAfter=1.5*mm, alignment=0)
    s_bold = ParagraphStyle("BodyBold", parent=s_body, fontName=FONT_BOLD)
    s_small = ParagraphStyle("Small", fontSize=8, leading=10,
                              fontName=FONT, textColor=colors.HexColor("#666"))
    s_tcell = ParagraphStyle("TCell", fontSize=8.5, leading=11, fontName=FONT, alignment=1)
    s_tcell_l = ParagraphStyle("TCellLeft", fontSize=8.5, leading=11, fontName=FONT, alignment=0)
    s_tcell_bold = ParagraphStyle("TCellBold", parent=s_tcell, fontName=FONT_BOLD)
    s_tcell_bold_l = ParagraphStyle("TCellBoldLeft", parent=s_tcell_l, fontName=FONT_BOLD)

    elements = []

    # ===== HEADER =====
    elements.append(Paragraph("Zhongshan Bohui", s_title))
    elements.append(Paragraph("Advertising Craft Products", s_subtitle))
    elements.append(Paragraph("Co., Ltd.", s_subtitle))
    elements.append(Spacer(1, 1*mm))
    elements.append(Paragraph(
        "Factory: No.239 Fumin Avenue, Xiaolan Town, Zhongshan, Guangdong, China", s_small))
    elements.append(Paragraph(
        "Official Site: wa.bohui-sign.com | Tel: 13824779947", s_small))
    elements.append(Spacer(1, 2*mm))
    elements.append(HRFlowable(width="100%", thickness=0.75, color=colors.HexColor("#222")))
    elements.append(Spacer(1, 3*mm))

    # ===== REFERENCE =====
    month_names = ["January","February","March","April","May","June",
                   "July","August","September","October","November","December"]
    now = datetime.now()
    date_formatted = f"{month_names[now.month-1]} {now.day}, {now.year}"

    ref_data = [
        [Paragraph(f"<b>ORDER REF NO:</b>", s_ref), Paragraph(quote_no, s_ref)],
        [Paragraph(f"<b>DATE:</b>", s_ref), Paragraph(date_formatted, s_ref)],
    ]
    if customer_name:
        ref_data.append([Paragraph(f"<b>CLIENT:</b>", s_ref), Paragraph(customer_name, s_ref)])
    ref_table = Table(ref_data, colWidths=[38*mm, doc.width-38*mm])
    ref_table.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('TOPPADDING', (0,0), (-1,-1), 1),
        ('BOTTOMPADDING', (0,0), (-1,-1), 1),
    ]))
    elements.append(ref_table)
    elements.append(Spacer(1, 2*mm))

    # Status
    elements.append(Paragraph(
        "QUOTATION (All-Inclusive Price)", s_bold))
    elements.append(Spacer(1, 3*mm))

    # ===== SECTION I: SPECIFICATIONS =====
    elements.append(Paragraph("I. TECHNICAL SPECIFICATIONS & CRAFT", s_section))

    # Product name
    elements.append(Paragraph(f"<b>Product Name:</b> {p['name']}", s_body))

    # Specs
    specs = p.get("specs", {})
    if specs and isinstance(specs, dict):
        spec_lines = []
        if specs.get("material"):
            spec_lines.append(f"Material: {specs['material']}")
        if specs.get("thickness"):
            spec_lines.append(f"Thickness: {specs['thickness']}")
        if specs.get("size_range"):
            spec_lines.append(f"Dimensions: {specs['size_range']}")
        if specs.get("color_options"):
            spec_lines.append(f"Color Options: {specs['color_options']}")
        for line in spec_lines:
            elements.append(Paragraph(line, s_body))
        elements.append(Spacer(1, 1*mm))

    # Description
    if p.get("description"):
        elements.append(Paragraph(f"<b>Description:</b> {p['description']}", s_body))
        elements.append(Spacer(1, 2*mm))

    # ===== SECTION II: PRICING =====
    elements.append(Paragraph("II. PRICING", s_section))

    tiers = p.get("price_tiers", [])
    currency = p.get("currency", "USD")
    unit = p.get("unit", "个")
    min_order = p.get("min_order", 1)

    if tiers and isinstance(tiers, list) and len(tiers) > 0:
        # Header row
        hdr = [Paragraph("<b>Item</b>", s_tcell_bold_l),
               Paragraph("<b>Qty</b>", s_tcell_bold),
               Paragraph(f"<b>Net Price<br/>(excl. tax)</b>", s_tcell_bold),
               Paragraph(f"<b>10% Service<br/>Surcharge</b>", s_tcell_bold),
               Paragraph(f"<b>Final Unit<br/>Price (incl.)</b>", s_tcell_bold),
               Paragraph(f"<b>Total<br/>(incl. tax)</b>", s_tcell_bold)]

        rows = [hdr]
        for i, t in enumerate(tiers):
            qty = t.get("qty", "")
            price_str = t.get("price", "0")
            try:
                net_price = float(price_str)
            except:
                net_price = 0
            surcharge = round(net_price * 0.1, 2)
            unit_price = round(net_price + surcharge, 2)
            total = round(unit_price * (float(qty) if qty else 1), 2)

            item_name = f"{p['name']} — Tier {i+1}" if len(tiers) > 1 else p['name']
            rows.append([
                Paragraph(item_name, s_tcell_l),
                Paragraph(f"{qty} {unit}", s_tcell),
                Paragraph(f"{currency} {net_price:,.2f}", s_tcell),
                Paragraph(f"{currency} {surcharge:,.2f}", s_tcell),
                Paragraph(f"{currency} {unit_price:,.2f}", s_tcell),
                Paragraph(f"{currency} {total:,.2f}", s_tcell),
            ])

        # Summary rows
        net_total = sum(float(t.get("price", 0)) * (float(t.get("qty", 1)) if t.get("qty") else 1) for t in tiers)
        grand_total = round(net_total * 1.1, 2)

        rows.append([Paragraph("<b>Net Total (excl. tax)</b>", s_tcell_bold_l), "", "", "", "",
                     Paragraph(f"<b>{currency} {net_total:,.2f}</b>", s_tcell_bold)])
        rows.append([Paragraph(f"<b>Grand Total (Tax & Service Included)</b>", s_tcell_bold_l), "", "", "", "",
                     Paragraph(f"<b>{currency} {grand_total:,.2f}</b>", s_tcell_bold)])

        col_widths = [doc.width*0.26, doc.width*0.12, doc.width*0.16, doc.width*0.14, doc.width*0.16, doc.width*0.16]
        price_table = Table(rows, colWidths=col_widths, repeatRows=1)
        price_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#2c2c2c")),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('ALIGN', (0,0), (0,-1), 'LEFT'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('GRID', (0,0), (-1,-3), 0.4, colors.HexColor("#aaa")),
            ('LINEABOVE', (0,-2), (-1,-2), 0.8, colors.HexColor("#333")),
            ('LINEABOVE', (0,-1), (-1,-1), 0.8, colors.HexColor("#333")),
            ('TOPPADDING', (0,0), (-1,-1), 4),
            ('BOTTOMPADDING', (0,0), (-1,-1), 4),
            ('LEFTPADDING', (0,0), (-1,-1), 4),
            ('RIGHTPADDING', (0,0), (-1,-1), 4),
        ]))
        elements.append(price_table)
        elements.append(Spacer(1, 3*mm))

    # Min order
    if min_order > 1:
        elements.append(Paragraph(f"Minimum Order: {min_order} {unit}", s_body))
        elements.append(Spacer(1, 2*mm))

    # Notes
    if notes:
        elements.append(Paragraph(f"Notes: {notes}", s_body))
        elements.append(Spacer(1, 3*mm))

    # ===== SIGNATURE =====
    elements.append(Spacer(1, 5*mm))
    elements.append(HRFlowable(width="60%", thickness=0.5, color=colors.HexColor("#999")))
    elements.append(Spacer(1, 1*mm))
    elements.append(Paragraph("<b>Authorized Signature:</b>", s_body))
    elements.append(Paragraph("Yang Junliang", s_body))
    elements.append(Paragraph("COO | Zhongshan Bohui Advertising Craft Products Co., Ltd.", s_small))

    # ===== FOOTER =====
    elements.append(Spacer(1, 8*mm))
    elements.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor("#ccc")))
    elements.append(Spacer(1, 1*mm))
    footer = (f"Zhongshan Bohui Advertising Craft Products Co., Ltd. | "
              f"Factory: No.239 Fumin Avenue, Xiaolan Town, Zhongshan, Guangdong, China | "
              f"Official Site: wa.bohui-sign.com | Tel: 13824779947 | {quote_no}")
    elements.append(Paragraph(footer, ParagraphStyle("Footer", fontSize=7, leading=9,
                                                      fontName="Helvetica", textColor=colors.HexColor("#888"),
                                                      alignment=1)))

    doc.build(elements)

    # 注册到文件库
    fsize = os.path.getsize(pdf_path)
    fname = f"{quote_no}.pdf"
    add_media(f"Quote-{p['name']}-{quote_no}.pdf", pdf_path, "document", fsize)

    return jsonify({
        "ok": True,
        "quote_no": quote_no,
        "path": pdf_path,
        "url": f"/uploads/quotes/{fname}",
        "filename": f"Quote-{p['name']}-{quote_no}.pdf",
    })


# ========= 报价管理 =========
@app.route("/api/quotes")
def api_quotes():
    status = request.args.get("status")
    cid = request.args.get("customer_id", type=int)
    # 用户隔离：非管理员默认只看自己的报价
    uid = session.get("user_id")
    role = session.get("role")
    created_by = request.args.get("created_by", type=int)
    if created_by:
        pass  # 明确指定了创建人
    elif request.args.get("all") and role == "admin":
        created_by = None  # 管理员看全部
    elif uid:
        created_by = uid  # 默认只看自己的
    return jsonify(get_quotes(status, cid, created_by))

@app.route("/api/quotes/<int:qid>")
def api_quote(qid):
    q = get_quote(qid)
    return jsonify(q) if q else (jsonify({"error": "not found"}), 404)

@app.route("/api/quotes", methods=["POST"])
def api_add_quote():
    data = request.json
    data["created_by"] = session.get("user_id")
    result = add_quote(data)
    uid = session.get("user_id")
    if uid and result.get("id"):
        add_activity_log(uid, "create", "quote", result["id"],
            f"创建了报价 {result.get('quote_no','')} (ID:{result['id']})")
    return jsonify(result)

@app.route("/api/quotes/<int:qid>", methods=["PUT"])
def api_update_quote(qid):
    data = request.json
    q = get_quote(qid)
    update_quote(qid, data)
    uid = session.get("user_id")
    if uid and q:
        add_activity_log(uid, "update", "quote", qid,
            f"修改了报价 {q.get('quote_no','')} (ID:{qid})")
    return jsonify({"ok": True})

@app.route("/api/quotes/<int:qid>", methods=["DELETE"])
def api_delete_quote(qid):
    q = get_quote(qid)
    delete_quote(qid)
    uid = session.get("user_id")
    if uid and q:
        add_activity_log(uid, "delete", "quote", qid,
            f"删除了报价 {q.get('quote_no','')} (ID:{qid})")
    return jsonify({"ok": True})

@app.route("/api/quotes/<int:qid>/send", methods=["POST"])
def api_send_quote(qid):
    q = get_quote(qid)
    if not q:
        return jsonify({"error": "报价单不存在"}), 404
    data = request.json or {}
    contact_name = data.get("contact_name", "")
    text = data.get("text", f"Quote {q['quote_no']} — Please check the attached quotation.")
    uid = session.get("user_id")
    if uid and q:
        add_activity_log(uid, "send", "quote", qid,
            f"发送了报价 {q.get('quote_no','')} 给客户 {q.get('customer_id','')}")
    def job():
        try:
            from whatsapp_engine import send_text, send_media_file
            if q.get("pdf_path") and os.path.exists(q["pdf_path"]):
                send_media_file(q["pdf_path"], text, contact_name=contact_name)
            else:
                send_text(text, contact_name=contact_name)
            update_quote(qid, {"status": "sent"})
        except Exception as e:
            print(f"[Quote] send error: {e}")
    threading.Thread(target=job, daemon=True).start()
    return jsonify({"ok": True, "msg": "发送中"})


# ==================== Lead State API ====================
@app.route("/api/leads/<int:cid>/state")
def api_lead_state(cid):
    """查看客户当前状态 + 历史"""
    from lead_state_engine import get_lead_state, get_state_history
    from action_router import get_action_history
    c = get_customer(cid)
    if not c:
        return jsonify({"error": "客户不存在"}), 404
    return jsonify({
        "customer": c["name"],
        "current_state": get_lead_state(cid),
        "state_history": get_state_history(cid),
        "action_history": get_action_history(cid),
    })


@app.route("/api/leads/<int:cid>/state", methods=["POST"])
def api_lead_set_state(cid):
    """手动设置客户状态"""
    from lead_state_engine import get_lead_state, set_lead_state, log_state_transition
    data = request.json or {}
    new_state = data.get("state", "")
    valid_states = ["NEW", "INTERESTED", "REQUESTED_PRICE", "QUOTED", "NEGOTIATING", "HOT", "COLD", "CLOSED_WON", "CLOSED_LOST"]
    if new_state not in valid_states:
        return jsonify({"error": f"无效状态，可选: {', '.join(valid_states)}"}), 400
    old = get_lead_state(cid)
    set_lead_state(cid, new_state)
    log_state_transition(cid, old, new_state, trigger_source="manual", trigger_detail=f"手动设置: {data.get('reason', '')}")
    return jsonify({"ok": True, "from": old, "to": new_state})


@app.route("/api/leads/states")
def api_lead_states_summary():
    """所有客户状态分布统计"""
    from lead_state_engine import LEAD_STATES
    import sqlite3
    conn = sqlite3.connect(os.path.join(BASE_DIR, "crm_data.db"))
    try:
        rows = conn.execute("SELECT lead_state, COUNT(*) as cnt FROM customers GROUP BY lead_state ORDER BY cnt DESC").fetchall()
    except:
        rows = []
    conn.close()
    dist = {r[0]: r[1] for r in rows}
    return jsonify({
        "states": {k: {"label": v, "count": dist.get(k, 0)} for k, v in LEAD_STATES.items()},
        "distribution": dist,
    })


# ========= 产品目录 =========
@app.route("/api/generate-catalog", methods=["POST"])
def api_generate_catalog():
    """生成高级感PDF产品目录"""
    data = request.json or {}
    cat = data.get("category", "__all__")
    lang = data.get("language", "bilingual")
    title = data.get("title", "GLOWFORGE Product Catalog")

    def job():
        try:
            generate_catalog(category_filter=cat, language=lang, title=title)
        except Exception as e:
            print(f"[Catalog] 生成失败: {e}")

    # Run in thread for large catalogs
    import threading
    threading.Thread(target=job, daemon=True).start()

    return jsonify({"ok": True, "msg": "目录生成中，请稍后查看 uploads/catalogs/ 目录"})


@app.route("/api/catalogs")
def api_catalogs():
    """列出已生成的产品目录"""
    cat_dir = os.path.join(UPLOAD_DIR, "catalogs")
    if not os.path.isdir(cat_dir):
        return jsonify([])
    files = []
    for f in sorted(os.listdir(cat_dir), reverse=True):
        if f.endswith(".pdf"):
            fpath = os.path.join(cat_dir, f)
            files.append({
                "name": f,
                "size": os.path.getsize(fpath),
                "mtime": datetime.fromtimestamp(os.path.getmtime(fpath)).isoformat(),
                "url": f"/uploads/catalogs/{f}",
            })
    return jsonify(files)


# ==================== 报价计算器 PDF 生成 ====================
CALC_PRODUCT_SPECS = {
    "front": {"name":"Front-Lit LED Letters","craft":"Front-illuminated LED letters. Face illumination via 5mm high-transparency milky white acrylic sheet. High-brightness SMD LED beads, uniform light distribution without dark spots.","install":"Bottom stainless steel base with M4×40mm threaded rod mounting + industrial adhesive. Includes 1:1 installation template."},
    "back": {"name":"Halo Backlit LED Letters","craft":"Halo backlit LED letters. Light emitted from the back creating a soft halo/glow effect on the mounting surface. 5mm high-transparency acrylic for even light diffusion.","install":"Stand-off mounting from wall surface. Includes M4 threaded rods + 1:1 installation template."},
    "double": {"name":"Double-Sided Illuminated Letters","craft":"Double-sided illuminated letters. Front & back both fitted with 5mm high-transparency acrylic for even light transmission. Overall thickness with lighting chamber.","install":"Suspended or wall-mounted. Includes M4 hardware + 1:1 installation template."},
    "bottom": {"name":"Bottom-Lit LED Letters","craft":"Bottom-lit LED letters. Light emitted from the base channel creating a floating illusion effect on the mounting surface.","install":"Surface mounted with concealed base channel. Includes mounting hardware + template."},
    "rgb": {"name":"RGB Full-Color Dynamic Letters","craft":"RGB full-color dynamic LED letters. Built-in waterproof LED light strips with T8 controller for color management — static colors, fading, flashing, and gradient effects.","install":"Bottom stainless steel base. Includes RGB controller + waterproof connectors + 1:1 installation template."},
    "neon": {"name":"LED Neon Flex Letters","craft":"LED neon flex letters. Flexible silicone LED neon tube, energy-efficient, uniform brightness, 360° seamless bend. Durable for outdoor use.","install":"Aluminum backing frame with mounting clips. Pre-wired with LED driver."},
    "mini": {"name":"Mini LED Illuminated Letters","craft":"Mini LED illuminated letters. Compact SMD LED, high brightness density. Suitable for indoor signage and detail-oriented applications.","install":"Surface mounting with mini stand-offs. Includes 1:1 template."},
    "metal": {"name":"Solid Metal Letters (Non-Illuminated)","craft":"Solid metal letters, non-illuminated. Precision laser-cut, seamless welding, grinding, polishing to mirror/brushed finish.","install":"Direct surface mounting with industrial adhesive and/or mechanical fixings."},
    "acrylic_display": {"name":"Acrylic Display Signage","craft":"Acrylic display signage. Premium acrylic sheet with reverse UV printing available for vibrant graphics. High transparency and gloss.","install":"Stand-off wall mounting or free-standing base option."},
    "flat": {"name":"Flat Metal Letters (Non-Illuminated)","craft":"Flat metal letters, non-illuminated. Laser-cut from metal plate. Clean edges, deburred finish.","install":"Direct surface mounting with adhesive pads or screws."},
}

@app.route("/api/calc/pdf", methods=["POST"])
def api_calc_pdf():
    """Generate professional ReportLab PDF quotation from calculator data"""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.colors import HexColor, white
    from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable, PageBreak
    )
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    # Register fonts
    if sys.platform == "win32":
        FONT_DIRS = ["C:/Windows/Fonts"]
    else:
        FONT_DIRS = ["/usr/share/fonts/crm", "/usr/share/fonts"]
    for f in [("SimSun","simsun.ttc"),("SimHei","simhei.ttf"),
               ("MSYH","msyh.ttc"),("MSYHBD","msyhbd.ttc"),
               ("Arial","arial.ttf"),("ArialBD","arialbd.ttf")]:
        for fd in FONT_DIRS:
            fp = os.path.join(fd, f[1])
            if os.path.exists(fp):
                try:
                    pdfmetrics.registerFont(TTFont(f[0], fp))
                    break
                except:
                    pass

    # Colours matching Bohui quote style
    C_PRI    = HexColor("#1a237e")
    C_ACC    = HexColor("#283593")
    C_THDR   = HexColor("#1a237e")
    C_TALT   = HexColor("#f5f5f5")
    C_DARK   = HexColor("#212121")
    C_MID    = HexColor("#616161")
    C_BORD   = HexColor("#c5cae9")
    C_RED    = HexColor("#c62828")
    C_TBG    = HexColor("#eef2ff")
    C_GBG    = HexColor("#fff3e0")

    PAGE_W, PAGE_H = A4
    ML, MR, MT, MB = 50, 50, 50, 45
    CW = PAGE_W - ML - MR

    def ps(name, font, size, color=C_DARK, align=TA_LEFT, leading=None, before=0, after=0):
        return ParagraphStyle(name, fontName=font, fontSize=size, textColor=color,
                              alignment=align, leading=leading or size*1.4,
                              spaceBefore=before, spaceAfter=after)

    S = {
        "co_name":ps("co_name","ArialBD",22,C_PRI,leading=28),
        "co_addr":ps("co_addr","Arial",9,C_DARK,leading=13),
        "ref":ps("ref","ArialBD",12,C_PRI,TA_RIGHT,leading=15),
        "date":ps("date","Arial",9,C_MID,TA_RIGHT,leading=12),
        "status":ps("status","ArialBD",13,C_RED,TA_CENTER,leading=17,before=2,after=6),
        "sec_hdr":ps("sec_hdr","ArialBD",11,C_PRI,leading=16,before=10,after=4),
        "spec_lbl":ps("spec_lbl","ArialBD",8.5,C_ACC,before=4,after=1,leading=13),
        "spec_it":ps("spec_it","Arial",8.5,C_DARK,after=1,leading=13),
        "spec_det":ps("spec_det","Arial",8,C_DARK,after=1,leading=12),
        "tbl_hdr":ps("tbl_hdr","ArialBD",8,white,TA_CENTER,leading=10),
        "tbl_cell":ps("tbl_cell","Arial",8,C_DARK,TA_CENTER,leading=10),
        "tbl_l":ps("tbl_l","Arial",8,C_DARK,TA_LEFT,leading=10),
        "net_lbl":ps("net_lbl","ArialBD",8.5,C_DARK,TA_RIGHT,leading=11),
        "net_val":ps("net_val","ArialBD",9,C_PRI,TA_CENTER,leading=11),
        "grand_lbl":ps("grand_lbl","ArialBD",11,C_DARK,TA_RIGHT,leading=14),
        "grand_val":ps("grand_val","ArialBD",12,C_RED,TA_CENTER,leading=15),
        "sig_lbl":ps("sig_lbl","Arial",8,C_MID,TA_RIGHT,leading=10),
        "sig_txt":ps("sig_txt","ArialBD",10,C_DARK,leading=13),
        "footer":ps("footer","Arial",7,C_MID,TA_CENTER,leading=9),
        "note":ps("note","Arial",8,C_DARK,leading=12,before=2,after=1),
        "note_bold":ps("note_bold","ArialBD",8.5,C_ACC,leading=13,before=4,after=1),
    }

    data = request.json or {}
    ref_no = data.get("ref", "BH-000000-000")
    date_str = data.get("date", "")
    client_name = data.get("client", "—")
    currency = data.get("currency", "CNY")
    markup_pct = int(data.get("markupPct", 30))
    exchange_rate = float(data.get("exchangeRate", 6.8))

    ptype = data.get("productType", "front")
    mat_type = data.get("material", "ss304")
    mat_label = data.get("materialLabel", "304 Stainless Steel")
    electro_color = data.get("electroColor", "")
    is_electro = data.get("isElectro", False)
    kelvin = data.get("kelvin", "6000K-6500K / Cool White")
    items = data.get("items", [{"w":50,"h":50,"thick":8,"qty":1}])
    total_qty = sum(it.get("qty",1) for it in items)
    total_len_cm = sum(max(it.get("w",0),it.get("h",0)) * it.get("qty",1) for it in items)

    box_size = data.get("boxSize", "0×0×0cm")
    sea_freight = data.get("seaFreight", "3969")
    air_freight = data.get("airFreight", "7677")

    cost_items = data.get("costBreakdown", {})
    formal_total = data.get("formalTotal", "0")

    # Build spec descriptions from product type
    spec = CALC_PRODUCT_SPECS.get(ptype, CALC_PRODUCT_SPECS["front"])
    prod_name = spec["name"]
    craft_desc = spec["craft"]
    install_desc = spec["install"]

    # Material description
    if is_electro:
        mat_desc = f"{mat_label} stainless steel, {electro_color} electroplating finish, corrosion-resistant, premium appearance."
    elif "ss" in mat_type:
        mat_desc = f"{mat_label}, precision laser-cut, seamless welding, grinding and polishing fine finish."
    elif mat_type == "zn":
        mat_desc = "Galvanized steel plate, durable and cost-effective."
    elif mat_type == "zn_paint":
        mat_desc = "Galvanized steel plate with painted finish, weather-resistant."
    else:
        mat_desc = mat_label

    # Build specs list for PDF
    specs_list = []
    specs_list.append(("Product Type:", prod_name))
    specs_list.append(("Material:", mat_desc))
    if is_electro:
        specs_list.append(("Plating Finish:", f"{electro_color} electroplating — uniform color, anti-corrosion."))
    specs_list.append(("Craft & Construction:", craft_desc))
    specs_list.append(("LED & Lighting:", f"Color temperature: {kelvin}. Even illumination."))

    # Power description
    power_desc = data.get("powerDesc", "Mean Well 400W 12V power supply")
    specs_list.append(("Power Supply:", power_desc))
    specs_list.append(("Dimensions:", ', '.join(f"{it['w']}×{it['h']}×{it['thick']}cm ×{it['qty']}" for it in items)))
    specs_list.append(("Installation:", install_desc))
    specs_list.append(("Packaging:", "15mm thick plywood crate, pearl cotton foam interior. Shockproof, moisture-proof for international shipping. External size: " + box_size))

    # Build table rows
    price_items = data.get("priceItems", [])
    net_total_val = data.get("netTotal", "0")
    grand_total_val = data.get("grandTotal", "0")

    def cell(t):
        return Paragraph(str(t), S["tbl_cell"])

    def spec_bold(t):
        return Paragraph(f"<b>{t}</b>", S["spec_lbl"])

    def spec_item(l, d):
        return Paragraph(f"<b>{l}</b> {d}", S["spec_it"])

    def spec_detail(d):
        return Paragraph(f"\xa0\xa0\xa0{d}", S["spec_det"])

    # Generate PDF
    pdf_dir = os.path.join(UPLOAD_DIR, "quotes")
    os.makedirs(pdf_dir, exist_ok=True)
    safe_ref = ref_no.replace("/","-").replace("\\","-")
    output_path = os.path.join(pdf_dir, f"QUOTE_{safe_ref}.pdf")

    doc = SimpleDocTemplate(output_path, pagesize=A4,
                            leftMargin=ML, rightMargin=MR,
                            topMargin=MT, bottomMargin=MB,
                            title=f"Bohui Quote {ref_no}",
                            author="Bohui Advertising Craft")
    story = []

    # ── HEADER ──
    left = [
        Paragraph("Zhongshan Bohui Advertising Craft Products Co., Ltd.", S["co_name"]),
        Paragraph("Factory: No.239 Fumin Avenue, Xiaolan Town, Zhongshan, Guangdong, China 528415", S["co_addr"]),
        Paragraph("Official Site: wa.bohui-sign.com | Tel: +86 13824779947", S["co_addr"]),
    ]
    right = [
        Paragraph(f"ORDER REF NO: {ref_no}", S["ref"]),
        Paragraph(f"DATE: {date_str}", S["date"]),
    ]
    hdr = Table([[left, right]], colWidths=[CW*0.6, CW*0.4])
    hdr.setStyle(TableStyle([
        ('VALIGN',(0,0),(-1,-1),'TOP'),
        ('LEFTPADDING',(0,0),(0,0),0),
        ('RIGHTPADDING',(1,1),(1,1),0),
        ('TOPPADDING',(0,0),(-1,-1),0),
        ('BOTTOMPADDING',(0,0),(-1,-1),0),
    ]))
    story.append(hdr)
    story.append(Spacer(1,4))
    story.append(HRFlowable(width="100%", thickness=1, color=C_PRI, spaceAfter=6, spaceBefore=2))
    story.append(Paragraph("QUOTATION (All-Inclusive Price)", S["status"]))
    story.append(HRFlowable(width="100%", thickness=0.5, color=C_BORD, spaceAfter=8, spaceBefore=2))

    # CLIENT
    story.append(Paragraph(f"<b>CLIENT:</b> {client_name}", ps("client","Arial",10,C_DARK,leading=14,before=4,after=8)))

    # ── SECTION I: SPECS ──
    story.append(Paragraph("I. TECHNICAL SPECIFICATIONS & CRAFT", S["sec_hdr"]))
    story.append(Spacer(1,2))
    for label, desc in specs_list:
        story.append(Paragraph(f"<b>{label}</b> {desc}", S["spec_it"]))
    story.append(Spacer(1,10))

    # ── SECTION II: PRICING ──
    story.append(Paragraph("II. PRICING", S["sec_hdr"]))
    story.append(Spacer(1,6))

    col_w = [CW*0.20, CW*0.09, CW*0.15, CW*0.13, CW*0.16, CW*0.27]

    def h(t):
        return Paragraph(t, S["tbl_hdr"])

    sym = "¥" if currency == "CNY" else "$"
    rate_note = f" (1 USD = {exchange_rate} CNY)" if currency == "CNY" else ""

    tbl_rows = [
        [h("Item"), h("Qty"),
         h(f"Net Price<br/><font size=6>(excl. tax){rate_note}</font>"),
         h("Service Fee<br/><font size=6>(incl.)</font>"),
         h("Final Unit<br/><font size=6>Price (incl.)</font>"),
         h("Total Subtotal<br/><font size=6>(incl. tax)</font>")],
    ]

    for pi in price_items:
        item_desc = pi.get("item") or pi.get("desc", "Item")
        net_val   = pi.get("netUnit") or pi.get("net", "0")
        unit_val  = pi.get("unitPrice") or pi.get("unit", "0")
        tbl_rows.append([
            Paragraph(str(item_desc), S["tbl_l"]),
            Paragraph(str(pi.get("qty","")), S["tbl_l"]),
            cell(f"{sym}{net_val}"),
            cell(f"{sym}{pi.get('service','0')}"),
            cell(f"{sym}{unit_val}"),
            cell(f"{sym}{pi.get('total','0')}"),
        ])

    n_data = len(tbl_rows)
    tbl_rows.append([
        Paragraph("Net Total (excl. tax)", S["net_lbl"]),
        cell(""), cell(""), cell(""), cell(""),
        Paragraph(f"{sym}{net_total_val} (RMB)", S["net_val"])])
    tbl_rows.append([
        Paragraph("Grand Total (Tax & Service Included)", S["grand_lbl"]),
        cell(""), cell(""), cell(""), cell(""),
        Paragraph(f"{sym}{grand_total_val} (RMB)", S["grand_val"])])

    tbl = Table(tbl_rows, colWidths=col_w, repeatRows=1)
    tbl_style = [
        ('BACKGROUND',(0,0),(-1,0),C_THDR),
        ('TEXTCOLOR',(0,0),(-1,0),white),
        ('ALIGN',(0,0),(-1,-1),'CENTER'),
        ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
        ('GRID',(0,0),(-1,0),0.5,C_BORD),
        ('GRID',(0,1),(-1,n_data-1),0.5,C_BORD),
        ('TOPPADDING',(0,0),(-1,-1),5),
        ('BOTTOMPADDING',(0,0),(-1,-1),5),
        ('LEFTPADDING',(0,0),(-1,-1),5),
        ('RIGHTPADDING',(0,0),(-1,-1),5),
        ('ALIGN',(0,1),(0,-1),'LEFT'),
    ]
    for idx in range(1, n_data):
        if idx % 2 == 1:
            tbl_style.append(('BACKGROUND',(0,idx),(-1,idx),C_TALT))
    tbl_style += [
        ('SPAN',(0,-2),(4,-2)),
        ('BACKGROUND',(0,-2),(-1,-2),C_TBG),
        ('LINEABOVE',(0,-2),(-1,-2),1,C_PRI),
        ('LINEBELOW',(0,-2),(-1,-2),0.5,C_BORD),
        ('SPAN',(0,-1),(4,-1)),
        ('BACKGROUND',(0,-1),(-1,-1),C_GBG),
        ('LINEABOVE',(0,-1),(-1,-1),1.5,C_RED),
        ('LINEBELOW',(0,-1),(-1,-1),1.5,C_RED),
        ('TOPPADDING',(0,-1),(-1,-1),8),
        ('BOTTOMPADDING',(0,-1),(-1,-1),8),
    ]
    tbl.setStyle(tbl_style)
    story.append(tbl)
    story.append(Spacer(1,8))

    # ── Formal Total ──
    formal_total_val = data.get("formalTotal", grand_total_val)
    formal_sym = "¥" if currency == "CNY" else "$"
    story.append(HRFlowable(width="100%", thickness=0.3, color=C_BORD, spaceAfter=4, spaceBefore=2))
    story.append(Paragraph(
        f"<b>TOTAL QUOTATION PRICE:</b>"
        f" <font color='#c62828' size=14>{formal_sym}{formal_total_val}</font>",
        ps("formal","ArialBD",10,C_DARK,TA_CENTER,leading=16,before=4,after=4)))

    # ── Notes ──
    notes = [
        ("Coverage:", "All-inclusive: materials, processing, accessories, crate, shipping, customs, delivery."),
        ("Payment:", "50% deposit to start production. 50% balance before shipment."),
        ("Lead Time:", "10 working days (counted from day after deposit received)."),
        ("Validity:", "Quote valid 30 days from date of issue."),
        ("", None),
        ("SHIPPING OPTIONS (DDU — Delivered Duty Unpaid):", None),
        ("", f"Sea freight: {sym}{sea_freight} — Transit: 28-35 days after sailing."),
        ("", f"Air freight: {sym}{air_freight} — Transit: 3-5 working days after departure."),
    ]

    story.append(HRFlowable(width="100%", thickness=0.3, color=C_BORD, spaceAfter=4, spaceBefore=2))
    for label, desc in notes:
        if label and desc:
            story.append(Paragraph(f"<b>{label}</b> {desc}", S["note"]))
        elif label and not desc:
            story.append(Paragraph(f"<b>{label}</b>", S["note_bold"]))
        elif desc:
            story.append(Paragraph(f"\xa0\xa0\xa0{desc}", S["note"]))
    story.append(Spacer(1,12))

    # ── Signature ──
    sig = Table([
        [Paragraph("Authorized Signature:", S["sig_lbl"]),
         Paragraph("Yang Junliang<br/><font size=7>COO | Zhongshan Bohui Advertising Craft Products Co., Ltd.</font>",
                   S["sig_txt"])]
    ], colWidths=[CW*0.30, CW*0.70])
    sig.setStyle(TableStyle([
        ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
        ('LEFTPADDING',(0,0),(-1,-1),0),
        ('RIGHTPADDING',(0,0),(-1,-1),0),
        ('TOPPADDING',(0,0),(-1,-1),0),
        ('BOTTOMPADDING',(0,0),(-1,-1),0),
        ('LINEABOVE',(0,0),(-1,0),1,C_PRI),
    ]))
    story.append(sig)
    story.append(Spacer(1,18))

    # ── Footer ──
    story.append(HRFlowable(width="100%", thickness=0.5, color=C_BORD, spaceAfter=4, spaceBefore=2))
    story.append(Paragraph(
        "Zhongshan Bohui Advertising Craft Products Co., Ltd. | "
        "Factory: No.239 Fumin Avenue, Xiaolan Town, Zhongshan, Guangdong, China<br/>"
        f"Official Site: wa.bohui-sign.com | Tel: +86 13824779947 | {ref_no}",
        S["footer"]
    ))

    doc.build(story)

    # Return the PDF file
    return send_from_directory(pdf_dir, f"QUOTE_{safe_ref}.pdf",
                               mimetype="application/pdf",
                               as_attachment=True,
                               download_name=f"QUOTE_{safe_ref}.pdf")


# ==================== AI 润色正式报价 ====================

def _build_ai_polish_prompt(ptype, mat_label, is_electro, electro_color, kelvin,
                             items, total_qty, total_len_cm, power_desc, has_acrylic, is_rgb):
    """Build a prompt for AI to write professional product description."""
    item_str = ', '.join(f"{it['w']}×{it['h']}×{it['thick']}cm ×{it['qty']}pcs" for it in items)
    prompt = f"""You are a quotation writer for Zhongshan Bohui Advertising Craft Products Co., Ltd., a premium LED sign manufacturer. Write a clear, natural product description for this quotation in English. The tone should be professional yet approachable — imagine explaining the product to a potential customer in a friendly, confident way.

PRODUCT TYPE: {ptype}
MATERIAL: {mat_label}{f', {electro_color} electroplating finish' if is_electro else ''}
COLOR TEMPERATURE: {kelvin}
DIMENSIONS: {item_str} (Total {total_qty} characters)
POWER: {power_desc}
ACRYLIC: {"Yes, 5mm high-transparency milky white acrylic" if has_acrylic else "No"}
RGB: {"Yes, full-color dynamic LED" if is_rgb else "No"}

Please generate a clean quotation description with these 6 sections. Each section should be 1-2 sentences, natural business English — write it the way a helpful salesperson would speak to a customer:

1) PRODUCT: One-line product name
2) MATERIAL & FINISH: Describe the material and any surface treatment
3) CRAFT & CONSTRUCTION: How it's made (laser cutting, welding, polishing, etc.)
4) LED & LIGHTING: LED specs, color temp, illumination effect
5) INSTALLATION: How it's mounted
6) PACKAGING: 15mm plywood crate with foam interior

IMPORTANT: Do NOT mention any markup, discount, or internal pricing. This is for a customer-facing quotation. Sound natural and confident, not stiff.
Respond in this format:
PRODUCT: <text>
MATERIAL: <text>
CRAFT: <text>
LIGHTING: <text>
INSTALLATION: <text>
PACKAGING: <text>"""
    return prompt


@app.route("/api/calc/ai-quote", methods=["POST"])
def api_calc_ai_quote():
    """Generate AI-polished formal quotation PDF (customer-ready)."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.colors import HexColor, white
    from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
    )
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    # Register fonts
    if sys.platform == "win32":
        FONT_DIRS = ["C:/Windows/Fonts"]
    else:
        FONT_DIRS = ["/usr/share/fonts/crm", "/usr/share/fonts"]
    for f in [("SimSun","simsun.ttc"),("SimHei","simhei.ttf"),
               ("MSYH","msyh.ttc"),("MSYHBD","msyhbd.ttc"),
               ("Arial","arial.ttf"),("ArialBD","arialbd.ttf")]:
        for fd in FONT_DIRS:
            fp = os.path.join(fd, f[1])
            if os.path.exists(fp):
                try:
                    pdfmetrics.registerFont(TTFont(f[0], fp))
                    break
                except:
                    pass

    C_PRI = HexColor("#1a237e")
    C_ACC = HexColor("#283593")
    C_TALT = HexColor("#f5f5f5")
    C_DARK = HexColor("#212121")
    C_MID = HexColor("#616161")
    C_BORD = HexColor("#c5cae9")
    C_RED = HexColor("#c62828")

    PAGE_W, PAGE_H = A4
    ML, MR, MT, MB = 50, 50, 50, 45
    CW = PAGE_W - ML - MR

    def ps(name, font, size, color=C_DARK, align=TA_LEFT, leading=None, before=0, after=0):
        return ParagraphStyle(name, fontName=font, fontSize=size, textColor=color,
                              alignment=align, leading=leading or size*1.4,
                              spaceBefore=before, spaceAfter=after)

    S = {
        "co_name":ps("co_name","ArialBD",22,C_PRI,leading=28),
        "co_addr":ps("co_addr","Arial",9,C_DARK,leading=13),
        "ref":ps("ref","ArialBD",12,C_PRI,TA_RIGHT,leading=15),
        "date":ps("date","Arial",9,C_MID,TA_RIGHT,leading=12),
        "status":ps("status","ArialBD",13,C_PRI,TA_CENTER,leading=17),
        "client":ps("client","Arial",10,C_DARK,leading=14),
        "sec_hdr":ps("sec_hdr","ArialBD",11,C_PRI,leading=16),
        "spec_lbl":ps("spec_lbl","ArialBD",8.5,C_ACC,leading=13),
        "spec_it":ps("spec_it","Arial",8.5,C_DARK,leading=13),
        "tbl_hdr":ps("tbl_hdr","ArialBD",8,white,TA_CENTER,leading=10),
        "tbl_cell":ps("tbl_cell","Arial",8,C_DARK,TA_CENTER,leading=10),
        "tbl_l":ps("tbl_l","Arial",8,C_DARK,TA_LEFT,leading=10),
        "total_lbl":ps("total_lbl","ArialBD",10,C_DARK,TA_RIGHT,leading=13),
        "total_val":ps("total_val","ArialBD",14,C_PRI,TA_CENTER,leading=17),
        "sig_lbl":ps("sig_lbl","Arial",8,C_MID,TA_RIGHT,leading=10),
        "sig_txt":ps("sig_txt","ArialBD",10,C_DARK,leading=13),
        "footer":ps("footer","Arial",7,C_MID,TA_CENTER,leading=9),
        "note":ps("note","Arial",8,C_DARK,leading=12),
    }

    data = request.json or {}
    ref_no = data.get("ref", "BH-000000-000")
    date_str = data.get("date", "")
    client_name = data.get("client", "—")
    currency = data.get("currency", "CNY")
    cust_rate = float(data.get("custRate", 6.6))
    ptype = data.get("productType", "front")
    mat_label = data.get("materialLabel", "Stainless Steel")
    is_electro = data.get("isElectro", False)
    electro_color = data.get("electroColor", "")
    kelvin = data.get("kelvin", "6000K-6500K")
    items = data.get("items", [{"w":50,"h":50,"thick":8,"qty":1}])
    total_qty = data.get("totalQty", sum(it.get("qty",1) for it in items))
    total_len_cm = data.get("totalLenCm", 0)
    power_desc = data.get("powerDesc", "")
    has_acrylic = data.get("hasAcrylic", False)
    is_rgb = data.get("isRgb", False)
    formal_total = data.get("formalTotal", "0")
    product_label = data.get("productLabel", "LED Letters")

    # Call AI to polish descriptions
    prompt = _build_ai_polish_prompt(ptype, mat_label, is_electro, electro_color, kelvin,
                                      items, total_qty, total_len_cm, power_desc, has_acrylic, is_rgb)
    try:
        ai_result = ask_ali(prompt, "", max_tokens=2000, timeout=60)
    except Exception as e:
        ai_result = None
        print(f"[AI Quote] AI call failed: {e}")

    # Parse AI result
    ai_sections = {"PRODUCT": product_label, "MATERIAL": "", "CRAFT": "", "LIGHTING": "", "INSTALLATION": "", "PACKAGING": ""}
    if ai_result:
        for line in ai_result.split("\n"):
            line = line.strip()
            for key in ai_sections:
                if line.upper().startswith(key + ":"):
                    ai_sections[key] = line[len(key)+1:].strip()
                    break

    # Build specs list
    specs_list = [
        ("Product Type:", f"{ai_sections['PRODUCT']}"),
        ("Material & Finish:", ai_sections["MATERIAL"] or f"{mat_label}{' with '+electro_color+' electroplating' if is_electro else ''}"),
        ("Craft & Construction:", ai_sections["CRAFT"] or "Precision laser-cut, seamless welding, grinding and polishing fine finish."),
        ("LED & Lighting:", ai_sections["LIGHTING"] or f"Color temperature: {kelvin}. Even illumination."),
        ("Power Supply:", power_desc or "As per specification"),
        ("Dimensions:", ', '.join(f"{it['w']}×{it['h']}×{it['thick']}cm ×{it['qty']}" for it in items)),
        ("Installation:", ai_sections["INSTALLATION"] or "Includes mounting hardware and 1:1 installation template."),
        ("Packaging:", ai_sections["PACKAGING"] or "15mm thick plywood crate, foam interior. Suitable for international shipping."),
    ]

    # Generate clean PDF
    pdf_dir = os.path.join(UPLOAD_DIR, "quotes")
    os.makedirs(pdf_dir, exist_ok=True)
    safe_ref = ref_no.replace("/","-").replace("\\","-")
    output_path = os.path.join(pdf_dir, f"FORMAL_QUOTE_{safe_ref}.pdf")

    doc = SimpleDocTemplate(output_path, pagesize=A4,
                            leftMargin=ML, rightMargin=MR,
                            topMargin=MT, bottomMargin=MB,
                            title=f"Bohui Formal Quote {ref_no}",
                            author="Bohui Advertising Craft")
    story = []

    # ── HEADER ──
    sym = "$" if currency == "USD" else "¥"
    left = [
        Paragraph("Zhongshan Bohui Advertising Craft Products Co., Ltd.", S["co_name"]),
        Paragraph("Factory: No.239 Fumin Avenue, Xiaolan Town, Zhongshan, Guangdong, China 528415", S["co_addr"]),
        Paragraph(f"Official Site: wa.bohui-sign.com | Tel: +86 13824779947", S["co_addr"]),
    ]
    right = [
        Paragraph(f"QUOTATION REF: {ref_no}", S["ref"]),
        Paragraph(f"DATE: {date_str}", S["date"]),
    ]
    hdr = Table([[left, right]], colWidths=[CW*0.6, CW*0.4])
    hdr.setStyle(TableStyle([
        ('VALIGN',(0,0),(-1,-1),'TOP'),
        ('LEFTPADDING',(0,0),(0,0),0),
        ('RIGHTPADDING',(1,1),(1,1),0),
        ('TOPPADDING',(0,0),(-1,-1),0),
        ('BOTTOMPADDING',(0,0),(-1,-1),0),
    ]))
    story.append(hdr)
    story.append(Spacer(1,4))
    story.append(HRFlowable(width="100%", thickness=1, color=C_PRI, spaceAfter=6, spaceBefore=2))
    story.append(Paragraph("QUOTATION", S["status"]))
    story.append(HRFlowable(width="100%", thickness=0.5, color=C_BORD, spaceAfter=8, spaceBefore=2))

    # CLIENT
    story.append(Paragraph(f"<b>TO:</b> {client_name}", ps("client","Arial",10,C_DARK,leading=14,before=4,after=8)))

    # ── SECTION I: SPECIFICATIONS ──
    story.append(Paragraph("I. PRODUCT SPECIFICATIONS", S["sec_hdr"]))
    story.append(Spacer(1,2))
    for label, desc in specs_list:
        story.append(Paragraph(f"<b>{label}</b> {desc}", S["spec_it"]))
    story.append(Spacer(1,10))

    # ── SECTION II: PRICE ──
    story.append(Paragraph("II. QUOTATION PRICE", S["sec_hdr"]))
    story.append(Spacer(1,6))

    # Clean the formal total - remove currency symbols and commas
    clean_total = formal_total.replace("¥","").replace("$","").replace("￥","").replace(",","").strip()

    # Try to convert with customer rate
    try:
        total_num = float(clean_total)
        if currency == "USD":
            total_usd = round(total_num / cust_rate, 2)
            total_display = f"¥{clean_total} (≈ ${total_usd} @ {cust_rate})"
        else:
            total_display = f"¥{clean_total}"
    except:
        total_display = formal_total

    # Price table - clean, no markup
    price_col_w = [CW*0.50, CW*0.15, CW*0.35]
    tbl_data = [
        [Paragraph("Description", S["tbl_hdr"]),
         Paragraph("Quantity", S["tbl_hdr"]),
         Paragraph(f"Total Price ({currency})", S["tbl_hdr"])],
    ]
    item_desc = f"{product_label} - {mat_label}"
    if is_electro:
        item_desc += f" ({electro_color} electroplating)"
    tbl_data.append([
        Paragraph(item_desc, S["tbl_l"]),
        Paragraph(str(total_qty), S["tbl_cell"]),
        Paragraph(total_display, S["tbl_cell"]),
    ])

    # Net total row
    tbl_data.append([
        Paragraph("Total Amount", ps("net_lbl","ArialBD",9,C_DARK,TA_RIGHT,leading=11)),
        Paragraph("", S["tbl_cell"]),
        Paragraph(total_display, ps("net_val","ArialBD",12,C_PRI,TA_CENTER,leading=14)),
    ])

    tbl = Table(tbl_data, colWidths=price_col_w)
    tbl.setStyle(TableStyle([
        ('BACKGROUND',(0,0),(-1,0),C_PRI),
        ('TEXTCOLOR',(0,0),(-1,0),white),
        ('ALIGN',(0,0),(-1,-1),'CENTER'),
        ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
        ('GRID',(0,0),(-1,1),0.5,C_BORD),
        ('TOPPADDING',(0,0),(-1,-1),6),
        ('BOTTOMPADDING',(0,0),(-1,-1),6),
        ('LEFTPADDING',(0,0),(-1,-1),6),
        ('RIGHTPADDING',(0,0),(-1,-1),6),
        ('BACKGROUND',(0,1),(-1,1),C_TALT),
        ('SPAN',(0,-1),(1,-1)),
        ('LINEABOVE',(0,-1),(-1,-1),1.5,C_PRI),
        ('LINEBELOW',(0,-1),(-1,-1),0.5,C_BORD),
        ('TOPPADDING',(0,-1),(-1,-1),8),
        ('BOTTOMPADDING',(0,-1),(-1,-1),8),
    ]))
    story.append(tbl)
    story.append(Spacer(1,12))

    # ── Shipping Options ──
    story.append(HRFlowable(width="100%", thickness=0.3, color=C_BORD, spaceAfter=4, spaceBefore=2))
    story.append(Paragraph("<b>SHIPPING OPTIONS (DDU — Delivered Duty Unpaid):</b>", S["note"]))
    sea = "$3,969" if currency == "USD" else "¥3,969"
    air = "$7,677" if currency == "USD" else "¥7,677"
    story.append(Paragraph(f"   Sea freight: {sea} — Transit: 28-35 days after sailing.", S["note"]))
    story.append(Paragraph(f"   Air freight: {air} — Transit: 3-5 working days after departure.", S["note"]))
    story.append(Spacer(1,8))

    # ── Notes ──
    story.append(HRFlowable(width="100%", thickness=0.3, color=C_BORD, spaceAfter=4, spaceBefore=2))
    story.append(Paragraph("<b>TERMS:</b>", ps("nb","ArialBD",8.5,C_ACC,leading=13)))
    notes = [
        "Payment: 50% deposit to start production. 50% balance before shipment.",
        "Production Lead Time: 10 working days (counted from day after deposit received).",
        "Validity: This quotation is valid for 30 days from the date of issue.",
        "All-inclusive: Prices cover materials, processing, accessories, crate packing, and shipping as selected.",
        "Exchange Rate: Reference rate 1 USD = " + str(cust_rate) + " CNY. Final exchange rate may vary at time of payment.",
    ]
    for n in notes:
        story.append(Paragraph(f"• {n}", S["note"]))
    story.append(Spacer(1,12))

    # ── Signature ──
    sig = Table([
        [Paragraph("Authorized Signature:", S["sig_lbl"]),
         Paragraph("Yang Junliang<br/><font size=7>COO | Zhongshan Bohui Advertising Craft Products Co., Ltd.</font>",
                   S["sig_txt"])]
    ], colWidths=[CW*0.30, CW*0.70])
    sig.setStyle(TableStyle([
        ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
        ('LEFTPADDING',(0,0),(-1,-1),0),
        ('RIGHTPADDING',(0,0),(-1,-1),0),
        ('TOPPADDING',(0,0),(-1,-1),0),
        ('BOTTOMPADDING',(0,0),(-1,-1),0),
        ('LINEABOVE',(0,0),(-1,0),1,C_PRI),
    ]))
    story.append(sig)
    story.append(Spacer(1,18))

    # ── Footer ──
    story.append(HRFlowable(width="100%", thickness=0.5, color=C_BORD, spaceAfter=4, spaceBefore=2))
    story.append(Paragraph(
        "Zhongshan Bohui Advertising Craft Products Co., Ltd. | "
        "Factory: No.239 Fumin Avenue, Xiaolan Town, Zhongshan, Guangdong, China<br/>"
        f"Official Site: wa.bohui-sign.com | Tel: +86 13824779947 | {ref_no}",
        S["footer"]
    ))

    doc.build(story)

    return send_from_directory(pdf_dir, f"FORMAL_QUOTE_{safe_ref}.pdf",
                               mimetype="application/pdf",
                               as_attachment=True,
                               download_name=f"FORMAL_QUOTE_{safe_ref}.pdf")


@app.route("/api/calc/save-quote", methods=["POST"])
def api_calc_save_quote():
    """Auto-save AI-generated quotation to CRM quotes table."""
    data = request.json or {}
    client_id = data.get("clientId", "")
    if not client_id:
        return jsonify({"ok": False, "error": "No client selected"}), 400

    ref = data.get("ref", "BH-" + datetime.now().strftime("%Y%m%d") + "-001")
    formal_total = data.get("formalTotal", "0")
    clean = formal_total.replace("¥","").replace("$","").replace(",","").strip()
    try:
        amt = float(clean)
    except:
        amt = 0

    from datetime import date, timedelta
    valid_until = (date.today() + timedelta(days=30)).isoformat()

    quote_data = {
        "customer_id": client_id,
        "quote_no": ref,
        "currency": data.get("currency", "CNY"),
        "total_amount": amt,
        "valid_until": valid_until,
        "status": "draft",
        "notes": "AI-Generated Formal Quotation (auto-saved)",
        "created_by": session.get("user_id"),
        "items": [{
            "product_id": "",
            "name": f"{data.get('productLabel','LED Letters')} - {data.get('materialLabel','')}",
            "qty": data.get("totalQty", 1),
            "unit": "set",
            "unit_price": round(amt / max(data.get("totalQty", 1), 1), 2),
            "total": amt
        }],
    }
    try:
        qid = add_quote(quote_data)
        return jsonify({"ok": True, "id": qid})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/quotes/<int:qid>/ai-quote", methods=["POST"])
def api_quote_ai_quote(qid):
    """Generate AI-polished formal quotation PDF from a saved CRM quote."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.colors import HexColor, white
    from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
    )
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    FONT_DIR = "C:/Windows/Fonts"
    for f in [("SimSun","simsun.ttc"),("SimHei","simhei.ttc"),
               ("MSYH","msyh.ttc"),("MSYHBD","msyhbd.ttc"),
               ("Arial","arial.ttf"),("ArialBD","arialbd.ttf")]:
        try:
            pdfmetrics.registerFont(TTFont(f[0], os.path.join(FONT_DIR, f[1])))
        except:
            pass

    C_PRI = HexColor("#1a237e")
    C_ACC = HexColor("#283593")
    C_TALT = HexColor("#f5f5f5")
    C_DARK = HexColor("#212121")
    C_MID = HexColor("#616161")
    C_BORD = HexColor("#c5cae9")

    PAGE_W, PAGE_H = A4
    ML, MR, MT, MB = 50, 50, 50, 45
    CW = PAGE_W - ML - MR

    def ps(name, font, size, color=C_DARK, align=TA_LEFT, leading=None, before=0, after=0):
        return ParagraphStyle(name, fontName=font, fontSize=size, textColor=color,
                              alignment=align, leading=leading or size*1.4,
                              spaceBefore=before, spaceAfter=after)

    S = {
        "co_name":ps("co_name","ArialBD",22,C_PRI,leading=28),
        "co_addr":ps("co_addr","Arial",9,C_DARK,leading=13),
        "ref":ps("ref","ArialBD",12,C_PRI,TA_RIGHT,leading=15),
        "date":ps("date","Arial",9,C_MID,TA_RIGHT,leading=12),
        "status":ps("status","ArialBD",13,C_PRI,TA_CENTER,leading=17),
        "client":ps("client","Arial",10,C_DARK,leading=14),
        "sec_hdr":ps("sec_hdr","ArialBD",11,C_PRI,leading=16),
        "spec_it":ps("spec_it","Arial",8.5,C_DARK,leading=13),
        "tbl_hdr":ps("tbl_hdr","ArialBD",8,white,TA_CENTER,leading=10),
        "tbl_cell":ps("tbl_cell","Arial",8,C_DARK,TA_CENTER,leading=10),
        "tbl_l":ps("tbl_l","Arial",8,C_DARK,TA_LEFT,leading=10),
        "total_lbl":ps("total_lbl","ArialBD",10,C_DARK,TA_RIGHT,leading=13),
        "total_val":ps("total_val","ArialBD",14,C_PRI,TA_CENTER,leading=17),
        "sig_lbl":ps("sig_lbl","Arial",8,C_MID,TA_RIGHT,leading=10),
        "sig_txt":ps("sig_txt","ArialBD",10,C_DARK,leading=13),
        "footer":ps("footer","Arial",7,C_MID,TA_CENTER,leading=9),
        "note":ps("note","Arial",8,C_DARK,leading=12),
    }

    # Load quote from database
    quote = get_quote(qid)
    if not quote:
        return jsonify({"ok": False, "error": "Quote not found"}), 404

    ref_no = quote.get("quote_no", f"Q-{qid}")
    date_str = (quote.get("created_at") or "")[:10]
    client_name = quote.get("customer_name", "—")
    currency = quote.get("currency", "CNY")
    total_amt = float(quote.get("total_amount", 0))
    notes = quote.get("notes", "")
    items_raw = quote.get("items", "[]")
    try:
        items = json.loads(items_raw) if isinstance(items_raw, str) else items_raw
    except:
        items = []

    sym = "$" if currency == "USD" else "¥"
    total_display = f"{sym}{total_amt:.2f}"

    # Build quote items list for PDF
    if items and isinstance(items, list):
        item_desc = items[0].get("name", "LED Signage") if len(items) > 0 else "LED Signage"
        item_qty = str(items[0].get("qty", 1)) if len(items) > 0 else "1"
    else:
        item_desc = "LED Signage"
        item_qty = "1"

    # Generate PDF
    pdf_dir = os.path.join(UPLOAD_DIR, "quotes")
    os.makedirs(pdf_dir, exist_ok=True)
    safe_ref = ref_no.replace("/","-").replace("\\","-")
    output_path = os.path.join(pdf_dir, f"FORMAL_QUOTE_{safe_ref}.pdf")

    doc = SimpleDocTemplate(output_path, pagesize=A4,
                            leftMargin=ML, rightMargin=MR,
                            topMargin=MT, bottomMargin=MB,
                            title=f"Bohui Formal Quote {ref_no}",
                            author="Bohui Advertising Craft")
    story = []

    # ── HEADER ──
    left = [
        Paragraph("Zhongshan Bohui Advertising Craft Products Co., Ltd.", S["co_name"]),
        Paragraph("Factory: No.239 Fumin Avenue, Xiaolan Town, Zhongshan, Guangdong, China 528415", S["co_addr"]),
        Paragraph("Official Site: wa.bohui-sign.com | Tel: +86 13824779947", S["co_addr"]),
    ]
    right = [
        Paragraph(f"QUOTATION REF: {ref_no}", S["ref"]),
        Paragraph(f"DATE: {date_str}", S["date"]),
    ]
    hdr = Table([[left, right]], colWidths=[CW*0.6, CW*0.4])
    hdr.setStyle(TableStyle([
        ('VALIGN',(0,0),(-1,-1),'TOP'),
        ('LEFTPADDING',(0,0),(0,0),0),
        ('RIGHTPADDING',(1,1),(1,1),0),
        ('TOPPADDING',(0,0),(-1,-1),0),
        ('BOTTOMPADDING',(0,0),(-1,-1),0),
    ]))
    story.append(hdr)
    story.append(Spacer(1,4))
    story.append(HRFlowable(width="100%", thickness=1, color=C_PRI, spaceAfter=6, spaceBefore=2))
    story.append(Paragraph("QUOTATION", S["status"]))
    story.append(HRFlowable(width="100%", thickness=0.5, color=C_BORD, spaceAfter=8, spaceBefore=2))

    # CLIENT
    story.append(Paragraph(f"<b>TO:</b> {client_name}", ps("client","Arial",10,C_DARK,leading=14,before=4,after=8)))

    # ── SECTION I: PRODUCT ──
    story.append(Paragraph("I. PRODUCT SPECIFICATIONS", S["sec_hdr"]))
    story.append(Spacer(1,2))
    story.append(Paragraph(f"<b>Product:</b> {item_desc}", S["spec_it"]))

    # Try AI polish
    try:
        prompt = _build_ai_polish_prompt("general", "", False, "", "6000K",
                                          [{"w":50,"h":50,"thick":8,"qty":1}],
                                          1, 0, "Standard", False, False)
        ai_result = ask_ali(prompt, "", max_tokens=2000, timeout=60)
        if ai_result:
            for line in ai_result.split("\n"):
                line = line.strip()
                for key_start in ["PRODUCT:", "MATERIAL:", "CRAFT:", "LIGHTING:", "INSTALLATION:", "PACKAGING:"]:
                    if line.upper().startswith(key_start):
                        val = line[len(key_start):].strip()
                        if val:
                            lbl = key_start.replace(":", "")
                            story.append(Paragraph(f"<b>{lbl}:</b> {val}", S["spec_it"]))
    except Exception as e:
        print(f"[AI Quote] AI call failed for saved quote: {e}")
        story.append(Paragraph(f"<b>Craft & Construction:</b> Precision manufactured LED signage.", S["spec_it"]))

    if notes:
        story.append(Paragraph(f"<b>Notes:</b> {notes}", S["spec_it"]))
    story.append(Spacer(1,10))

    # ── SECTION II: PRICE ──
    story.append(Paragraph("II. QUOTATION PRICE", S["sec_hdr"]))
    story.append(Spacer(1,6))

    price_col_w = [CW*0.50, CW*0.15, CW*0.35]
    tbl_data = [
        [Paragraph("Description", S["tbl_hdr"]),
         Paragraph("Quantity", S["tbl_hdr"]),
         Paragraph(f"Total Price ({currency})", S["tbl_hdr"])],
    ]
    tbl_data.append([
        Paragraph(item_desc, S["tbl_l"]),
        Paragraph(item_qty, S["tbl_cell"]),
        Paragraph(total_display, S["tbl_cell"]),
    ])
    tbl_data.append([
        Paragraph("Total Amount", ps("nl","ArialBD",9,C_DARK,TA_RIGHT,leading=11)),
        Paragraph("", S["tbl_cell"]),
        Paragraph(total_display, ps("nv","ArialBD",12,C_PRI,TA_CENTER,leading=14)),
    ])

    tbl = Table(tbl_data, colWidths=price_col_w)
    tbl.setStyle(TableStyle([
        ('BACKGROUND',(0,0),(-1,0),C_PRI),
        ('TEXTCOLOR',(0,0),(-1,0),white),
        ('ALIGN',(0,0),(-1,-1),'CENTER'),
        ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
        ('GRID',(0,0),(-1,1),0.5,C_BORD),
        ('TOPPADDING',(0,0),(-1,-1),6),
        ('BOTTOMPADDING',(0,0),(-1,-1),6),
        ('LEFTPADDING',(0,0),(-1,-1),6),
        ('RIGHTPADDING',(0,0),(-1,-1),6),
        ('BACKGROUND',(0,1),(-1,1),C_TALT),
        ('SPAN',(0,-1),(1,-1)),
        ('LINEABOVE',(0,-1),(-1,-1),1.5,C_PRI),
        ('LINEBELOW',(0,-1),(-1,-1),0.5,C_BORD),
        ('TOPPADDING',(0,-1),(-1,-1),8),
        ('BOTTOMPADDING',(0,-1),(-1,-1),8),
    ]))
    story.append(tbl)
    story.append(Spacer(1,12))

    # ── Shipping ──
    story.append(HRFlowable(width="100%", thickness=0.3, color=C_BORD, spaceAfter=4, spaceBefore=2))
    story.append(Paragraph("<b>SHIPPING OPTIONS (DDU):</b>", S["note"]))
    sea = "$3,969" if currency == "USD" else "¥3,969"
    air = "$7,677" if currency == "USD" else "¥7,677"
    story.append(Paragraph(f"   Sea freight: {sea} — 28-35 days", S["note"]))
    story.append(Paragraph(f"   Air freight: {air} — 3-5 working days", S["note"]))
    story.append(Spacer(1,8))

    # ── Terms ──
    story.append(HRFlowable(width="100%", thickness=0.3, color=C_BORD, spaceAfter=4, spaceBefore=2))
    story.append(Paragraph("<b>TERMS:</b>", ps("nb","ArialBD",8.5,C_ACC,leading=13)))
    for n in [
        "Payment: 50% deposit to start production. 50% balance before shipment.",
        "Lead Time: 10 working days after deposit received.",
        "Validity: 30 days from date of issue.",
    ]:
        story.append(Paragraph(f"• {n}", S["note"]))
    story.append(Spacer(1,12))

    # ── Signature ──
    sig = Table([
        [Paragraph("Authorized Signature:", S["sig_lbl"]),
         Paragraph("Yang Junliang<br/><font size=7>COO | Zhongshan Bohui Advertising Craft Products Co., Ltd.</font>",
                   S["sig_txt"])]
    ], colWidths=[CW*0.30, CW*0.70])
    sig.setStyle(TableStyle([
        ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
        ('LEFTPADDING',(0,0),(-1,-1),0),
        ('RIGHTPADDING',(0,0),(-1,-1),0),
        ('TOPPADDING',(0,0),(-1,-1),0),
        ('BOTTOMPADDING',(0,0),(-1,-1),0),
        ('LINEABOVE',(0,0),(-1,0),1,C_PRI),
    ]))
    story.append(sig)
    story.append(Spacer(1,18))

    # ── Footer ──
    story.append(HRFlowable(width="100%", thickness=0.5, color=C_BORD, spaceAfter=4, spaceBefore=2))
    story.append(Paragraph(
        "Zhongshan Bohui Advertising Craft Products Co., Ltd. | "
        "Factory: No.239 Fumin Avenue, Xiaolan Town, Zhongshan, Guangdong, China<br/>"
        f"Official Site: wa.bohui-sign.com | Tel: +86 13824779947 | {ref_no}",
        S["footer"]
    ))

    doc.build(story)

    pdf_url = f"/uploads/quotes/FORMAL_QUOTE_{safe_ref}.pdf"
    return jsonify({
        "ok": True,
        "url": pdf_url,
        "filename": f"FORMAL_QUOTE_{safe_ref}.pdf"
    })


# ==================== 用户认证 ====================
from functools import wraps

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return jsonify({"error": "请先登录"}), 401
        return f(*args, **kwargs)
    return decorated


def role_required(*roles):
    """Decorator: require the logged-in user to have one of the given roles."""
    def decorator(f):
        @wraps(f)
        def decorated(*args, **kwargs):
            if 'user_id' not in session:
                return jsonify({"error": "请先登录"}), 401
            if session.get("role") not in roles:
                return jsonify({"error": "权限不足"}), 403
            return f(*args, **kwargs)
        return decorated
    return decorator


@app.route("/api/auth/login", methods=["POST"])
def api_auth_login():
    data = request.json or {}
    username = data.get("username", "").strip()
    password = data.get("password", "").strip()
    if not username or not password:
        return jsonify({"error": "用户名和密码不能为空"}), 400
    from werkzeug.security import check_password_hash
    user = get_user_by_username(username)
    if not user or not check_password_hash(user["password_hash"], password):
        return jsonify({"error": "用户名或密码错误"}), 401
    if not user.get("active", 1):
        return jsonify({"error": "账号已被禁用"}), 403
    session["user_id"] = user["id"]
    session["username"] = user["username"]
    session["display_name"] = user["display_name"]
    session["role"] = user["role"]
    return jsonify({
        "ok": True,
        "user": {
            "id": user["id"],
            "username": user["username"],
            "display_name": user["display_name"],
            "role": user["role"],
            "title": user.get("title", ""),
        }
    })


@app.route("/api/auth/logout", methods=["POST"])
def api_auth_logout():
    session.clear()
    return jsonify({"ok": True})


@app.route("/api/auth/me")
def api_auth_me():
    if 'user_id' not in session:
        return jsonify({"error": "未登录"}), 401
    user = get_user_by_username(session.get("username", ""))
    if not user:
        return jsonify({"error": "用户不存在"}), 404
    return jsonify({
        "id": user["id"],
        "username": user["username"],
        "display_name": user["display_name"],
        "role": user["role"],
        "title": user.get("title", ""),
    })


# ==================== 通知 API ====================
@app.route("/api/notifications")
@login_required
def api_notifications():
    return jsonify(get_notifications(session["user_id"]))


@app.route("/api/notifications/unread-count")
@login_required
def api_unread_count():
    return jsonify({"count": get_unread_count(session["user_id"])})


@app.route("/api/notifications/<int:nid>/read", methods=["POST"])
@login_required
def api_mark_read(nid):
    mark_notification_read(nid, session["user_id"])
    return jsonify({"ok": True})


@app.route("/api/notifications/read-all", methods=["POST"])
@login_required
def api_mark_all_read():
    mark_all_notifications_read(session["user_id"])
    return jsonify({"ok": True})


@app.route("/api/auth/users")
@login_required
def api_auth_users():
    if session.get("role") != "admin":
        return jsonify({"error": "仅管理员可查看用户列表"}), 403
    return jsonify(get_users())


@app.route("/api/auth/users", methods=["POST"])
@login_required
def api_auth_add_user():
    if session.get("role") != "admin":
        return jsonify({"error": "仅管理员可创建用户"}), 403
    data = request.json or {}
    result = add_user(data)
    if "error" in result:
        return jsonify(result), 400
    logged_uid = session.get("user_id")
    if logged_uid and result.get("id"):
        add_activity_log(logged_uid, "create", "user", result["id"],
            f"创建了用户 {data.get('username','')} ({data.get('display_name','')})")
    return jsonify(result)


@app.route("/api/auth/users/<int:uid>", methods=["PUT"])
@login_required
def api_auth_update_user(uid):
    if session.get("role") != "admin":
        return jsonify({"error": "仅管理员可修改用户"}), 403
    data = request.json or {}
    update_user(uid, data)
    logged_uid = session.get("user_id")
    if logged_uid:
        changed = [f"{k}={v}" for k, v in data.items()]
        add_activity_log(logged_uid, "update", "user", uid,
            f"修改了用户 ID:{uid}: {', '.join(changed)}")
    return jsonify({"ok": True})


# ==================== 订单管理 ====================
@app.route("/api/orders")
@login_required
def api_orders():
    status = request.args.get("status")
    customer_id = request.args.get("customer_id")
    if status:
        status = status.split(",")
    return jsonify(get_orders(status=status, customer_id=customer_id))


@app.route("/api/orders/<int:oid>")
@login_required
def api_order(oid):
    order = get_order(oid)
    if not order:
        return jsonify({"error": "订单不存在"}), 404
    return jsonify(order)


@app.route("/api/orders", methods=["POST"])
@login_required
def api_create_order():
    data = request.json or {}
    if not data.get("customer_id"):
        return jsonify({"error": "缺少 customer_id"}), 400
    data["created_by"] = session.get("user_id")
    result = add_order(data)
    if "error" in result:
        return jsonify(result), 400
    uid = session.get("user_id")
    if uid and result.get("id"):
        add_activity_log(uid, "create", "order", result["id"],
            f"创建了订单 {result.get('order_no','')} (ID:{result['id']})")
    return jsonify({"ok": True, "id": result["id"], "order_no": result["order_no"]})


@app.route("/api/orders/<int:oid>", methods=["PUT"])
@login_required
def api_update_order(oid):
    data = request.json or {}
    existing = get_order(oid)
    uid = session.get("user_id")
    # Track status change
    if "status" in data and existing:
        if existing.get("status") != data["status"]:
            old_status = existing.get("status", "")
            new_status = data["status"]
            add_timeline_entry(oid, new_status, session.get("user_id", 0),
                               f"状态变更: {old_status} → {new_status}")
            # Send WhatsApp notification for meaningful transitions
            if new_status in ("confirmed", "in_production", "shipped", "delivered"):
                customer = get_customer(existing.get("customer_id"))
                if customer and customer.get("whatsapp"):
                    def _notify_job():
                        _notify_order_status_change(existing, customer, old_status, new_status)
                    threading.Thread(target=_notify_job, daemon=True).start()
            # 通知所有活跃用户订单状态变更
            for auid in get_active_user_ids():
                add_notification(auid, 'order',
                    f'订单 {existing["order_no"]} 状态变更',
                    f'{old_status} → {new_status}',
                    link=f'/orders/{oid}',
                    related_type='order', related_id=oid)
            # Log the status change
            if uid:
                add_activity_log(uid, "update", "order", oid,
                    f"修改了订单 {existing.get('order_no','')}: 状态 {old_status}→{new_status}")
    update_order(oid, data)
    return jsonify({"ok": True})


@app.route("/api/orders/<int:oid>", methods=["DELETE"])
@login_required
def api_delete_order(oid):
    existing = get_order(oid)
    delete_order(oid)
    uid = session.get("user_id")
    if uid and existing:
        add_activity_log(uid, "delete", "order", oid,
            f"删除了订单 {existing.get('order_no','')} (ID:{oid})")
    return jsonify({"ok": True})


@app.route("/api/orders/<int:oid>/payment", methods=["POST"])
@login_required
def api_order_payment(oid):
    data = request.json or {}
    ptype = data.get("type", "deposit")  # 'deposit' or 'balance'
    update_data = {}
    if ptype == "deposit":
        update_data["deposit_amount"] = data.get("amount", 0)
        update_data["deposit_date"] = data.get("date", datetime.now().strftime("%Y-%m-%d"))
        update_data["deposit_method"] = data.get("method", "")
        update_data["deposit_received"] = 1
    else:
        update_data["balance_amount"] = data.get("amount", 0)
        update_data["balance_date"] = data.get("date", datetime.now().strftime("%Y-%m-%d"))
        update_data["balance_method"] = data.get("method", "")
        update_data["balance_received"] = 1
    update_order(oid, update_data)
    # 同步写入付款流水表
    order_currency = "USD"
    o_check = get_order(oid)
    if o_check:
        order_currency = o_check.get("currency", "USD")
    add_payment(
        order_id=oid,
        ptype=ptype,
        amount=data.get("amount", 0),
        currency=order_currency,
        payment_date=data.get("date", datetime.now().strftime("%Y-%m-%d")),
        method=data.get("method", ""),
        recorded_by=session.get("user_id"),
    )
    add_timeline_entry(oid, "payment", session.get("user_id", 0),
                       f"收到{data.get('method','')}付款 {data.get('amount',0)} ({'定金' if ptype=='deposit' else '尾款'})")
    # Send WhatsApp notification for payment
    order = get_order(oid)
    if order:
        customer = get_customer(order.get("customer_id"))
        if customer and customer.get("whatsapp"):
            def _pay_notify_job():
                _notify_payment_received(order, customer, ptype, data.get("amount", 0))
            threading.Thread(target=_pay_notify_job, daemon=True).start()
    uid = session.get("user_id")
    if uid and order:
        add_activity_log(uid, "payment", "order", oid,
            f"订单 {order.get('order_no','')} 收款 {data.get('amount',0)} ({'定金' if ptype=='deposit' else '尾款'})")
    return jsonify({"ok": True})


# ==================== 生产步骤追踪 API ====================
@app.route("/api/orders/<int:oid>/tasks")
@login_required
def api_get_production_tasks(oid):
    """获取订单的生产任务列表"""
    return jsonify(get_production_tasks(oid))


@app.route("/api/orders/<int:oid>/tasks", methods=["POST"])
@login_required
@role_required('admin', 'production')
def api_save_production_tasks(oid):
    """批量保存生产任务（管理员预设步骤）"""
    data = request.json or {}
    tasks = data.get("tasks", [])
    save_production_tasks(oid, tasks)
    uid = session.get("user_id")
    if uid:
        order = get_order(oid)
        add_activity_log(uid, "update", "order", oid,
            f"更新生产步骤 ({len(tasks)} 项) - {order.get('order_no','')}")
    return jsonify({"ok": True})


@app.route("/api/orders/<int:oid>/tasks/<int:tid>", methods=["PUT"])
@login_required
def api_update_task_status(oid, tid):
    """勾选/取消勾选单个任务"""
    data = request.json or {}
    is_done = data.get("is_done", 0)
    done_by = session.get("user_id")
    update_production_task_status(tid, is_done, done_by)
    return jsonify({"ok": True})


@app.route("/api/orders/<int:oid>/generate-tasks", methods=["POST"])
@login_required
@role_required('admin', 'production')
def api_generate_tasks(oid):
    """按默认模板为订单生成初始任务列表"""
    defaults = get_production_task_defaults()
    save_production_tasks(oid, defaults)
    uid = session.get("user_id")
    if uid:
        order = get_order(oid)
        add_activity_log(uid, "create", "order", oid,
            f"生成默认生产步骤 - {order.get('order_no','')}")
    return jsonify({"ok": True, "tasks": get_production_tasks(oid)})


# ==================== 发货管理 API ====================
@app.route("/api/orders/<int:oid>/shipments")
@login_required
def api_order_shipments(oid):
    """获取订单的发货记录列表"""
    return jsonify(get_shipments(oid))


@app.route("/api/orders/<int:oid>/shipments", methods=["POST"])
@login_required
def api_create_shipment(oid):
    """创建发货记录"""
    data = request.json or {}
    data["order_id"] = oid
    data["created_by"] = session.get("user_id")
    result = add_shipment(data)
    uid = session.get("user_id")
    if uid and result.get("id"):
        order = get_order(oid)
        add_activity_log(uid, "create", "shipment", result["id"],
            f"创建了发货记录 - 订单 {order.get('order_no','')}")
    return jsonify({"ok": True, "id": result["id"]})


@app.route("/api/shipments/<int:sid>")
@login_required
def api_get_shipment(sid):
    """获取单条发货记录"""
    s = get_shipment(sid)
    if not s:
        return jsonify({"error": "发货记录不存在"}), 404
    return jsonify(s)


@app.route("/api/shipments/<int:sid>", methods=["PUT"])
@login_required
def api_update_shipment(sid):
    """更新发货记录"""
    data = request.json or {}
    update_shipment(sid, data)
    uid = session.get("user_id")
    if uid:
        s = get_shipment(sid)
        if s:
            add_activity_log(uid, "update", "shipment", sid,
                f"更新了发货记录 - 订单 #{s.get('order_id','')}")
    return jsonify({"ok": True})


@app.route("/api/shipments/<int:sid>", methods=["DELETE"])
@login_required
def api_delete_shipment(sid):
    """删除发货记录"""
    s = get_shipment(sid)
    delete_shipment(sid)
    uid = session.get("user_id")
    if uid and s:
        add_activity_log(uid, "delete", "shipment", sid,
            f"删除了发货记录 - 订单 #{s.get('order_id','')}")
    return jsonify({"ok": True})


# ==================== 质检管理 API ====================
# ---- 质检模板 (admin only) ----
@app.route("/api/qc/templates")
@login_required
def api_qc_templates():
    """获取质检模板列表"""
    return jsonify(get_qc_templates())


@app.route("/api/qc/templates", methods=["POST"])
@login_required
@role_required('admin')
def api_create_qc_template():
    """创建质检模板"""
    data = request.json or {}
    save_qc_template(None, data.get("name", ""), data.get("items", []))
    uid = session.get("user_id")
    if uid:
        add_activity_log(uid, "create", "qc_template", 0,
            f"创建质检模板: {data.get('name','')}")
    return jsonify({"ok": True})


@app.route("/api/qc/templates/<int:tid>", methods=["PUT"])
@login_required
@role_required('admin')
def api_update_qc_template(tid):
    """更新质检模板"""
    data = request.json or {}
    save_qc_template(tid, data.get("name", ""), data.get("items", []))
    uid = session.get("user_id")
    if uid:
        add_activity_log(uid, "update", "qc_template", tid,
            f"更新质检模板: {data.get('name','')}")
    return jsonify({"ok": True})


@app.route("/api/qc/templates/<int:tid>", methods=["DELETE"])
@login_required
@role_required('admin')
def api_delete_qc_template(tid):
    """删除质检模板"""
    delete_qc_template(tid)
    uid = session.get("user_id")
    if uid:
        add_activity_log(uid, "delete", "qc_template", tid, "删除了质检模板")
    return jsonify({"ok": True})


# ---- 质检记录 ----
@app.route("/api/orders/<int:oid>/qc")
@login_required
def api_order_qc(oid):
    """获取订单的质检记录列表"""
    return jsonify(get_order_qc_inspections(oid))


@app.route("/api/orders/<int:oid>/qc", methods=["POST"])
@login_required
def api_create_qc(oid):
    """创建质检记录"""
    data = request.json or {}
    data["order_id"] = oid
    data["inspector_id"] = session.get("user_id")
    data["inspected_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    result = add_qc_inspection(data)
    uid = session.get("user_id")
    if uid and result.get("id"):
        order = get_order(oid)
        add_activity_log(uid, "create", "qc_inspection", result["id"],
            f"创建了质检记录 - 订单 {order.get('order_no','')} ({data.get('result','pending')})")
    return jsonify({"ok": True, "id": result["id"]})


@app.route("/api/qc/<int:iid>")
@login_required
def api_get_qc(iid):
    """获取单条质检记录"""
    ins = get_qc_inspection(iid)
    if not ins:
        return jsonify({"error": "质检记录不存在"}), 404
    return jsonify(ins)


@app.route("/api/qc/<int:iid>", methods=["PUT"])
@login_required
def api_update_qc(iid):
    """更新质检记录"""
    data = request.json or {}
    if "inspected_at" not in data:
        data["inspected_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    update_qc_inspection(iid, data)
    uid = session.get("user_id")
    if uid:
        ins = get_qc_inspection(iid)
        order_id = ins.get("order_id", "") if ins else ""
        add_activity_log(uid, "update", "qc_inspection", iid,
            f"更新了质检记录 (ID:{iid}) - 订单 #{order_id}")
    return jsonify({"ok": True})


@app.route("/api/qc/<int:iid>", methods=["DELETE"])
@login_required
def api_delete_qc(iid):
    """删除质检记录"""
    ins = get_qc_inspection(iid)
    delete_qc_inspection(iid)
    uid = session.get("user_id")
    if uid and ins:
        add_activity_log(uid, "delete", "qc_inspection", iid,
            f"删除了质检记录 - 订单 #{ins.get('order_id','')}")
    return jsonify({"ok": True})


# ==================== 报表中心 API ====================
@app.route("/api/reports/sales")
@login_required
def api_reports_sales():
    """综合销售报表"""
    months = request.args.get("months", 12, type=int)
    sales_detail = get_monthly_sales_detail(months)
    customer_acq = get_customer_acquisition_stats(months)
    prod_stats = get_production_stats()
    commission = get_commission_stats()
    return jsonify({
        "sales_detail": sales_detail,
        "customer_acquisition": customer_acq,
        "production_stats": prod_stats,
        "commission_summary": commission.get("summary", {}),
        "sales_performance": commission.get("sales_summary", []),
    })


@app.route("/api/export/report-sales")
@login_required
def api_export_report_sales():
    """导出销售报表Excel"""
    import openpyxl, io
    from openpyxl.styles import Font, PatternFill, Alignment

    months = request.args.get("months", 12, type=int)
    sales_detail = get_monthly_sales_detail(months)
    commission = get_commission_stats()

    wb = openpyxl.Workbook()

    # Sheet 1: 月度明细
    ws1 = wb.active
    ws1.title = "月度销售明细"
    ws1.append(["月份", "订单数", "营收(USD)", "成本(USD)", "利润(USD)"])
    for r in sales_detail:
        ws1.append([r["month"], r["order_count"], r["revenue"], r["cost"], r["profit"]])

    # Sheet 2: 销售排行
    ws2 = wb.create_sheet("销售排行")
    ws2.append(["销售员", "订单数", "营收(USD)", "成本(USD)", "利润(USD)", "提成率", "提成金额"])
    for r in (commission.get("sales_summary") or []):
        ws2.append([
            r.get("sales_name", ""), r.get("order_count", 0),
            r.get("total_revenue", 0), r.get("total_cost", 0),
            r.get("total_profit", 0), f"{r.get('commission_rate', 0)}%",
            r.get("total_commission", 0)
        ])

    output = io.BytesIO()
    wb.save(output)
    return _send_excel(output, "销售报表.xlsx")


@app.route("/api/quotes/<int:qid>/convert", methods=["POST"])
@login_required
def api_convert_quote_to_order(qid):
    quote = get_quote(qid)
    if not quote:
        return jsonify({"error": "报价不存在"}), 404
    if quote.get("status") != "accepted":
        return jsonify({"error": "仅已接受的报价可以转订单"}), 400
    data = {
        "customer_id": quote["customer_id"],
        "quote_id": qid,
        "items": quote.get("items", []),
        "total_amount": quote.get("total_amount", 0),
        "currency": quote.get("currency", "USD"),
        "status": "pending_approval",
        "created_by": session.get("user_id"),
        "order_no": quote.get("quote_no", "").replace("BH-", "ORD-"),
        "notes": f"源自报价 {quote.get('quote_no','')}",
    }
    result = add_order(data)
    if "error" in result:
        return jsonify(result), 400
    uid = session.get("user_id")
    if uid and result.get("id"):
        add_activity_log(uid, "convert", "order", result["id"],
            f"报价 {quote.get('quote_no','')} 转为订单 {result.get('order_no','')}")
    return jsonify({"ok": True, "id": result["id"], "order_no": result["order_no"]})


@app.route("/api/payment-dashboard")
@login_required
def api_payment_dashboard():
    return jsonify(get_payment_dashboard())


# ==================== 应收账款 API ====================
@app.route("/api/ar/summary")
@login_required
@role_required('admin', 'sales')
def api_ar_summary():
    return jsonify(get_ar_summary())


@app.route("/api/ar/by-customer")
@login_required
@role_required('admin', 'sales')
def api_ar_by_customer():
    return jsonify(get_ar_by_customer())


@app.route("/api/ar/payment-history")
@login_required
@role_required('admin', 'sales')
def api_ar_payment_history():
    limit = request.args.get("limit", 50, type=int)
    return jsonify(get_payment_history(limit))


@app.route("/api/ar/aging")
@login_required
@role_required('admin', 'sales')
def api_ar_aging():
    return jsonify(get_aging_analysis())


@app.route("/api/ar/migrate", methods=["POST"])
@login_required
@role_required('admin', 'sales')
def api_ar_migrate():
    count = migrate_payments_from_orders()
    return jsonify({"ok": True, "migrated": count})

# ==================== 潜客管理 API (Leads) ====================
@app.route("/api/leads")
@login_required
@role_required('admin', 'sales')
def api_leads():
    source = request.args.get("source")
    status = request.args.get("status")
    assigned_to = request.args.get("assigned_to", type=int)
    country = request.args.get("country")
    pool = request.args.get("pool", type=int)
    if pool:
        return jsonify(get_leads_with_pool_info(source=source, status=status, assigned_to=assigned_to, country=country))
    return jsonify(get_leads(source=source, status=status, assigned_to=assigned_to, country=country))


@app.route("/api/leads/summary")
@login_required
@role_required('admin', 'sales')
def api_leads_summary():
    return jsonify(get_lead_summary())


@app.route("/api/leads/funnel")
@login_required
@role_required('admin', 'sales')
def api_leads_funnel():
    return jsonify(get_lead_funnel())


@app.route("/api/leads/<int:cid>/assign", methods=["PUT"])
@login_required
@role_required('admin', 'sales')
def api_leads_assign(cid):
    data = request.json or {}
    uid = data.get("user_id")
    if uid is None:
        return jsonify({"error": "缺少 user_id"}), 400
    assign_lead(cid, uid)
    logged_uid = session.get("user_id")
    customer = get_customer(cid)
    cname = customer["name"] if customer else f"ID:{cid}"
    if logged_uid:
        target_user = get_user_by_id(uid)
        tname = target_user["display_name"] if target_user else str(uid)
        add_activity_log(logged_uid, "assign", "lead", cid,
            f"分配潜客 {cname} 给 {tname}")
    # 通知被分配者
    if uid:
        if customer:
            add_notification(int(uid), 'assign',
                f'新潜客分配: {customer["name"]}',
                f'来源: {customer.get("source","")} | {customer.get("country","")}',
                link=f'/leads/{cid}',
                related_type='lead', related_id=cid)
    return jsonify({"ok": True})


@app.route("/api/leads/<int:cid>/status", methods=["PUT"])
@login_required
@role_required('admin', 'sales')
def api_leads_status(cid):
    data = request.json or {}
    status = data.get("lead_status")
    if not status:
        return jsonify({"error": "缺少 lead_status"}), 400
    customer = get_customer(cid)
    update_lead_status(cid, status)
    uid = session.get("user_id")
    if uid and customer:
        add_activity_log(uid, "status_change", "lead", cid,
            f"修改潜客 {customer.get('name','')} 阶段为 {status}")
    return jsonify({"ok": True})


@app.route("/api/leads/<int:cid>/source", methods=["PUT"])
@login_required
@role_required('admin', 'sales')
def api_leads_source(cid):
    data = request.json or {}
    source = data.get("source", "")
    campaign = data.get("campaign", "")
    update_lead_source(cid, source, campaign)
    uid = session.get("user_id")
    customer = get_customer(cid)
    if uid and customer:
        add_activity_log(uid, "update", "lead", cid,
            f"修改潜客 {customer.get('name','')} 来源为 {source}")
    return jsonify({"ok": True})


# ==================== 公海自动流转 API ====================

@app.route("/api/users")
@login_required
@role_required('admin', 'sales')
def api_users():
    """获取所有销售用户（供分配弹窗和过滤下拉使用）"""
    return jsonify(get_users())


@app.route("/api/leads/<int:cid>/unassign", methods=["PUT"])
@login_required
@role_required('admin', 'sales')
def api_leads_unassign(cid):
    """退回公海"""
    customer = get_customer(cid)
    unassign_lead(cid)
    uid = session.get("user_id")
    if uid and customer:
        add_activity_log(uid, "unassign", "lead", cid,
            f"将潜客 {customer.get('name','')} 退回公海")
    return jsonify({"ok": True})


@app.route("/api/leads/<int:cid>/claim", methods=["PUT"])
@login_required
@role_required('admin', 'sales')
def api_leads_claim(cid):
    """认领潜客"""
    uid = session.get("user_id")
    if not uid:
        return jsonify({"error": "未登录"}), 401
    ok = claim_lead(cid, uid)
    if not ok:
        return jsonify({"error": "该潜客已被他人认领"}), 409
    customer = get_customer(cid)
    if uid and customer:
        add_activity_log(uid, "claim", "lead", cid,
            f"认领了潜客 {customer.get('name','')}")
    return jsonify({"ok": True})


@app.route("/api/leads/reclaim", methods=["POST"])
@login_required
@role_required('admin', 'sales')
def api_leads_reclaim():
    """手动触发回收检查"""
    reclaimed = reclaim_expired_leads(days=14)
    return jsonify({"ok": True, "count": len(reclaimed), "items": reclaimed})


# ==================== 跟进提醒 API ====================

@app.route("/api/leads/followup")
@login_required
@role_required('admin', 'sales')
def api_leads_followup():
    """获取需要跟进的潜客列表"""
    return jsonify(get_leads_due_followup())


@app.route("/api/leads/followup/summary")
@login_required
@role_required('admin', 'sales')
def api_leads_followup_summary():
    """获取跟进汇总数据"""
    return jsonify(get_today_followup_summary())


@app.route("/api/leads/<int:cid>/contacted", methods=["POST"])
@login_required
@role_required('admin', 'sales')
def api_leads_contacted(cid):
    """标记潜客为已联系（更新 last_contacted_at）"""
    update_last_contacted(cid)
    return jsonify({"ok": True})


@app.route("/api/leads/<int:cid>/followup-msg", methods=["POST"])
@login_required
@role_required('admin', 'sales')
def api_leads_followup_msg(cid):
    """AI生成跟进消息"""
    from database import get_customer
    customer = get_customer(cid)
    if not customer:
        return jsonify({"error": "客户不存在"}), 404
    data = request.json or {}
    followup_type = data.get("followup_type", "3day")
    result = get_ai_followup_message(
        customer_name=customer["name"],
        language=customer.get("language", "English"),
        followup_type=followup_type,
        sales_name="Philip"
    )
    return jsonify({"ok": True, "followup_type": followup_type, "message": result})


# ==================== V4 — 自动销售网络 API ====================

@app.route("/api/v4/dashboard/kpi")
@login_required
@role_required('admin', 'sales')
def api_v4_dashboard_kpi():
    """V4 仪表盘KPI"""
    from dashboard_engine import DashboardEngine
    return jsonify(DashboardEngine().get_kpi_summary())


@app.route("/api/v4/dashboard/funnel")
@login_required
@role_required('admin', 'sales')
def api_v4_dashboard_funnel():
    """V4 转化漏斗"""
    from dashboard_engine import DashboardEngine
    return jsonify(DashboardEngine().get_conversion_funnel())


@app.route("/api/v4/dashboard/campaigns")
@login_required
@role_required('admin', 'sales')
def api_v4_dashboard_campaigns():
    """V4 活动表现"""
    from dashboard_engine import DashboardEngine
    return jsonify(DashboardEngine().get_campaign_performance())


@app.route("/api/v4/dashboard/geo")
@login_required
@role_required('admin', 'sales')
def api_v4_dashboard_geo():
    """V4 国家分布"""
    from dashboard_engine import DashboardEngine
    return jsonify(DashboardEngine().get_geo_distribution())


@app.route("/api/v4/dashboard/industry")
@login_required
@role_required('admin', 'sales')
def api_v4_dashboard_industry():
    """V4 行业分布"""
    from dashboard_engine import DashboardEngine
    return jsonify(DashboardEngine().get_industry_distribution())


@app.route("/api/v4/dashboard/top-leads")
@login_required
@role_required('admin', 'sales')
def api_v4_dashboard_top_leads():
    """V4 高分线索"""
    limit = request.args.get("limit", 10, type=int)
    from dashboard_engine import DashboardEngine
    return jsonify(DashboardEngine().get_top_leads(limit=limit))


@app.route("/api/v4/dashboard/revenue-trend")
@login_required
@role_required('admin', 'sales')
def api_v4_dashboard_revenue_trend():
    """V4 收入趋势"""
    days = request.args.get("days", 30, type=int)
    from dashboard_engine import DashboardEngine
    return jsonify(DashboardEngine().get_revenue_trend(days=days))


@app.route("/api/v4/campaigns")
@login_required
@role_required('admin', 'sales')
def api_v4_campaigns_list():
    """V4 活动列表"""
    status = request.args.get("status")
    from campaign_engine import CampaignEngine
    return jsonify(CampaignEngine().list_campaigns(status=status))


@app.route("/api/v4/campaigns", methods=["POST"])
@login_required
@role_required('admin', 'sales')
def api_v4_campaigns_create():
    """V4 创建活动"""
    data = request.json or {}
    from campaign_engine import Campaign, CampaignEngine
    c = Campaign(
        name=data.get("name", ""),
        target_countries=data.get("target_countries"),
        target_industries=data.get("target_industries"),
        message_template=data.get("message_template", ""),
        max_outreach_per_day=data.get("max_outreach_per_day", 10),
    )
    cid = CampaignEngine().create_campaign(c)
    return jsonify({"ok": True, "campaign_id": cid})


@app.route("/api/v4/campaigns/<int:cid>")
@login_required
@role_required('admin', 'sales')
def api_v4_campaigns_get(cid):
    """V4 活动详情"""
    from campaign_engine import CampaignEngine
    camp = CampaignEngine().get_campaign(cid)
    if not camp:
        return jsonify({"error": "活动不存在"}), 404
    return jsonify(camp)


@app.route("/api/v4/campaigns/<int:cid>/status", methods=["PUT"])
@login_required
@role_required('admin', 'sales')
def api_v4_campaigns_status(cid):
    """V4 更新活动状态"""
    data = request.json or {}
    status = data.get("status", "")
    if status not in ("draft", "active", "paused", "completed"):
        return jsonify({"error": "无效状态"}), 400
    from campaign_engine import CampaignEngine
    CampaignEngine().update_status(cid, status)
    return jsonify({"ok": True})


@app.route("/api/v4/campaigns/<int:cid>/assign", methods=["POST"])
@login_required
@role_required('admin', 'sales')
def api_v4_campaigns_assign(cid):
    """V4 分配线索到活动"""
    data = request.json or {}
    customer_ids = data.get("customer_ids", [])
    if not customer_ids:
        return jsonify({"error": "缺少 customer_ids"}), 400
    from campaign_engine import CampaignEngine
    CampaignEngine().assign_leads(cid, customer_ids)
    return jsonify({"ok": True, "assigned": len(customer_ids)})


@app.route("/api/v4/campaigns/<int:cid>/leads")
@login_required
@role_required('admin', 'sales')
def api_v4_campaigns_leads(cid):
    """V4 活动下线索列表"""
    from campaign_engine import CampaignEngine
    return jsonify(CampaignEngine().get_campaign_leads(cid))


@app.route("/api/v4/campaigns/<int:cid>/reply", methods=["POST"])
@login_required
@role_required('admin', 'sales')
def api_v4_campaigns_reply(cid):
    """V4 记录回复"""
    from campaign_engine import CampaignEngine
    CampaignEngine().record_reply(cid)
    return jsonify({"ok": True})


@app.route("/api/v4/campaigns/<int:cid>/convert", methods=["POST"])
@login_required
@role_required('admin', 'sales')
def api_v4_campaigns_convert(cid):
    """V4 记录转化"""
    data = request.json or {}
    revenue = data.get("revenue", 0)
    from campaign_engine import CampaignEngine
    CampaignEngine().record_conversion(cid, revenue)
    return jsonify({"ok": True})


@app.route("/api/v4/leads/import", methods=["POST"])
@login_required
@role_required('admin', 'sales')
def api_v4_leads_import():
    """V4 CSV导入线索 + 自动评分"""
    import tempfile
    file = request.files.get("file")
    if not file:
        return jsonify({"error": "请上传CSV文件"}), 400
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".csv")
    file.save(tmp.name)
    from lead_router import LeadRouter
    result = LeadRouter().import_csv(tmp.name)
    os.unlink(tmp.name)
    status_code = 200 if result["imported"] > 0 else 400
    return jsonify(result), status_code


@app.route("/api/v4/leads/score/<int:cid>")
@login_required
@role_required('admin', 'sales')
def api_v4_leads_score(cid):
    """V4 评分单个线索"""
    from lead_router import LeadRouter
    from database import get_customer
    cust = get_customer(cid)
    if not cust:
        return jsonify({"error": "客户不存在"}), 404
    result = LeadRouter().score_lead(cid, cust)
    route = LeadRouter().route_lead(cid, cust)
    return jsonify({"score_result": result, "route": route})


@app.route("/api/v4/leads/score-batch", methods=["POST"])
@login_required
@role_required('admin', 'sales')
def api_v4_leads_score_batch():
    """V4 批量评分"""
    data = request.json or {}
    ids = data.get("customer_ids", [])
    if not ids:
        return jsonify({"error": "缺少 customer_ids"}), 400
    from lead_router import LeadRouter
    results = LeadRouter().score_batch(ids)
    return jsonify(results)


@app.route("/api/v4/outreach/process", methods=["POST"])
@login_required
@role_required('admin', 'sales')
def api_v4_outreach_process():
    """V4 处理待触达线索"""
    data = request.json or {}
    limit = data.get("limit", 10)
    from outreach_engine import OutreachEngine
    result = OutreachEngine().process_new_leads(limit=limit)
    return jsonify(result)


@app.route("/api/v4/outreach/followups")
@login_required
@role_required('admin', 'sales')
def api_v4_outreach_followups():
    """V4 获取待跟进线索"""
    limit = request.args.get("limit", 20, type=int)
    from outreach_engine import OutreachEngine
    return jsonify(OutreachEngine().get_due_followups(limit=limit))


@app.route("/api/v4/outreach/schedule", methods=["POST"])
@login_required
@role_required('admin', 'sales')
def api_v4_outreach_schedule():
    """V4 执行跟进调度"""
    data = request.json or {}
    limit = data.get("limit", 20)
    from outreach_engine import OutreachEngine
    result = OutreachEngine().schedule_followups(limit=limit)
    return jsonify(result)


@app.route("/api/v4/leads/export-template")
@login_required
@role_required('admin', 'sales')
def api_v4_leads_export_template():
    """V4 下载CSV导入模板"""
    import csv, io as csv_io
    output = csv_io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["name", "company", "whatsapp", "country", "industry", "source", "campaign", "notes"])
    writer.writerow(["Example Cafe", "Sunny Cafe", "+1234567890", "US", "restaurant", "import", "Q3 US Campaign", "needs LED storefront sign"])
    output.seek(0)
    from flask import Response
    return Response(
        output.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment;filename=lead_import_template.csv"}
    )


# ==================== V4 — Autonomous Sales System API ====================

# ---- V4: 客户优先级 (Deal Prioritizer) ----

@app.route("/api/v4/priorities", methods=["GET"])
@login_required
@role_required('admin', 'sales')
def api_v4_priorities():
    """获取所有活跃客户的优先级评分列表"""
    from ai_engine.deal_prioritizer import DealPrioritizer
    prioritizer = DealPrioritizer()
    results = prioritizer.batch_reprioritize(limit=200)
    return jsonify(results)


@app.route("/api/v4/prioritize", methods=["POST"])
@login_required
@role_required('admin')
def api_v4_prioritize():
    """手动触发重新评分"""
    from ai_engine.deal_prioritizer import DealPrioritizer
    result = DealPrioritizer().batch_reprioritize(limit=200)
    return jsonify({"ok": True, "count": len(result)})


@app.route("/api/v4/priorities/summary", methods=["GET"])
@login_required
@role_required('admin', 'sales')
def api_v4_priorities_summary():
    """获取 A/B/C 三类客户数量汇总"""
    from ai_engine.deal_prioritizer import DealPrioritizer
    summary = DealPrioritizer().get_priority_summary()
    return jsonify(summary)


# ---- V4: 自动发送 (Autonomous Sender) ----

@app.route("/api/v4/send/<int:customer_id>", methods=["POST"])
@login_required
@role_required('admin', 'sales')
def api_v4_send(customer_id):
    """手动触发自动发送

    Body: {"type": "followup_light|push_close|risk_message|wake_up|final_close|quote"}
    """
    data = request.json or {}
    msg_type = data.get("type", "followup_light")
    from ai_engine.autonomous_sender import AutonomousSender
    sender = AutonomousSender()

    if msg_type == "quote":
        result = sender.send_quote(customer_id, data.get("price_tier", "MID"))
    else:
        result = sender.send_followup(customer_id, msg_type)

    return jsonify(result)


# ---- V4: 动态定价 (Dynamic Pricing) ----

@app.route("/api/v4/pricing/<int:customer_id>", methods=["GET"])
@login_required
@role_required('admin', 'sales')
def api_v4_get_pricing(customer_id):
    """查看客户当前定价"""
    from ai_engine.dynamic_pricing import DynamicPricing
    result = DynamicPricing().get_pricing_for_customer(customer_id)
    return jsonify(result)


@app.route("/api/v4/pricing/<int:customer_id>", methods=["POST"])
@login_required
@role_required('admin')
def api_v4_override_pricing(customer_id):
    """手动覆盖客户定价档位

    Body: {"price_tier": "LOW|MID|HIGH", "reason": "..."}
    """
    data = request.json or {}
    tier = data.get("price_tier", "MID")
    reason = data.get("reason", "manual override")
    from ai_engine.dynamic_pricing import DynamicPricing
    result = DynamicPricing().apply_override(customer_id, tier, reason)
    return jsonify(result)


# ---- V4: 决策大脑 (Conversion AI Brain) ----

@app.route("/api/v4/decide/<int:customer_id>", methods=["GET"])
@login_required
@role_required('admin', 'sales')
def api_v4_decide(customer_id):
    """查看决策大脑对该客户的决策结果（dry run）"""
    from ai_engine.conversion_ai_brain import ConversionBrain
    result = ConversionBrain().decide(customer_id)
    return jsonify(result)


# ---- V4: 收单调度器 (Revenue Scheduler) ----

@app.route("/api/v4/scheduler/status", methods=["GET"])
@login_required
@role_required('admin')
def api_v4_scheduler_status():
    """查看调度器状态和今日日志"""
    from ai_engine.revenue_scheduler import RevenueScheduler
    result = RevenueScheduler().get_status()
    return jsonify(result)


@app.route("/api/v4/scheduler/run", methods=["POST"])
@login_required
@role_required('admin')
def api_v4_scheduler_run():
    """手动触发调度 tick

    Body: {"slot": "09:00|12:00|18:00|22:00"} (可选)
    """
    data = request.json or {}
    slot = data.get("slot")
    from ai_engine.revenue_scheduler import RevenueScheduler
    result = RevenueScheduler().manual_run(slot)
    return jsonify(result)


@app.route("/api/v4/scheduler/log", methods=["GET"])
@login_required
@role_required('admin')
def api_v4_scheduler_log():
    """获取调度器历史日志"""
    limit = request.args.get("limit", 50, type=int)
    from ai_engine.revenue_scheduler import RevenueScheduler
    result = RevenueScheduler().get_today_log()
    return jsonify(result[:limit])

# ---- V5: 区域引擎 ----

@app.route("/api/v5/regions")
@login_required
@role_required('admin', 'sales')
def api_v5_regions():
    """获取所有区域"""
    from region_engine import RegionEngine
    return jsonify(RegionEngine().get_all_regions())


@app.route("/api/v5/regions/<country_code>")
@login_required
@role_required('admin', 'sales')
def api_v5_region_for_country(country_code):
    """获取国家所属区域"""
    from region_engine import RegionEngine
    return jsonify(RegionEngine().get_region_for_country(country_code))


@app.route("/api/v5/exchange-rates")
@login_required
@role_required('admin', 'sales')
def api_v5_exchange_rates():
    """获取汇率列表"""
    from_currency = request.args.get("from", "EUR")
    to_currency = request.args.get("to", "USD")
    from region_engine import RegionEngine
    rate = RegionEngine().get_rate(from_currency, to_currency)
    history = RegionEngine().get_exchange_rate_history(from_currency, to_currency)
    return jsonify({"rate": rate, "history": history})


@app.route("/api/v5/exchange-rates", methods=["POST"])
@login_required
@role_required('admin')
def api_v5_update_exchange_rate():
    """更新汇率"""
    data = request.json
    if not data or "from_currency" not in data or "rate" not in data:
        return jsonify({"error": "缺少 from_currency 或 rate"}), 400
    from region_engine import RegionEngine
    ok = RegionEngine().update_rate(
        data["from_currency"], data["rate"],
        data.get("to_currency", "USD"),
        data.get("date"),
        data.get("source", "manual"),
    )
    return jsonify({"ok": ok})


@app.route("/api/v5/currency-convert", methods=["POST"])
@login_required
@role_required('admin', 'sales')
def api_v5_currency_convert():
    """币种转换"""
    data = request.json
    if not data or "amount" not in data or "from" not in data:
        return jsonify({"error": "缺少 amount 或 from"}), 400
    from region_engine import RegionEngine
    result = RegionEngine().convert(
        data["amount"], data["from"],
        data.get("to", "USD"),
        data.get("date"),
    )
    return jsonify(result)


@app.route("/api/v5/seed-regions", methods=["POST"])
@login_required
@role_required('admin')
def api_v5_seed_regions():
    """种子区域数据"""
    from region_engine import RegionEngine
    re = RegionEngine()
    regions = re.seed_default_regions()
    rates = re.seed_default_rates()
    return jsonify({"regions_seeded": regions, "rates_seeded": rates})

# ---- V5: 市场定价 ----

@app.route("/api/v5/market-pricing")
@login_required
@role_required('admin', 'sales')
def api_v5_market_pricing():
    """获取市场定价策略"""
    from region_engine import RegionEngine
    region_id = request.args.get("region_id", type=int)
    return jsonify(RegionEngine().get_market_margin_targets(region_id))


@app.route("/api/v5/market-pricing/<int:region_id>/<product_category>")
@login_required
@role_required('admin', 'sales')
def api_v5_pricing_coefficient(region_id, product_category):
    """获取特定区域+产品的定价系数"""
    from region_engine import RegionEngine
    return jsonify(RegionEngine().get_pricing_coefficient(region_id, product_category))


# ---- V5: 全球线索路由 ----

@app.route("/api/v5/lead-router/score/<int:customer_id>")
@login_required
@role_required('admin', 'sales')
def api_v5_lead_score(customer_id):
    """V5 全球评分"""
    from global_lead_router import GlobalLeadRouter
    return jsonify(GlobalLeadRouter().score_lead_global(customer_id))


@app.route("/api/v5/lead-router/route/<int:customer_id>")
@login_required
@role_required('admin', 'sales')
def api_v5_lead_route(customer_id):
    """V5 全球路由"""
    from global_lead_router import GlobalLeadRouter
    return jsonify(GlobalLeadRouter().route_lead_global(customer_id))


@app.route("/api/v5/lead-router/export")
@login_required
@role_required('admin', 'sales')
def api_v5_lead_export():
    """V5 导出全量线索分析"""
    from global_lead_router import GlobalLeadRouter
    limit = request.args.get("limit", 50, type=int)
    return jsonify(GlobalLeadRouter().export_lead_analysis(limit))


# ---- V5: 多Agent团队 ----

@app.route("/api/v5/agents")
@login_required
@role_required('admin', 'sales')
def api_v5_agents():
    """获取所有Agent"""
    from multi_agent_team import MultiAgentTeam
    return jsonify(MultiAgentTeam().get_all_agents())


@app.route("/api/v5/agents/select")
@login_required
@role_required('admin', 'sales')
def api_v5_agent_select():
    """选择适合国家的Agent"""
    country = request.args.get("country", "").upper()
    language = request.args.get("language", "English")
    from multi_agent_team import MultiAgentTeam
    return jsonify(MultiAgentTeam().select_agent(country, language))


@app.route("/api/v5/agents/stats")
@login_required
@role_required('admin', 'sales')
def api_v5_agent_stats():
    """Agent触达统计"""
    from multi_agent_team import MultiAgentTeam
    days = request.args.get("days", 30, type=int)
    return jsonify(MultiAgentTeam().get_agent_stats(days))


# ---- V5: Agent Competition System ----

@app.route("/api/v5/competition/run", methods=["POST"])
@login_required
@role_required('admin', 'sales')
def api_v5_competition_run():
    """运行Agent竞争（测试用）"""
    data = request.json or {}
    msg = data.get("message", "Hello, I'm interested in your signs.")
    state = data.get("state", "NEW")
    priority = data.get("priority", "C")
    country = data.get("country", "")
    from ai_engine.agents.agent_competition import AgentCompetition
    result = AgentCompetition().run_competition(
        customer_msg=msg,
        context={"state": state, "priority": priority, "country": country},
    )
    return jsonify(result)


@app.route("/api/v5/competition/agents")
@login_required
@role_required('admin', 'sales')
def api_v5_competition_agents():
    """获取所有Agent类型"""
    from ai_engine.agents.agent_competition import AgentCompetition
    comp = AgentCompetition()
    agents = []
    for aid in comp.get_agent_ids():
        agent_class = comp.agents.get(aid)
        if agent_class:
            agents.append({
                "agent_id": aid,
                "name": getattr(agent_class, "NAME", aid),
                "strategy": getattr(agent_class, "STRATEGY", ""),
                "pricing_mode": getattr(agent_class, "PRICING_MODE", ""),
            })
    return jsonify(agents)


@app.route("/api/v5/competition/weights")
@login_required
@role_required('admin', 'sales')
def api_v5_competition_weights():
    """获取Agent权重/学习数据"""
    scene_state = request.args.get("state", "")
    scene_priority = request.args.get("priority", "")
    from ai_engine.agents.winner_selector import WinnerSelector
    weights = WinnerSelector().get_agent_weights(scene_state, scene_priority)
    return jsonify(weights)


@app.route("/api/v5/competition/evolution")
@login_required
@role_required('admin', 'sales')
def api_v5_competition_evolution():
    """获取Agent进化报告"""
    from ai_engine.agents.winner_selector import WinnerSelector
    report = WinnerSelector().get_agent_evolution_report()
    return jsonify(report)


@app.route("/api/v5/competition/schedule")
@login_required
@role_required('admin', 'sales')
def api_v5_competition_schedule():
    """获取Agent调度计划"""
    date_str = request.args.get("date", "")
    from ai_engine.agents.agent_manager import AgentManager
    schedule = AgentManager().get_daily_schedule(date_str or None)
    return jsonify(schedule)


@app.route("/api/v5/competition/route", methods=["POST"])
@login_required
@role_required('admin', 'sales')
def api_v5_competition_route():
    """Agent路由：客户特征→最佳Agent"""
    data = request.json or {}
    from ai_engine.agents.agent_router import AgentRouter
    result = AgentRouter.route(data)
    return jsonify(result)


@app.route("/api/v5/competition/load")
@login_required
@role_required('admin', 'sales')
def api_v5_competition_load():
    """获取Agent负载"""
    from ai_engine.agents.agent_manager import AgentManager
    load = AgentManager().get_agent_load()
    return jsonify(load)


# =================== V6 — Global Revenue Network ===================

@app.route("/api/v6/route", methods=["POST"])
@login_required
@role_required('admin', 'sales')
def api_v6_route():
    """Global Router: 国家→区域路由"""
    data = request.json or {}
    from ai_engine.global_router import GlobalRouter
    return jsonify(GlobalRouter.route(
        data.get("country", ""),
        data.get("intent", ""),
        data.get("product", ""),
        data.get("urgency", "medium"),
    ))


@app.route("/api/v6/regions")
@login_required
@role_required('admin', 'sales')
def api_v6_regions():
    """获取所有区域配置"""
    from ai_engine.global_router import GlobalRouter
    return jsonify(GlobalRouter.get_all_regions())


@app.route("/api/v6/culture/<country>")
@login_required
@role_required('admin', 'sales')
def api_v6_culture(country):
    """Culture Adaptor: 获取国家文化上下文"""
    from ai_engine.culture_adaptor import CultureAdaptor
    return jsonify(CultureAdaptor.get_culture_context(country))


@app.route("/api/v6/adapt-message", methods=["POST"])
@login_required
@role_required('admin', 'sales')
def api_v6_adapt_message():
    """Culture Adaptor: 适配消息到目标文化"""
    data = request.json or {}
    from ai_engine.culture_adaptor import CultureAdaptor
    adapted = CultureAdaptor.adapt(
        data.get("text", ""),
        data.get("country", ""),
        data.get("product_type", ""),
    )
    return jsonify({"adapted": adapted})


@app.route("/api/v6/strategy/<region>")
@login_required
@role_required('admin', 'sales')
def api_v6_strategy(region):
    """Regional Sales Brain: 获取区域销售策略"""
    from ai_engine.regional_sales_brain import RegionalSalesBrain
    return jsonify(RegionalSalesBrain.get_strategy(region.upper()))


@app.route("/api/v6/strategies")
@login_required
@role_required('admin', 'sales')
def api_v6_strategies():
    """获取所有区域策略"""
    from ai_engine.regional_sales_brain import RegionalSalesBrain
    return jsonify(RegionalSalesBrain.get_all_strategies())


@app.route("/api/v6/profit/calculate", methods=["POST"])
@login_required
@role_required('admin', 'sales')
def api_v6_profit_calculate():
    """Profit Engine: 计算价格"""
    data = request.json or {}
    from ai_engine.profit_engine import ProfitEngine
    return jsonify(ProfitEngine.calculate_price(
        float(data.get("base_cost", 0)),
        data.get("country", "US"),
        data.get("product_category", "general"),
        int(data.get("quantity", 1)),
    ))


@app.route("/api/v6/profit/countries")
@login_required
@role_required('admin', 'sales')
def api_v6_profit_countries():
    """获取所有国家利润配置"""
    from ai_engine.profit_engine import ProfitEngine
    return jsonify(ProfitEngine.get_all_country_profiles())


@app.route("/api/v6/currency/convert", methods=["POST"])
@login_required
@role_required('admin', 'sales')
def api_v6_currency_convert():
    """Currency Optimizer: 汇率换算"""
    data = request.json or {}
    from ai_engine.currency_optimizer import CurrencyOptimizer
    return jsonify(CurrencyOptimizer.convert(
        float(data.get("amount_usd", 0)),
        data.get("to_currency", "USD"),
    ))


@app.route("/api/v6/currency/localize", methods=["POST"])
@login_required
@role_required('admin', 'sales')
def api_v6_currency_localize():
    """Currency Optimizer: 本地化定价"""
    data = request.json or {}
    from ai_engine.currency_optimizer import CurrencyOptimizer
    return jsonify(CurrencyOptimizer.get_localized_price(
        float(data.get("amount_usd", 0)),
        data.get("country", "US"),
    ))


@app.route("/api/v6/currency/all-prices", methods=["POST"])
@login_required
@role_required('admin', 'sales')
def api_v6_currency_all_prices():
    """获取所有货币价格"""
    data = request.json or {}
    from ai_engine.currency_optimizer import CurrencyOptimizer
    return jsonify(CurrencyOptimizer.get_all_prices(
        float(data.get("amount_usd", 0)),
    ))


@app.route("/api/v6/production/allocate", methods=["POST"])
@login_required
@role_required('admin', 'sales')
def api_v6_production_allocate():
    """Production Allocator: 排产分配"""
    data = request.json or {}
    from ai_engine.production_allocator import ProductionAllocator
    return jsonify(ProductionAllocator.allocate(data.get("orders", [])))


@app.route("/api/v6/production/factories")
@login_required
@role_required('admin', 'sales')
def api_v6_production_factories():
    """获取工厂配置"""
    from ai_engine.production_allocator import ProductionAllocator
    return jsonify(ProductionAllocator.get_factory_utilization())


@app.route("/api/v6/production/shipping", methods=["POST"])
@login_required
@role_required('admin', 'sales')
def api_v6_production_shipping():
    """估算运费"""
    data = request.json or {}
    from ai_engine.production_allocator import ProductionAllocator
    return jsonify(ProductionAllocator.estimate_shipping(
        data.get("country", ""),
        data.get("factory_id", "f1"),
        int(data.get("quantity", 1)),
        float(data.get("weight_kg", 10)),
    ))


# ---- V5: 利润引擎 ----

@app.route("/api/v5/margin/cost", methods=["POST"])
@login_required
@role_required('admin', 'sales')
def api_v5_margin_cost():
    """计算完整成本"""
    data = request.json or {}
    from margin_engine import MarginEngine
    return jsonify(MarginEngine().calculate_full_cost(
        data.get("factory_id"),
        data.get("product_category", "general"),
        data.get("quantity", 1),
    ))


@app.route("/api/v5/margin/evaluate", methods=["POST"])
@login_required
@role_required('admin', 'sales')
def api_v5_margin_evaluate():
    """评估利润率"""
    data = request.json or {}
    if "selling_price" not in data:
        return jsonify({"error": "缺少 selling_price"}), 400
    from margin_engine import MarginEngine
    return jsonify(MarginEngine().evaluate_margin(
        data.get("factory_id"),
        data.get("product_category", "general"),
        data["selling_price"],
        data.get("currency", "USD"),
        data.get("customer_country", ""),
        data.get("quantity", 1),
    ))


@app.route("/api/v5/margin/optimize", methods=["POST"])
@login_required
@role_required('admin', 'sales')
def api_v5_margin_optimize():
    """最优定价建议"""
    data = request.json or {}
    if "customer_country" not in data or "product_category" not in data:
        return jsonify({"error": "缺少 customer_country 或 product_category"}), 400
    from margin_engine import MarginEngine
    return jsonify(MarginEngine().optimize_price(
        data.get("factory_id"),
        data["product_category"],
        data["customer_country"],
        data.get("target_margin"),
        data.get("quantity", 1),
    ))


@app.route("/api/v5/margin/summary")
@login_required
@role_required('admin', 'sales')
def api_v5_margin_summary():
    """利润汇总"""
    from margin_engine import MarginEngine
    region = request.args.get("region")
    days = request.args.get("days", 30, type=int)
    return jsonify(MarginEngine().get_profit_summary(region, days))


# ---- V5: 工厂分配 ----

@app.route("/api/v5/factories")
@login_required
@role_required('admin', 'sales')
def api_v5_factories():
    """获取所有工厂"""
    from factory_allocator import FactoryAllocator
    return jsonify(FactoryAllocator().get_all_factories())


@app.route("/api/v5/factories/utilization")
@login_required
@role_required('admin', 'sales')
def api_v5_factory_utilization():
    """工厂产能利用率"""
    from factory_allocator import FactoryAllocator
    return jsonify(FactoryAllocator().get_factory_utilization())


@app.route("/api/v5/factories/best", methods=["POST"])
@login_required
@role_required('admin', 'sales')
def api_v5_factory_best():
    """寻找最优工厂"""
    data = request.json or {}
    if "product_category" not in data or "destination_country" not in data:
        return jsonify({"error": "缺少 product_category 或 destination_country"}), 400
    from factory_allocator import FactoryAllocator
    return jsonify(FactoryAllocator().find_best_factory(
        data["product_category"],
        data["destination_country"],
        data.get("preference", "cost"),
    ))


@app.route("/api/v5/factories/shipping", methods=["POST"])
@login_required
@role_required('admin', 'sales')
def api_v5_shipping_estimate():
    """运费估算"""
    data = request.json or {}
    if "destination_country" not in data:
        return jsonify({"error": "缺少 destination_country"}), 400
    from factory_allocator import FactoryAllocator
    return jsonify(FactoryAllocator().estimate_shipping(
        data["destination_country"],
        data.get("volume_m3", 0.1),
        data.get("mode", "sea"),
    ))


# ---- V5: 收入仪表盘 ----

@app.route("/api/v5/dashboard/global-kpi")
@login_required
@role_required('admin', 'sales')
def api_v5_global_kpi():
    """全球KPI汇总"""
    from revenue_dashboard_v5 import RevenueDashboardV5
    return jsonify(RevenueDashboardV5().get_global_kpi_summary())


@app.route("/api/v5/dashboard/revenue-by-region")
@login_required
@role_required('admin', 'sales')
def api_v5_revenue_by_region():
    """按区域收入"""
    from revenue_dashboard_v5 import RevenueDashboardV5
    days = request.args.get("days", 30, type=int)
    return jsonify(RevenueDashboardV5().get_revenue_by_region(days))


@app.route("/api/v5/dashboard/profit-by-region")
@login_required
@role_required('admin', 'sales')
def api_v5_profit_by_region():
    """按区域利润"""
    from revenue_dashboard_v5 import RevenueDashboardV5
    days = request.args.get("days", 30, type=int)
    return jsonify(RevenueDashboardV5().get_profit_margin_by_region(days))


@app.route("/api/v5/dashboard/lead-stats")
@login_required
@role_required('admin', 'sales')
def api_v5_lead_stats():
    """区域线索统计"""
    from revenue_dashboard_v5 import RevenueDashboardV5
    return jsonify(RevenueDashboardV5().get_region_lead_stats())


@app.route("/api/v5/dashboard/ranking")
@login_required
@role_required('admin', 'sales')
def api_v5_ranking():
    """全球绩效排名"""
    from revenue_dashboard_v5 import RevenueDashboardV5
    return jsonify(RevenueDashboardV5().get_global_performance_ranking())


@app.route("/api/v5/dashboard/multi-currency")
@login_required
@role_required('admin', 'sales')
def api_v5_multi_currency():
    """多币种收入"""
    from revenue_dashboard_v5 import RevenueDashboardV5
    days = request.args.get("days", 30, type=int)
    return jsonify(RevenueDashboardV5().get_multi_currency_revenue(days))


# ==================== V6 — Financial Intelligence OS API ====================

# ---- V6: P&L Engine ----

@app.route("/api/v6/pl/accounts")
@login_required
@role_required('admin', 'sales')
def api_v6_pl_accounts():
    """获取会计科目表"""
    from pl_engine import PLEngine
    return jsonify(PLEngine().get_pl_accounts() if hasattr(PLEngine(), 'get_pl_accounts') else {"accounts": PLEngine()._import_db().get_pl_accounts()})


@app.route("/api/v6/pl/seed-accounts", methods=["POST"])
@login_required
@role_required('admin')
def api_v6_pl_seed_accounts():
    """初始化会计科目"""
    from pl_engine import PLEngine
    count = PLEngine().seed_accounts()
    return jsonify({"seeded": count})


@app.route("/api/v6/pl/periods")
@login_required
@role_required('admin', 'sales')
def api_v6_pl_periods():
    """获取会计期间"""
    from pl_engine import PLEngine
    return jsonify(PLEngine().get_periods())


@app.route("/api/v6/pl/periods", methods=["POST"])
@login_required
@role_required('admin')
def api_v6_pl_create_period():
    """创建会计期间"""
    data = request.json
    if not data or "code" not in data:
        return jsonify({"error": "缺少 code"}), 400
    from pl_engine import PLEngine
    result = PLEngine().create_period(
        data["code"], data.get("type", "monthly"),
        data["start_date"], data["end_date"], data.get("notes", "")
    )
    return jsonify(result)


@app.route("/api/v6/pl/close-period", methods=["POST"])
@login_required
@role_required('admin')
def api_v6_pl_close_period():
    """关账"""
    data = request.json
    if not data or "period_id" not in data:
        return jsonify({"error": "缺少 period_id"}), 400
    from pl_engine import PLEngine
    result = PLEngine().close_period(data["period_id"])
    return jsonify(result)


@app.route("/api/v6/pl/summary")
@login_required
@role_required('admin', 'sales')
def api_v6_pl_summary():
    """P&L 损益表"""
    period_id = request.args.get("period_id", type=int)
    from pl_engine import PLEngine
    if not period_id:
        periods = PLEngine().get_periods()
        if not periods:
            return jsonify({"error": "无会计期间，请先创建"}), 400
        period_id = periods[0]["id"]
    return jsonify(PLEngine().generate_pl(period_id))


@app.route("/api/v6/pl/trend")
@login_required
@role_required('admin', 'sales')
def api_v6_pl_trend():
    """P&L 趋势"""
    months = request.args.get("months", 6, type=int)
    from pl_engine import PLEngine
    return jsonify(PLEngine().get_pl_trend(months))


@app.route("/api/v6/pl/by-dimension")
@login_required
@role_required('admin', 'sales')
def api_v6_pl_by_dimension():
    """按维度钻取"""
    period_id = request.args.get("period_id", type=int)
    dimension = request.args.get("dimension", "category")
    from pl_engine import PLEngine
    if not period_id:
        periods = PLEngine().get_periods()
        if not periods:
            return jsonify([])
        period_id = periods[0]["id"]
    return jsonify(PLEngine().get_pl_by_dimension(period_id, dimension))


@app.route("/api/v6/pl/compare")
@login_required
@role_required('admin', 'sales')
def api_v6_pl_compare():
    """环比分析"""
    current_id = request.args.get("current_id", type=int)
    previous_id = request.args.get("previous_id", type=int)
    from pl_engine import PLEngine
    if not current_id or not previous_id:
        return jsonify({"error": "需要 current_id 和 previous_id"}), 400
    return jsonify(PLEngine().period_over_period(current_id, previous_id))


@app.route("/api/v6/pl/margin-analysis")
@login_required
@role_required('admin', 'sales')
def api_v6_pl_margin():
    """毛利率分析"""
    period_id = request.args.get("period_id", type=int)
    from pl_engine import PLEngine
    if not period_id:
        periods = PLEngine().get_periods()
        if not periods:
            return jsonify({"error": "无会计期间"}), 400
        period_id = periods[0]["id"]
    return jsonify(PLEngine().get_margin_analysis(period_id))


# ---- V6: Invoices ----

@app.route("/api/v6/invoices")
@login_required
@role_required('admin', 'sales')
def api_v6_invoices():
    """获取发票列表"""
    status = request.args.get("status")
    limit = request.args.get("limit", 50, type=int)
    from invoice_engine import InvoiceEngine
    return jsonify(InvoiceEngine().get_invoices(status, limit))


@app.route("/api/v6/invoices/stats")
@login_required
@role_required('admin', 'sales')
def api_v6_invoice_stats():
    """发票统计"""
    from invoice_engine import InvoiceEngine
    return jsonify(InvoiceEngine().get_invoice_stats())


@app.route("/api/v6/invoices/overdue")
@login_required
@role_required('admin', 'sales')
def api_v6_invoices_overdue():
    """逾期发票"""
    from invoice_engine import InvoiceEngine
    return jsonify(InvoiceEngine().get_overdue_invoices())


@app.route("/api/v6/invoices/<int:invoice_id>")
@login_required
@role_required('admin', 'sales')
def api_v6_invoice_detail(invoice_id):
    """发票详情"""
    from invoice_engine import InvoiceEngine
    return jsonify(InvoiceEngine().get_invoice(invoice_id))


@app.route("/api/v6/invoices/create/<int:order_id>", methods=["POST"])
@login_required
@role_required('admin')
def api_v6_invoice_create(order_id):
    """创建发票"""
    from invoice_engine import InvoiceEngine
    user_id = session.get("user_id")
    result = InvoiceEngine().create_invoice(order_id, created_by=user_id)
    return jsonify(result)


@app.route("/api/v6/invoices/<int:invoice_id>/send", methods=["PUT"])
@login_required
@role_required('admin')
def api_v6_invoice_send(invoice_id):
    """发送发票"""
    from invoice_engine import InvoiceEngine
    return jsonify(InvoiceEngine().send_invoice(invoice_id))


@app.route("/api/v6/invoices/<int:invoice_id>/pay", methods=["PUT"])
@login_required
@role_required('admin')
def api_v6_invoice_pay(invoice_id):
    """标记已付"""
    data = request.json or {}
    from invoice_engine import InvoiceEngine
    return jsonify(InvoiceEngine().mark_paid(invoice_id, data.get("payment_id")))


@app.route("/api/v6/invoices/<int:invoice_id>/cancel", methods=["PUT"])
@login_required
@role_required('admin')
def api_v6_invoice_cancel(invoice_id):
    """取消发票"""
    from invoice_engine import InvoiceEngine
    return jsonify(InvoiceEngine().cancel_invoice(invoice_id))


@app.route("/api/v6/invoices/check-overdue", methods=["POST"])
@login_required
@role_required('admin')
def api_v6_invoice_check_overdue():
    """检查并更新逾期"""
    from invoice_engine import InvoiceEngine
    return jsonify(InvoiceEngine().check_and_update_overdue())


# ---- V6: Expenses ----

@app.route("/api/v6/expenses")
@login_required
@role_required('admin', 'sales')
def api_v6_expenses():
    """获取费用列表"""
    category = request.args.get("category")
    status = request.args.get("status")
    start_date = request.args.get("start_date")
    end_date = request.args.get("end_date")
    limit = request.args.get("limit", 50, type=int)
    from expense_engine import ExpenseEngine
    return jsonify(ExpenseEngine().get_expenses(category, status, start_date, end_date, limit))


@app.route("/api/v6/expenses/summary")
@login_required
@role_required('admin', 'sales')
def api_v6_expense_summary():
    """费用汇总"""
    group_by = request.args.get("group_by", "category")
    start_date = request.args.get("start_date")
    end_date = request.args.get("end_date")
    from expense_engine import ExpenseEngine
    return jsonify(ExpenseEngine().get_expense_summary(group_by, start_date, end_date))


@app.route("/api/v6/expenses/trend")
@login_required
@role_required('admin', 'sales')
def api_v6_expense_trend():
    """费用趋势"""
    months = request.args.get("months", 6, type=int)
    from expense_engine import ExpenseEngine
    return jsonify(ExpenseEngine().get_expense_trend(months))


@app.route("/api/v6/expenses", methods=["POST"])
@login_required
@role_required('admin')
def api_v6_add_expense():
    """添加费用"""
    data = request.json
    if not data or "category" not in data or "amount" not in data:
        return jsonify({"error": "缺少 category 或 amount"}), 400
    from expense_engine import ExpenseEngine
    result = ExpenseEngine().add_expense(
        category=data["category"],
        amount=data["amount"],
        currency=data.get("currency", "USD"),
        expense_date=data.get("expense_date"),
        vendor=data.get("vendor", ""),
        paid_by=data.get("paid_by"),
        notes=data.get("notes", ""),
        description=data.get("description", ""),
    )
    return jsonify(result)


@app.route("/api/v6/expenses/<int:expense_id>/approve", methods=["PUT"])
@login_required
@role_required('admin')
def api_v6_expense_approve(expense_id):
    """审批通过"""
    from expense_engine import ExpenseEngine
    user_id = session.get("user_id")
    return jsonify(ExpenseEngine().approve_expense(expense_id, approved_by=user_id))


@app.route("/api/v6/expenses/<int:expense_id>/reject", methods=["PUT"])
@login_required
@role_required('admin')
def api_v6_expense_reject(expense_id):
    """驳回"""
    data = request.json or {}
    from expense_engine import ExpenseEngine
    user_id = session.get("user_id")
    return jsonify(ExpenseEngine().reject_expense(expense_id, approved_by=user_id, reason=data.get("reason", "")))


# ---- V6: Budgets ----

@app.route("/api/v6/budgets")
@login_required
@role_required('admin', 'sales')
def api_v6_budgets():
    """获取预算"""
    period = request.args.get("period")
    category = request.args.get("category")
    from budget_engine import BudgetEngine
    return jsonify(BudgetEngine().get_budgets(period, category))


@app.route("/api/v6/budgets/vs-actual")
@login_required
@role_required('admin', 'sales')
def api_v6_budget_vs_actual():
    """预算 vs 实际"""
    period = request.args.get("period")
    if not period:
        return jsonify({"error": "缺少 period"}), 400
    from budget_engine import BudgetEngine
    return jsonify(BudgetEngine().get_budget_vs_actual(period))


@app.route("/api/v6/budgets/alerts")
@login_required
@role_required('admin', 'sales')
def api_v6_budget_alerts():
    """预算预警"""
    threshold = request.args.get("threshold", 10, type=int)
    from budget_engine import BudgetEngine
    return jsonify(BudgetEngine().get_budget_alerts(threshold))


@app.route("/api/v6/budgets", methods=["POST"])
@login_required
@role_required('admin')
def api_v6_set_budget():
    """设置预算"""
    data = request.json
    if not data or "period" not in data or "category" not in data or "amount" not in data:
        return jsonify({"error": "缺少 period/category/amount"}), 400
    from budget_engine import BudgetEngine
    result = BudgetEngine().set_budget(data["period"], data["category"], data["amount"], data.get("notes", ""))
    return jsonify(result)


@app.route("/api/v6/budgets/auto-generate", methods=["POST"])
@login_required
@role_required('admin')
def api_v6_budget_auto():
    """自动生成预算模板"""
    data = request.json
    if not data or "year" not in data or "months" not in data:
        return jsonify({"error": "缺少 year 或 months"}), 400
    from budget_engine import BudgetEngine
    count = BudgetEngine().auto_generate_months(data["year"], data["months"])
    return jsonify({"generated": count})


# ---- V6: Executive Dashboard ----

@app.route("/api/v6/exec/summary")
@login_required
@role_required('admin', 'sales')
def api_v6_exec_summary():
    """CEO 综合摘要"""
    from executive_dash import ExecutiveDashboard
    return jsonify(ExecutiveDashboard().get_ceo_summary())


@app.route("/api/v6/exec/health-score")
@login_required
@role_required('admin', 'sales')
def api_v6_exec_health():
    """财务健康评分"""
    from executive_dash import ExecutiveDashboard
    return jsonify(ExecutiveDashboard().get_financial_health_score())


@app.route("/api/v6/exec/revenue-waterfall")
@login_required
@role_required('admin', 'sales')
def api_v6_exec_waterfall():
    """收入瀑布图"""
    period_id = request.args.get("period_id", type=int)
    from executive_dash import ExecutiveDashboard
    return jsonify(ExecutiveDashboard().get_revenue_waterfall(period_id))


@app.route("/api/v6/exec/profit-drivers")
@login_required
@role_required('admin', 'sales')
def api_v6_exec_profit_drivers():
    """利润驱动因素"""
    period_id = request.args.get("period_id", type=int)
    from executive_dash import ExecutiveDashboard
    return jsonify(ExecutiveDashboard().get_profit_driver_analysis(period_id))


@app.route("/api/v6/exec/cash-position")
@login_required
@role_required('admin', 'sales')
def api_v6_exec_cash():
    """现金流状况"""
    from executive_dash import ExecutiveDashboard
    return jsonify(ExecutiveDashboard().get_cash_position())


# ==================== V7 — AI Evolution API ====================

@app.route("/api/v7/ai-feedback/rules", methods=["GET"])
@login_required
@role_required('admin', 'sales')
def api_v7_feedback_rules():
    """获取 AI 进化规则列表"""
    cat = request.args.get("category")
    from database import get_ai_feedback_rules
    rules = get_ai_feedback_rules(active_only=False, category=cat)
    return jsonify(rules)


@app.route("/api/v7/ai-feedback/rules", methods=["POST"])
@login_required
@role_required('admin', 'finance')
def api_v7_feedback_add():
    """添加 AI 进化规则"""
    data = request.get_json(force=True)
    from database import add_ai_feedback_rule
    rid = add_ai_feedback_rule(data)
    from ai_evolution import clear_rules_cache
    clear_rules_cache()
    return jsonify({"id": rid, "ok": True})


@app.route("/api/v7/ai-feedback/rules/<int:rule_id>", methods=["PUT"])
@login_required
@role_required('admin', 'finance')
def api_v7_feedback_update(rule_id):
    """更新 AI 进化规则"""
    data = request.get_json(force=True)
    from database import update_ai_feedback_rule
    update_ai_feedback_rule(rule_id, **data)
    from ai_evolution import clear_rules_cache
    clear_rules_cache()
    return jsonify({"ok": True})


@app.route("/api/v7/ai-feedback/rules/<int:rule_id>", methods=["DELETE"])
@login_required
@role_required('admin')
def api_v7_feedback_delete(rule_id):
    """删除 AI 进化规则"""
    from database import delete_ai_feedback_rule
    delete_ai_feedback_rule(rule_id)
    from ai_evolution import clear_rules_cache
    clear_rules_cache()
    return jsonify({"ok": True})


@app.route("/api/v7/ai-feedback/categories", methods=["GET"])
@login_required
def api_v7_feedback_categories():
    """获取规则分类列表"""
    return jsonify([
        {"key": "b2b_rule", "label": "B2B 销售规则", "color": "#e74c3c"},
        {"key": "objection_handling", "label": "异议处理", "color": "#e67e22"},
        {"key": "regional_strategy", "label": "区域策略", "color": "#2ecc71"},
        {"key": "sales_tactic", "label": "销售策略", "color": "#3498db"},
        {"key": "auto_improvement", "label": "自动改进", "color": "#9b59b6"},
        {"key": "general", "label": "通用", "color": "#95a5a6"},
    ])


@app.route("/api/v7/ai-feedback/auto-improve", methods=["POST"])
@login_required
@role_required('admin')
def api_v7_feedback_auto_improve():
    """从最新评估结果自动生成改进规则"""
    from ai_evolution import generate_improvements_from_eval
    count = generate_improvements_from_eval()
    return jsonify({"generated": count, "ok": True})


@app.route("/api/v7/ai-feedback/clear-cache", methods=["POST"])
@login_required
@role_required('admin')
def api_v7_feedback_clear_cache():
    """清除规则缓存（强制重新读取）"""
    from ai_evolution import clear_rules_cache
    clear_rules_cache()
    return jsonify({"ok": True})


# ==================== V7 — AI Test Center API ====================

@app.route("/api/v7/ai-test/personas", methods=["GET"])
@login_required
@role_required('admin', 'sales')
def api_v7_test_personas():
    """获取可用的 AI 客户画像列表"""
    from ai_customer import list_persona_options
    return jsonify(list_persona_options())


@app.route("/api/v7/ai-test/run", methods=["POST"])
@login_required
@role_required('admin', 'sales')
def api_v7_test_run():
    """运行 AI 客户 vs AI 销售模拟对话"""
    data = request.get_json(force=True) or {}
    persona_id = data.get("persona_id", "dubai_distributor")
    max_rounds = int(data.get("max_rounds", 4))

    from ai_customer import get_persona, generate_customer_response
    from ai_engine import analyze_customer_message
    from database import save_ai_test_result
    from ai_evolution import clear_rules_cache

    persona = get_persona(persona_id)
    if not persona:
        return jsonify({"error": f"Persona not found: {persona_id}"}), 404

    conversation = []
    customer_msg = persona["opening"]

    for round_num in range(max_rounds):
        try:
            sales_result = analyze_customer_message(
                text=customer_msg,
                country=persona["country"],
                history=[{"role": "received", "content_en": m["content"]}
                         for m in conversation if m["role"] == "customer"]
            )
        except Exception as e:
            print(f"[TestCenter] AI分析异常: {e}")
            break
        sales_reply = sales_result.get("reply_en", "") or sales_result.get("reply", "") or ""
        if not sales_reply.strip():
            break

        conversation.append({"round": round_num + 1, "role": "sales", "content": sales_reply})

        customer_msg = generate_customer_response(
            persona=persona,
            conversation_history=conversation,
            sales_last_message=sales_reply,
        )
        conversation.append({"round": round_num + 1, "role": "customer", "content": customer_msg, "persona": persona["name"]})

    scores = _evaluate_test_conversation(conversation, persona)
    summary = {
        "persona_id": persona_id, "persona_name": persona["name"],
        "country": persona["country"], "role": persona["role"],
        "rounds": max_rounds, "total_messages": len(conversation),
    }
    generated_count = _auto_generate_rules_from_test(conversation, scores, persona_id, persona)
    clear_rules_cache()

    test_id = save_ai_test_result(persona_id, persona["name"], conversation, summary, scores, generated_count)

    return jsonify({
        "test_id": test_id, "persona": persona["name"],
        "rounds": len(conversation) // 2, "messages": len(conversation),
        "scores": scores, "generated_rules": generated_count,
        "conversation": conversation,
    })


def _evaluate_test_conversation(conversation, persona):
    """评估测试对话质量（规则评分，不调 LLM）"""
    sales_msgs = [m for m in conversation if m["role"] == "sales"]
    all_text = " ".join(m["content"].lower() for m in sales_msgs)

    # B2B 思维 — 是否先问需求再报价
    b2b = 10
    asked_spec = any(w in all_text for w in ["drawing", "dimension", "spec", "measurement", "size", "width", "height", "图纸", "尺寸"])
    asked_env = any(w in all_text for w in ["outdoor", "indoor", "location", "wall", "facade", "环境", "安装"])
    gave_price = any(w in all_text for w in ["$", "usd", "price is", "costs "])
    if gave_price and not asked_spec:
        b2b -= 5
    if not asked_spec:
        b2b -= 3
    if not asked_env:
        b2b -= 2

    # 异议处理
    obj = 10
    has_empathy = any(w in all_text for w in ["understand", "i see", "good point", "fair", "get it", "理解"])
    has_evidence = any(w in all_text for w in ["because", "difference", "compare", "quality", "warranty", "certified", "certification"])
    if not has_empathy:
        obj -= 3
    if not has_evidence:
        obj -= 3

    # 区域适配
    reg = 10
    r = persona.get("region", "")
    if r == "中东" and not any(w in all_text for w in ["shipping", "fob", "dubai", "dhl", "logistics"]):
        reg -= 3
    elif r == "欧洲" and not any(w in all_text for w in ["ce", "ul", "certification", "rohs", "iso"]):
        reg -= 3
    elif r == "北美" and not any(w in all_text for w in ["ul", "lead time", "shipping"]):
        reg -= 2

    # 话术自然度 (扣 ABC 分)
    abc_count = sum(1 for m in sales_msgs if any(w in m["content"].lower() for w in ["reply a", "reply b", "reply c", "choose a", "choose b", "option a", "option b"]))
    sales_score = max(0, 10 - abc_count * 3)

    overall = round((max(0, b2b) + max(0, obj) + max(0, reg) + sales_score) / 4, 1)

    return {
        "b2b_score": max(0, b2b),
        "objection_score": max(0, obj),
        "regional_score": max(0, reg),
        "sales_score": sales_score,
        "abc_count": abc_count,
        "overall": overall,
    }


def _auto_generate_rules_from_test(conversation, scores, persona_id, persona):
    """根据测试结果自动生成改进规则"""
    from database import add_ai_feedback_rule
    count = 0

    if scores.get("b2b_score", 10) < 6:
        add_ai_feedback_rule({
            "category": "auto_improvement",
            "trigger_condition": persona["country"].lower(),
            "action_rule": f"与{persona.get('region','')}客户对话时，必须优先问需求再报价。B2B评分: {scores['b2b_score']}/10",
            "severity": "suggestion",
            "source_scenario": f"test:{persona_id}",
        })
        count += 1

    # ABC 话术检测
    for m in conversation:
        if m["role"] != "sales":
            continue
        if any(w in m["content"].lower() for w in ["reply a", "reply b", "reply c", "choose a", "choose b"]):
            add_ai_feedback_rule({
                "category": "auto_improvement",
                "trigger_condition": persona["name"].lower(),
                "action_rule": f"禁止用ABC选项结尾。问题回复: {m['content'][:120]}",
                "severity": "hard_rule",
                "source_scenario": f"test:{persona_id}",
            })
            count += 1
            break

    return count


@app.route("/api/v7/ai-test/results", methods=["GET"])
@login_required
@role_required('admin', 'sales')
def api_v7_test_results():
    """获取测试历史列表"""
    limit = request.args.get("limit", 20, type=int)
    from database import get_ai_test_results
    return jsonify(get_ai_test_results(limit=limit))


@app.route("/api/v7/ai-test/results/<int:test_id>", methods=["GET"])
@login_required
@role_required('admin', 'sales')
def api_v7_test_result_detail(test_id):
    """获取单次测试详情"""
    from database import get_ai_test_result
    result = get_ai_test_result(test_id)
    if not result:
        return jsonify({"error": "Not found"}), 404
    return jsonify(result)


@app.route("/api/v7/ai-test/generate-rules/<int:test_id>", methods=["POST"])
@login_required
@role_required('admin')
def api_v7_test_generate_rules(test_id):
    """从测试结果生成进化规则"""
    from database import get_ai_test_result, add_ai_feedback_rule
    from ai_evolution import clear_rules_cache

    result = get_ai_test_result(test_id)
    if not result:
        return jsonify({"error": "Not found"}), 404

    conv = result.get("conversation", []) or []
    scores = result.get("scores", {}) or {}
    pid = result.get("persona_id", "")
    pname = result.get("persona_name", "")
    count = 0

    for m in conv:
        if m.get("role") != "sales":
            continue
        c = m.get("content", "")
        if any(w in c.lower() for w in ["reply a", "reply b", "reply c", "choose a", "choose b"]):
            add_ai_feedback_rule({
                "category": "auto_improvement",
                "trigger_condition": pname,
                "action_rule": f"禁止ABC结尾。问题: {c[:120]}",
                "severity": "hard_rule",
                "source_scenario": f"test:{pid}",
            })
            count += 1
            break

    if isinstance(scores, dict) and scores.get("b2b_score", 10) < 6:
        add_ai_feedback_rule({
            "category": "auto_improvement",
            "trigger_condition": pname,
            "action_rule": f"B2B评分{scores['b2b_score']}/10 — 必须先问需求再报价",
            "severity": "suggestion",
            "source_scenario": f"test:{pid}",
        })
        count += 1

    clear_rules_cache()
    return jsonify({"generated": count, "ok": True})


@app.route("/api/v7/sales-state/detect", methods=["POST"])
@login_required
@role_required('admin', 'sales')
def api_v7_sales_state_detect():
    """检测客户销售阶段 + 执行器输出（用于调试/前端展示）"""
    data = request.json or {}
    message = data.get("message", "")
    intent = data.get("intent", "")
    country = data.get("country", "")

    from sales_state import detect_sales_state
    from sales_executor import SalesExecutor

    state_info = detect_sales_state(message=message, intent=intent)
    behavior = SalesExecutor().execute(state_info, message=message, country=country)

    return jsonify({
        "state": state_info["state"],
        "confidence": state_info["confidence"],
        "deal_probability": state_info["deal_probability"],
        "price_tier": state_info["price_tier"],
        "next_action": state_info["next_action"],
        "recommended_quote_type": state_info["recommended_quote_type"],
        "matched_keywords": state_info["matched_keywords"],
        "reason": state_info["reason"],
        "execution": {
            "reply_type": behavior["reply_type"],
            "quote_trigger": behavior["quote_trigger"],
            "whatsapp_action": behavior["whatsapp_action"],
            "urgency_level": behavior["urgency_level"],
            "price_anchor": behavior["price_anchor"],
            "anchor_price": behavior["anchor_price"],
            "requires_risk_framing": behavior["requires_risk_framing"],
            "customer_type": behavior["customer_type"],
        },
    })


@app.route("/api/v7/sales-state/persona-test", methods=["POST"])
@login_required
@role_required('admin')
def api_v7_sales_state_persona_test():
    """批量测试：对每个客户画像的消息检测销售阶段"""
    data = request.json or {}
    persona_id = data.get("persona_id", "")
    from ai_customer import get_persona, generate_customer_response

    results = []
    if persona_id:
        personas = [get_persona(persona_id)]
    else:
        from ai_customer import get_personas
        personas = get_personas()

    for p in personas:
        if not p:
            continue
        msg = p["opening"]
        from sales_state import detect_sales_state
        from sales_executor import SalesExecutor
        si = detect_sales_state(message=msg)
        be = SalesExecutor().execute(si, message=msg, country=p.get("country", ""))
        results.append({
            "persona": p["id"],
            "name": p["name"],
            "country": p["country"],
            "message": msg[:80],
            "state": si["state"],
            "price_tier": si["price_tier"],
            "deal_probability": si["deal_probability"],
            "next_action": si["next_action"],
            "exec_action": be["reply_type"],
            "urgency": be["urgency_level"],
            "quote_trigger": be["quote_trigger"],
        })

    return jsonify({"results": results, "count": len(results)})


@app.route("/api/v7/sales-state/exec-instruction", methods=["POST"])
@login_required
@role_required('admin')
def api_v7_sales_state_instruction():
    """获取销售执行指令（用于查看 AI 会收到什么指令）"""
    data = request.json or {}
    message = data.get("message", "")
    country = data.get("country", "")

    from sales_state import detect_sales_state
    from sales_executor import SalesExecutor

    si = detect_sales_state(message=message)
    be = SalesExecutor().execute(si, message=message, country=country)

    return jsonify({
        "instruction": be["sales_instruction"],
        "risk_framing": be.get("risk_framing", ""),
        "requires_risk_framing": be["requires_risk_framing"],
        "abc_text": be.get("abc_text", ""),
    })


@app.route("/api/v7/revenue/process", methods=["POST"])
@login_required
@role_required('admin', 'sales')
def api_v7_revenue_process():
    """收单引擎：处理客户消息，返回完整成交动作"""
    data = request.json or {}
    message = data.get("message", "")
    intent = data.get("intent", "")
    country = data.get("country", "")
    customer_name = data.get("customer_name", "")

    from revenue_engine import RevenueEngine
    engine = RevenueEngine()
    result = engine.process(
        message=message, intent=intent,
        country=country, customer_name=customer_name,
    )
    return jsonify(result)


@app.route("/api/v7/revenue/conversion-score", methods=["POST"])
@login_required
@role_required('admin')
def api_v7_revenue_conversion_score():
    """计算成交评分（调试用）"""
    data = request.json or {}
    state = data.get("state", "NEW")
    intent = data.get("intent", "")
    price_tier = data.get("price_tier", "UNKNOWN")
    urgency = data.get("urgency", "low")
    from revenue_engine import _calculate_conversion_score
    score = _calculate_conversion_score(state, intent, price_tier, urgency)
    return jsonify({"conversion_score": score, "state": state, "intent": intent})


# ==================== V3 — 自优化销售系统 API ====================

# --- Conversions ---

@app.route("/api/v3/conversions", methods=["GET"])
@login_required
@role_required('admin', 'sales')
def api_v3_conversions():
    """获取转换记录列表"""
    from conversion_tracker import ConversionTracker
    tracker = ConversionTracker()
    state = request.args.get("state", "")
    result = request.args.get("result", "")
    days = request.args.get("days", "")
    limit = int(request.args.get("limit", 50))
    rows = tracker.get_conversions(state=state, result=result, days=days, limit=limit)
    return jsonify(rows)


@app.route("/api/v3/conversions/stats", methods=["GET"])
@login_required
@role_required('admin', 'sales')
def api_v3_conversions_stats():
    """获取聚合成交统计"""
    from conversion_tracker import ConversionTracker
    tracker = ConversionTracker()
    filters = {}
    for k in ("state", "intent", "price_tier", "ab_version", "days"):
        v = request.args.get(k, "")
        if v:
            filters[k] = v
    stats = tracker.get_conversion_rate(filters=filters or None)
    return jsonify(stats)


@app.route("/api/v3/conversions/close", methods=["POST"])
@login_required
@role_required('admin')
def api_v3_conversions_close():
    """关闭一条转换记录（成交/失败）"""
    data = request.json or {}
    cid = data.get("id", 0)
    result = data.get("result", "won")
    revenue = float(data.get("revenue", 0))
    profit = float(data.get("profit", 0))
    lost_reason = data.get("lost_reason", "")
    from revenue_engine import RevenueEngine
    RevenueEngine().close_conversation(cid, result, revenue, profit, lost_reason)
    return jsonify({"ok": True})


# --- A/B Test ---

@app.route("/api/v3/ab-test/performance", methods=["GET"])
@login_required
@role_required('admin', 'sales')
def api_v3_ab_performance():
    """获取 A/B 版本性能数据"""
    from a_b_optimizer import ABOptimizer
    optimizer = ABOptimizer()
    state = request.args.get("state", "")
    days = int(request.args.get("days", 30))
    perf = optimizer.analyze_performance(days=days)
    weights = optimizer.get_current_weights()
    return jsonify({"performance": perf, "weights": weights})


@app.route("/api/v3/ab-test/optimize", methods=["POST"])
@login_required
@role_required('admin')
def api_v3_ab_optimize():
    """运行 A/B 权重调优"""
    from a_b_optimizer import ABOptimizer
    optimizer = ABOptimizer()
    days = int((request.json or {}).get("days", 30))
    result = optimizer.adjust_weights(days=days)
    # 记录优化日志
    from database import get_db
    conn = get_db()
    conn.execute(
        "INSERT INTO v3_optimization_log (optimizer_type, summary, changes_made, details, triggered_by) VALUES (?,?,?,?,?)",
        ("ab_optimizer", result.get("summary", ""), result.get("changes_made", 0),
         '{}', 'manual')
    )
    conn.commit()
    conn.close()
    return jsonify(result)


@app.route("/api/v3/ab-test/weights", methods=["PUT"])
@login_required
@role_required('admin')
def api_v3_ab_weights():
    """手动设置 A/B 权重"""
    data = request.json or {}
    state = data.get("state", "NEEDS_ANALYSIS")
    version = data.get("version", "B")
    weight = float(data.get("weight", 1.0))
    from a_b_optimizer import ABOptimizer
    ABOptimizer().set_manual_weight(state, version, weight)
    return jsonify({"ok": True, "state": state, "version": version, "weight": weight})


# --- Price ---

@app.route("/api/v3/price/performance", methods=["GET"])
@login_required
@role_required('admin', 'sales')
def api_v3_price_performance():
    """获取价格性能数据"""
    from price_optimizer import PriceOptimizer
    optimizer = PriceOptimizer()
    days = int(request.args.get("days", 90))
    perf = optimizer.analyze_price_performance(days=days)
    current = optimizer.get_current_anchors()
    return jsonify({"performance": perf, "current_anchors": current})


@app.route("/api/v3/price/optimize", methods=["POST"])
@login_required
@role_required('admin')
def api_v3_price_optimize():
    """运行价格优化"""
    from price_optimizer import PriceOptimizer
    optimizer = PriceOptimizer()
    recs = optimizer.compute_optimal_prices()
    from database import get_db
    conn = get_db()
    changes = sum(1 for v in recs.values() if v.get("status") == "optimized")
    conn.execute(
        "INSERT INTO v3_optimization_log (optimizer_type, summary, changes_made, details, triggered_by) VALUES (?,?,?,?,?)",
        ("price_optimizer", f"Optimized {changes} tiers", changes, json.dumps(recs, ensure_ascii=False), 'manual')
    )
    conn.commit()
    conn.close()
    return jsonify(recs)


@app.route("/api/v3/price/apply", methods=["POST"])
@login_required
@role_required('admin')
def api_v3_price_apply():
    """将优化后的价格应用到引擎"""
    from price_optimizer import PriceOptimizer
    optimizer = PriceOptimizer()
    data = request.json or {}
    recs = data.get("recommendations")
    result = optimizer.update_price_anchors(recommendations=recs)
    return jsonify(result)


@app.route("/api/v3/price/reset", methods=["POST"])
@login_required
@role_required('admin')
def api_v3_price_reset():
    """重置价格为默认值"""
    from price_optimizer import PriceOptimizer
    PriceOptimizer().reset_to_defaults()
    return jsonify({"ok": True})


# --- Intent ---

@app.route("/api/v3/intent/weights", methods=["GET"])
@login_required
@role_required('admin', 'sales')
def api_v3_intent_weights():
    """获取当前意图权重"""
    from intent_weight_tuner import IntentWeightTuner
    tuner = IntentWeightTuner()
    weights = tuner.get_current_weights()
    return jsonify(weights)


@app.route("/api/v3/intent/performance", methods=["GET"])
@login_required
@role_required('admin', 'sales')
def api_v3_intent_performance():
    """获取意图成交性能"""
    from intent_weight_tuner import IntentWeightTuner
    tuner = IntentWeightTuner()
    days = int(request.args.get("days", 60))
    analysis = tuner.analyze_intent_performance(days=days)
    return jsonify(analysis)


@app.route("/api/v3/intent/optimize", methods=["POST"])
@login_required
@role_required('admin')
def api_v3_intent_optimize():
    """运行意图权重调优"""
    from intent_weight_tuner import IntentWeightTuner
    tuner = IntentWeightTuner()
    dry_run = (request.json or {}).get("dry_run", False)
    result = tuner.update_weights(dry_run=dry_run)
    # 记录优化日志
    from database import get_db
    conn = get_db()
    conn.execute(
        "INSERT INTO v3_optimization_log (optimizer_type, summary, changes_made, details, triggered_by) VALUES (?,?,?,?,?)",
        ("intent_tuner", result.get("summary", ""), result.get("changes_made", 0),
         json.dumps(result.get("changes", []), ensure_ascii=False), 'manual')
    )
    conn.commit()
    conn.close()
    return jsonify(result)


@app.route("/api/v3/intent/weights", methods=["POST"])
@login_required
@role_required('admin')
def api_v3_intent_weights_set():
    """手动设置单个意图权重"""
    data = request.json or {}
    intent = data.get("intent", "")
    weight = int(data.get("weight", 0))
    from intent_weight_tuner import IntentWeightTuner
    IntentWeightTuner().set_manual_weight(intent, weight)
    return jsonify({"ok": True, "intent": intent, "weight": weight})


@app.route("/api/v3/intent/reset", methods=["POST"])
@login_required
@role_required('admin')
def api_v3_intent_reset():
    """重置所有意图权重为默认值"""
    from intent_weight_tuner import IntentWeightTuner
    IntentWeightTuner().reset_weights()
    return jsonify({"ok": True})


# --- Deal Analysis ---

@app.route("/api/v3/deal-analysis", methods=["GET"])
@login_required
@role_required('admin', 'sales')
def api_v3_deal_analysis():
    """获取成交分析摘要"""
    from deal_analyzer import DealAnalyzer
    analyzer = DealAnalyzer()
    days = int(request.args.get("days", 30))
    result = analyzer.analyze_batch(days=days)
    return jsonify(result)


@app.route("/api/v3/deal-analysis/batch", methods=["POST"])
@login_required
@role_required('admin')
def api_v3_deal_analysis_batch():
    """运行批量成交分析"""
    from deal_analyzer import DealAnalyzer
    analyzer = DealAnalyzer()
    days = int((request.json or {}).get("days", 30))
    result = analyzer.analyze_batch(days=days)
    return jsonify(result)


@app.route("/api/v3/deal-analysis/insights", methods=["GET"])
@login_required
@role_required('admin', 'sales')
def api_v3_deal_insights():
    """获取聚合洞察"""
    from deal_analyzer import DealAnalyzer
    analyzer = DealAnalyzer()
    days = int(request.args.get("days", 90))
    insights = analyzer.get_insights(days=days)
    recs = analyzer.generate_recommendations()
    return jsonify({"insights": insights, "recommendations": recs})


# --- Full Optimization ---

@app.route("/api/v3/optimize", methods=["POST"])
@login_required
@role_required('admin')
def api_v3_optimize():
    """运行全优化流程（A/B + 价格 + 意图）"""
    results = {}
    errors = []

    # 1. A/B 优化
    try:
        from a_b_optimizer import ABOptimizer
        ab_result = ABOptimizer().adjust_weights()
        results["ab_optimizer"] = ab_result
    except Exception as e:
        errors.append(f"ab_optimizer: {e}")
        results["ab_optimizer"] = {"error": str(e)}

    # 2. 价格优化
    try:
        from price_optimizer import PriceOptimizer
        po = PriceOptimizer()
        recs = po.compute_optimal_prices()
        price_result = po.update_price_anchors(recommendations=recs)
        results["price_optimizer"] = price_result
    except Exception as e:
        errors.append(f"price_optimizer: {e}")
        results["price_optimizer"] = {"error": str(e)}

    # 3. 意图权重优化
    try:
        from intent_weight_tuner import IntentWeightTuner
        intent_result = IntentWeightTuner().update_weights(dry_run=False)
        results["intent_tuner"] = intent_result
    except Exception as e:
        errors.append(f"intent_tuner: {e}")
        results["intent_tuner"] = {"error": str(e)}

    # 记录优化日志
    from database import get_db
    conn = get_db()
    total_changes = sum(
        r.get("changes_made", 0) for r in results.values() if isinstance(r, dict)
    )
    conn.execute(
        "INSERT INTO v3_optimization_log (optimizer_type, summary, changes_made, details, triggered_by) VALUES (?,?,?,?,?)",
        ("full_optimize", f"Total changes: {total_changes}, Errors: {len(errors)}",
         total_changes, json.dumps(results, ensure_ascii=False), 'manual')
    )
    conn.commit()
    conn.close()

    return jsonify({"results": results, "errors": errors, "total_changes": total_changes})


@app.route("/api/v3/optimize/logs", methods=["GET"])
@login_required
@role_required('admin', 'sales')
def api_v3_optimize_logs():
    """获取优化历史日志"""
    from database import get_db
    conn = get_db()
    limit = int(request.args.get("limit", 20))
    rows = conn.execute(
        "SELECT * FROM v3_optimization_log ORDER BY id DESC LIMIT ?",
        (limit,)
    ).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@app.route("/api/v3/dashboard", methods=["GET"])
@login_required
@role_required('admin', 'sales')
def api_v3_dashboard():
    """V3 仪表盘聚合数据"""
    from conversion_tracker import ConversionTracker
    from a_b_optimizer import ABOptimizer
    from price_optimizer import PriceOptimizer
    from intent_weight_tuner import IntentWeightTuner
    from deal_analyzer import DealAnalyzer
    from database import get_db

    tracker = ConversionTracker()

    # 基础统计
    stats_30d = tracker.get_conversion_rate(filters={"days": "30"})
    stats_7d = tracker.get_conversion_rate(filters={"days": "7"})

    # A/B 权重
    ab_weights = ABOptimizer().get_current_weights()

    # 当前意图权重
    intent_weights = IntentWeightTuner().get_current_weights()

    # 优化日志
    conn = get_db()
    log_rows = conn.execute(
        "SELECT * FROM v3_optimization_log ORDER BY id DESC LIMIT 10"
    ).fetchall()
    conn.close()
    opt_logs = [dict(r) for r in log_rows]

    return jsonify({
        "stats_7d": stats_7d,
        "stats_30d": stats_30d,
        "ab_weights": ab_weights,
        "intent_weights": intent_weights,
        "optimization_logs": opt_logs,
    })


@app.route("/api/reload-knowledge-base", methods=["POST"])
@login_required
@role_required('admin', 'sales')
def api_reload_knowledge_base():
    """重新加载销售知识库"""
    clear_knowledge_base_cache()
    return jsonify({"ok": True, "message": "知识库已重新加载"})


@app.route("/api/regenerate-catalog", methods=["POST"])
@login_required
@role_required('admin')
def api_regenerate_catalog():
    """从数据库重新生成产品目录 → knowledge/prod_产品目录.txt"""
    try:
        path = generate_product_catalog()
        return jsonify({"ok": True, "path": path, "message": "产品目录已重新生成"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/ai/greeting", methods=["POST"])
@login_required
@role_required('admin', 'sales')
def api_ai_greeting():
    """测试AI招呼（用于调试）"""
    data = request.json or {}
    name = data.get("name", "Test Customer")
    country = data.get("country", "")
    result = get_ai_greeting(customer_name=name, country=country)
    return jsonify(result)


@app.route("/api/orders/profit-stats")
@login_required
def api_order_profit_stats():
    group = request.args.get("group")
    return jsonify(get_order_profit_stats(group=group))


@app.route("/api/orders/stats")
@login_required
def api_order_stats():
    return jsonify(get_order_stats())


@app.route("/api/commission-stats")
@login_required
@role_required('admin', 'sales')
def api_commission_stats():
    return jsonify(get_commission_stats())


@app.route("/api/users/<int:uid>/commission", methods=["PUT"])
@login_required
@role_required('admin', 'sales')
def api_update_commission_rate(uid):
    data = request.json or {}
    rate = data.get("commission_rate")
    if rate is not None:
        update_user(uid, {"commission_rate": float(rate)})
        return jsonify({"ok": True})
    return jsonify({"error": "缺少 commission_rate"}), 400


@app.route("/api/production/schedule")
@login_required
def api_production_schedule():
    return jsonify(get_production_schedule())


# ==================== 库存管理 ====================
@app.route("/api/inventory")
@login_required
def api_inventory_items():
    category = request.args.get("category")
    low_stock = request.args.get("low_stock", "false").lower() == "true"
    return jsonify(get_inventory_items(category=category, low_stock=low_stock))


@app.route("/api/inventory/summary")
@login_required
def api_inventory_summary():
    return jsonify(get_inventory_summary())


@app.route("/api/inventory/<int:iid>")
@login_required
def api_inventory_item(iid):
    item = get_inventory_item(iid)
    if not item:
        return jsonify({"error": "物料不存在"}), 404
    return jsonify(item)


@app.route("/api/inventory", methods=["POST"])
@login_required
def api_add_inventory_item():
    data = request.json or {}
    result = add_inventory_item(data)
    return jsonify({"ok": True, "id": result["id"]})


@app.route("/api/inventory/<int:iid>", methods=["PUT"])
@login_required
def api_update_inventory_item(iid):
    data = request.json or {}
    update_inventory_item(iid, data)
    return jsonify({"ok": True})


@app.route("/api/inventory/<int:iid>", methods=["DELETE"])
@login_required
def api_delete_inventory_item(iid):
    delete_inventory_item(iid)
    return jsonify({"ok": True})


@app.route("/api/inventory/movements")
@login_required
def api_all_stock_movements():
    return jsonify(get_stock_movements(item_id=None))


@app.route("/api/inventory/<int:iid>/movements")
@login_required
def api_stock_movements(iid):
    return jsonify(get_stock_movements(item_id=iid))


@app.route("/api/inventory/<int:iid>/movements", methods=["POST"])
@login_required
def api_add_stock_movement(iid):
    data = request.json or {}
    mtype = data.get("type", "in")
    quantity = float(data.get("quantity", 0))
    if quantity <= 0:
        return jsonify({"error": "数量必须大于0"}), 400
    add_stock_movement(
        item_id=iid,
        mtype=mtype,
        quantity=quantity,
        reference=data.get("reference", ""),
        notes=data.get("notes", ""),
        created_by=session.get("user_id"),
    )
    return jsonify({"ok": True})




# ==================== 数据导出 ====================
@app.route("/api/export/customers")
@login_required
def api_export_customers():
    customers = get_customers()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "客户列表"
    ws.append(["ID", "姓名", "公司", "WhatsApp", "国家", "语言", "状态",
               "来源", "活动", "潜客阶段", "潜客评分", "负责人",
               "首次联系", "最近联系", "创建人", "备注", "创建时间", "更新时间"])
    for c in customers:
        # 查找负责人姓名
        assigned_name = ""
        if c.get("assigned_to"):
            u = get_user_by_id(c["assigned_to"])
            if u: assigned_name = u.get("display_name", "")
        ws.append([
            c.get("id", ""),
            c.get("name", ""),
            c.get("company", ""),
            c.get("whatsapp", ""),
            c.get("country", ""),
            c.get("language", "English"),
            c.get("status", ""),
            c.get("source", ""),
            c.get("campaign", ""),
            c.get("lead_status", ""),
            c.get("lead_score", 0),
            assigned_name,
            c.get("first_contacted_at", ""),
            c.get("last_contacted_at", ""),
            c.get("created_by", ""),
            c.get("notes", ""),
            c.get("created_at", ""),
            c.get("updated_at", ""),
        ])
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return _send_excel(output, "客户导出.xlsx")


@app.route("/api/export/orders")
@login_required
def api_export_orders():
    orders = get_orders()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "订单列表"
    ws.append(["订单号", "客户", "金额", "币种", "状态", "成本", "利润", "利润率%", "创建时间"])
    for o in orders:
        ws.append([
            o.get("order_no", ""),
            o.get("customer_name", ""),
            o.get("total_amount", 0),
            o.get("currency", "USD"),
            o.get("status", ""),
            o.get("partner_cost", 0),
            o.get("profit", 0),
            o.get("profit_margin", 0),
            o.get("created_at", ""),
        ])
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return _send_excel(output, "订单导出.xlsx")


@app.route("/api/export/inventory")
@login_required
def api_export_inventory():
    items = get_inventory_items()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "库存清单"
    ws.append(["物料名称", "分类", "单位", "库存量", "警戒线", "单价", "库存价值", "备注"])
    for item in items:
        ws.append([
            item.get("name", ""),
            item.get("category", ""),
            item.get("unit", ""),
            item.get("quantity", 0),
            item.get("reorder_level", 0),
            item.get("unit_cost", 0),
            item.get("stock_value", 0),
            item.get("notes", ""),
        ])
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return _send_excel(output, "库存导出.xlsx")


@app.route("/api/export/commission")
@login_required
@role_required('admin', 'sales')
def api_export_commission():
    data = get_commission_stats()
    wb = openpyxl.Workbook()
    # Sheet 1: 按销售员汇总
    ws1 = wb.active
    ws1.title = "提成汇总"
    ws1.append(["销售员", "订单数", "营收", "成本", "利润", "提成率%", "提成金额"])
    for s in data.get("sales_summary", []):
        ws1.append([
            s.get("sales_name", ""),
            s.get("order_count", 0),
            s.get("total_revenue", 0),
            s.get("total_cost", 0),
            s.get("total_profit", 0),
            s.get("commission_rate", 0),
            s.get("total_commission", 0),
        ])
    # Sheet 2: 明细
    ws2 = wb.create_sheet("提成明细")
    ws2.append(["订单号", "客户", "销售员", "金额", "成本", "利润", "提成率%", "提成", "日期"])
    for d in data.get("detail", []):
        ws2.append([
            d.get("order_no", ""),
            d.get("customer_name", ""),
            d.get("sales_name", ""),
            d.get("total_amount", 0),
            d.get("partner_cost", 0),
            d.get("profit", 0),
            d.get("commission_rate", 0),
            d.get("commission", 0),
            d.get("created_at", ""),
        ])
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return _send_excel(output, "提成导出.xlsx")


def _send_excel(output, filename):
    from flask import send_file
    output.seek(0)
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


# ==================== 客户导入 ====================
import csv

COLUMN_ALIASES = {
    "name": ["name", "姓名", "名字", "客户名", "客户名称"],
    "company": ["company", "公司", "企业", "单位"],
    "whatsapp": ["whatsapp", "whatsapp号", "手机", "phone", "电话"],
    "country": ["country", "国家", "地区", "region"],
    "language": ["language", "语言", "语种"],
    "status": ["status", "状态", "客户状态"],
    "source": ["source", "来源", "渠道"],
    "campaign": ["campaign", "活动", "营销活动"],
    "lead_status": ["lead_status", "潜客阶段", "阶段", "lead status"],
    "lead_score": ["lead_score", "潜客评分", "评分", "分数", "score"],
    "notes": ["notes", "备注", "备注说明"],
}

def _resolve_csv_col(header):
    h = header.strip().lower()
    for field, aliases in COLUMN_ALIASES.items():
        if h in [a.lower() for a in aliases]:
            return field
    return None

@app.route("/api/import/customers", methods=["POST"])
@login_required
def api_import_customers():
    if "file" not in request.files:
        return jsonify({"ok": False, "error": "请上传CSV文件"}), 400
    f = request.files["file"]
    if not f.filename.endswith(".csv"):
        return jsonify({"ok": False, "error": "仅支持CSV格式"}), 400
    try:
        stream = io.StringIO(f.stream.read().decode("utf-8-sig"))
    except Exception:
        stream = io.StringIO(f.stream.read().decode("gbk"))
    reader = csv.reader(stream)
    try:
        headers = next(reader)
    except StopIteration:
        return jsonify({"ok": False, "error": "CSV文件为空"}), 400
    # 映射列头到字段名
    col_map = {}
    for h in headers:
        fld = _resolve_csv_col(h)
        if fld:
            col_map[fld] = len(col_map)  # use index
    if "name" not in col_map:
        return jsonify({"ok": False, "error": "CSV缺少必要列: name/姓名"}), 400
    # 重新构建 col_map: 字段名 → 列索引
    col_map = {}
    for idx, h in enumerate(headers):
        fld = _resolve_csv_col(h)
        if fld:
            col_map[fld] = idx
    if "name" not in col_map:
        return jsonify({"ok": False, "error": "CSV缺少必要列: name/姓名"}), 400
    rows = []
    for row in reader:
        if not any(cell.strip() for cell in row):
            continue  # 跳过空行
        d = {}
        for fld, idx in col_map.items():
            if idx < len(row):
                d[fld] = row[idx].strip()
        rows.append(d)
    if not rows:
        return jsonify({"ok": False, "error": "CSV无有效数据"}), 400
    imported, skipped, errors = bulk_add_customers(rows)
    return jsonify({
        "ok": True,
        "imported": imported,
        "skipped": skipped,
        "total": len(rows),
        "errors": errors[:10],  # 最多返回10个错误
    })


# ==================== 系统备份 ====================
import shutil

BACKUP_DIR = os.path.join(BASE_DIR, "backups")
os.makedirs(BACKUP_DIR, exist_ok=True)


@app.route("/api/backup", methods=["POST"])
@login_required
def api_create_backup():
    from database import DB_PATH
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"crm_backup_{ts}.db"
    dest = os.path.join(BACKUP_DIR, filename)
    try:
        shutil.copy2(DB_PATH, dest)
        # Keep only last 20 backups
        backups = sorted(
            [f for f in os.listdir(BACKUP_DIR) if f.endswith(".db")],
            reverse=True
        )
        for old in backups[20:]:
            try:
                os.remove(os.path.join(BACKUP_DIR, old))
            except:
                pass
        return jsonify({"ok": True, "filename": filename, "size": os.path.getsize(dest)})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/backups")
@login_required
def api_list_backups():
    backups = []
    if os.path.isdir(BACKUP_DIR):
        for f in sorted(os.listdir(BACKUP_DIR), reverse=True):
            if f.endswith(".db"):
                fp = os.path.join(BACKUP_DIR, f)
                backups.append({
                    "filename": f,
                    "size": os.path.getsize(fp),
                    "created_at": datetime.fromtimestamp(os.path.getmtime(fp)).strftime("%Y-%m-%d %H:%M:%S"),
                })
    return jsonify(backups)


# ==================== 合作方管理 ====================
@app.route("/api/partners")
@login_required
@role_required('admin', 'sales')
def api_partners():
    return jsonify(get_partners())


@app.route("/api/partners/<int:pid>")
@login_required
@role_required('admin', 'sales')
def api_partner(pid):
    partner = get_partner(pid)
    if not partner:
        return jsonify({"error": "合作方不存在"}), 404
    return jsonify(partner)


@app.route("/api/partners", methods=["POST"])
@login_required
@role_required('admin', 'sales')
def api_add_partner():
    data = request.json or {}
    result = add_partner(data)
    if "error" in result:
        return jsonify(result), 400
    return jsonify({"ok": True, "id": result["id"]})


@app.route("/api/partners/<int:pid>", methods=["PUT"])
@login_required
@role_required('admin', 'sales')
def api_update_partner(pid):
    data = request.json or {}
    update_partner(pid, data)
    return jsonify({"ok": True})


@app.route("/api/partners/<int:pid>", methods=["DELETE"])
@login_required
@role_required('admin', 'sales')
def api_delete_partner(pid):
    delete_partner(pid)
    return jsonify({"ok": True})


# ==================== 采购订单管理 ====================
@app.route("/api/purchase-orders")
@login_required
@role_required('admin', 'sales')
def api_purchase_orders():
    status = request.args.get("status")
    partner_id = request.args.get("partner_id")
    order_id = request.args.get("order_id")
    if status:
        status = status.split(",")
    return jsonify(get_purchase_orders(status=status, partner_id=partner_id, order_id=order_id))


@app.route("/api/purchase-orders/<int:oid>")
@login_required
@role_required('admin', 'sales')
def api_purchase_order(oid):
    po = get_purchase_order(oid)
    if not po:
        return jsonify({"error": "采购单不存在"}), 404
    return jsonify(po)


@app.route("/api/purchase-orders", methods=["POST"])
@login_required
@role_required('admin', 'sales')
def api_create_purchase_order():
    data = request.json or {}
    data["created_by"] = session.get("user_id")
    result = add_purchase_order(data)
    if "error" in result:
        return jsonify(result), 400
    return jsonify({"ok": True, "id": result["id"], "po_no": result["po_no"]})


@app.route("/api/purchase-orders/<int:oid>", methods=["PUT"])
@login_required
@role_required('admin', 'sales')
def api_update_purchase_order(oid):
    data = request.json or {}
    if "status" in data:
        existing = get_purchase_order(oid)
        if existing and existing.get("status") != data["status"]:
            add_po_timeline_entry(oid, data["status"], session.get("user_id", 0),
                                  f"状态变更: {existing.get('status','')} → {data['status']}")
    update_purchase_order(oid, data)
    return jsonify({"ok": True})


@app.route("/api/purchase-orders/<int:oid>", methods=["DELETE"])
@login_required
@role_required('admin', 'sales')
def api_delete_purchase_order(oid):
    delete_purchase_order(oid)
    return jsonify({"ok": True})


# ==================== 单实例保护 & 稳定启动 ====================
LOCK_DIR = os.path.join(BASE_DIR, ".crm_lock")
LOCK_FILE = os.path.join(LOCK_DIR, "app.lock")
PID_FILE = os.path.join(LOCK_DIR, "app.pid")


def _ensure_single_instance(port):
    """确保只有一个实例在运行。用socket端口锁定 + PID文件双保险"""
    os.makedirs(LOCK_DIR, exist_ok=True)

    # 1. 检查PID文件，杀掉旧进程
    if os.path.exists(PID_FILE):
        try:
            with open(PID_FILE) as f:
                old_pid = int(f.read().strip())
            # 检查进程是否还在
            if sys.platform == "win32":
                chk = subprocess.run(f"tasklist /FI \"PID eq {old_pid}\" /NH",
                                     shell=True, capture_output=True, text=True)
                if str(old_pid) in chk.stdout:
                    print(f"[启动] 发现旧进程 PID={old_pid}，正在终止...")
                    subprocess.run(f"taskkill /F /PID {old_pid}", shell=True)
                    time.sleep(1)
            else:
                try:
                    os.kill(old_pid, 0)  # 检查进程存在
                    print(f"[启动] 发现旧进程 PID={old_pid}，正在终止...")
                    os.kill(old_pid, signal.SIGTERM)
                    time.sleep(1)
                except OSError:
                    pass  # 进程已死
        except Exception:
            pass

    # 2. 尝试绑定端口（防冲突）
    sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    try:
        sock.bind(("127.0.0.1", port))
        sock.close()
    except OSError:
        print(f"[启动] 端口 {port} 已被占用！尝试关闭旧进程...")
        if sys.platform == "win32":
            subprocess.run(f"netstat -ano | findstr :{port}",
                           shell=True)
        sys.exit(1)

    # 3. 写当前PID
    with open(PID_FILE, "w") as f:
        f.write(str(os.getpid()))
    print(f"[启动] PID={os.getpid()}")


def _cleanup():
    """退出时清理锁文件"""
    for f in [LOCK_FILE, PID_FILE]:
        try:
            if os.path.exists(f):
                os.remove(f)
        except Exception:
            pass
    # 清理空的锁目录
    try:
        if os.path.exists(LOCK_DIR) and not os.listdir(LOCK_DIR):
            os.rmdir(LOCK_DIR)
    except Exception:
        pass


def _kill_orphan_playwright():
    """杀掉残留的Playwright浏览器进程（避免多个窗口）"""
    try:
        if sys.platform == "win32":
            result = subprocess.run(
                'wmic process where "name=\'chrome.exe\'" get commandline /format:csv',
                shell=True, capture_output=True, text=True, timeout=10
            )
            for line in result.stdout.split("\n"):
                if "playwright" in line.lower() and "persistent_profile" in line:
                    # 提取PID
                    pid_line = line.strip()
                    if pid_line:
                        print(f"[启动] 发现残留Playwright: {pid_line[:80]}...")
                        subprocess.run(
                            'wmic process where "name=\'chrome.exe\'" delete',
                            shell=True, capture_output=True, timeout=10
                        )
                        break
    except Exception:
        pass


# ==================== Lead State Engine 初始化 ====================
def _init_lead_state_engine():
    """初始化客户状态机：确保表字段存在 + 注册动作回调"""
    import lead_state_engine as lse
    import action_router as ar

    # 确保数据库有 lead_state 字段 + state_log 表
    lse._ensure_state_field()
    lse._ensure_state_log_table()
    ar._ensure_action_log()
    print("[LeadState] 客户状态引擎已就绪")

    # ---- 注册动作回调 ----
    def _callback_generate_quote(cid, ctx):
        """预生成报价（基于客户消息用假设引擎估算）"""
        chat_name = ctx.get("chat_name", "")
        intent = ctx.get("intent", "")
        analysis = ctx.get("analysis", {})
        print(f"[Action] GENERATE_QUOTE 客户#{cid} {chat_name} (intent={intent})")

        try:
            # 1. 获取客户信息和最新消息
            c = get_customer(cid)
            if not c:
                return {"ok": False, "error": "客户不存在", "detail": f"客户#{cid}不存在"}

            msgs = get_messages(cid, limit=5)
            latest_text = ""
            for m in msgs:
                if m["direction"] == "received":
                    latest_text = m.get("content_en") or m.get("content_cn") or m.get("text", "") or ""
                    break

            # 2. 用假设引擎生成报价
            from assumption_engine import generate_quote
            customer_input = {"text": latest_text or chat_name}
            # 如果有analysis里的翻译文本，用更完整的
            reply_text = ctx.get("reply_text", "")
            if reply_text and not latest_text:
                customer_input["text"] = reply_text

            q_result = generate_quote(customer_input)
            price_min, price_max, currency = q_result.get("price_range", (0, 0, "USD"))
            mode = q_result.get("mode", "ASSUMPTION_BASED")
            assumption = q_result.get("assumption", {})

            # 3. 用利润守卫检查价格区间
            try:
                from profit_guard import estimate_cost, ProfitGuardEngine
                cost = estimate_cost(assumption)
                guard = ProfitGuardEngine()
                safe_check = guard.evaluate_range(cost, price_min, price_max)
                level = safe_check.get("level", "SAFE")
            except Exception:
                level = "UNKNOWN"

            print(f"[Action] 报价估算: USD {price_min}–{price_max}, mode={mode}, profit_level={level}")

            # 4. 存入数据库（草稿状态）
            ref_no = f"BH-{datetime.now().strftime('%Y%m%d')}-{chat_name[:4].upper()}-AUTO"
            from datetime import date, timedelta
            quote_data = {
                "customer_id": cid,
                "quote_no": ref_no,
                "currency": currency,
                "total_amount": price_max,  # 用上限作为预估总额
                "valid_until": (date.today() + timedelta(days=15)).isoformat(),
                "status": "draft",
                "notes": f"AI自动生成（{mode}）: USD {price_min}–{price_max}, 利润等级={level}, 意图={intent}",
                "created_by": 1,  # admin
                "items": [{
                    "product_id": "",
                    "name": f"{assumption.get('sign_type', 'LED Sign')} ({assumption.get('description', '')})",
                    "qty": 1,
                    "unit": "set",
                    "unit_price": price_max,
                    "total": price_max
                }],
            }
            db_result = add_quote(quote_data)
            qid = db_result.get("id")
            quote_no = db_result.get("quote_no", ref_no)

            detail = (f"客户#{cid} {chat_name}: 估USD {price_min}–{price_max} "
                      f"({mode}, {level}), 报价单#{quote_no}")
            print(f"[Action] ✅ {detail}")

            return {
                "ok": True,
                "result": "报价已生成",
                "detail": detail,
                "quote_id": qid,
                "quote_no": quote_no,
                "price_min": price_min,
                "price_max": price_max,
                "currency": currency,
                "profit_level": level,
            }

        except Exception as e:
            print(f"[Action] ❌ GENERATE_QUOTE 失败: {e}")
            return {"ok": False, "error": str(e), "detail": f"报价生成异常: {e}"}

    def _callback_follow_up(cid, ctx):
        """标记跟进"""
        chat_name = ctx.get("chat_name", "")
        print(f"[Action] FOLLOW_UP 客户#{cid} {chat_name}")
        return {"ok": True, "result": "已标记跟进", "detail": f"客户#{cid}跟进已记录"}

    def _callback_escalate(cid, ctx):
        """升级到人工"""
        chat_name = ctx.get("chat_name", "")
        print(f"[Action] ESCALATE 客户#{cid} {chat_name} — 需人工处理")
        return {"ok": True, "result": "已升级", "detail": "通知管理员待处理"}

    def _callback_send_quote(cid, ctx):
        """生成报价并通过WhatsApp发送给客户"""
        chat_name = ctx.get("chat_name", "")
        # 先调用生成逻辑
        gen_result = _callback_generate_quote(cid, ctx)
        if not gen_result.get("ok"):
            return gen_result

        # 发送WhatsApp消息
        price_min = gen_result.get("price_min", 0)
        price_max = gen_result.get("price_max", 0)
        currency = gen_result.get("currency", "USD")
        quote_no = gen_result.get("quote_no", "")

        sym = "$" if currency == "USD" else "¥"
        msg = (
            f"Here's a quick estimate based on typical configuration:\n"
            f"📋 Reference: {quote_no}\n"
            f"💰 Estimated range: {sym}{price_min} – {sym}{price_max}\n\n"
            f"This is a preliminary quote. Once you confirm the details "
            f"(size, material, quantity), I'll send you the final formal quotation."
        )
        try:
            from whatsapp_engine import send_text
            send_text(msg, contact_name=chat_name)
            # 标记报价为已发送
            try:
                qid = gen_result.get("quote_id")
                if qid:
                    update_quote(qid, {"status": "sent"})
            except Exception:
                pass
            print(f"[Action] ✅ 报价已发送给 {chat_name} ({quote_no})")
            return {
                "ok": True,
                "result": "报价已生成并发送",
                "detail": f"已发送报价 {quote_no} 给 {chat_name} (USD {price_min}–{price_max})",
                "quote_no": quote_no,
            }
        except Exception as e:
            print(f"[Action] ⚠️ 报价已生成但发送失败: {e}")
            return {
                "ok": True,
                "result": "报价已生成，发送失败",
                "detail": f"报价 {quote_no} 已生成，但WhatsApp发送异常: {e}",
                "quote_no": quote_no,
            }

    # 注册
    ar.register_action("GENERATE_QUOTE", _callback_generate_quote)
    ar.register_action("SEND_QUOTE", _callback_send_quote)
    ar.register_action("FOLLOW_UP", _callback_follow_up)
    ar.register_action("ESCALATE", _callback_escalate)
    print(f"[LeadState] 动作回调已注册: GENERATE_QUOTE, SEND_QUOTE, FOLLOW_UP, ESCALATE")


_init_lead_state_engine()


# =================== V7 — Revenue Empire OS ===================

@app.route("/api/v7/acquisition/campaign", methods=["POST"])
@login_required
@role_required('admin', 'sales')
def api_v7_acquisition_campaign():
    """Acquisition Engine: 生成获客活动"""
    data = request.json or {}
    from ai_engine.acquisition_engine import AcquisitionEngine
    return jsonify(AcquisitionEngine.generate_campaign(
        data.get("product", ""), data.get("target", ""), data.get("country", ""),
    ))


@app.route("/api/v7/acquisition/budget", methods=["POST"])
@login_required
@role_required('admin', 'sales')
def api_v7_acquisition_budget():
    """Acquisition Engine: 估算预算"""
    data = request.json or {}
    from ai_engine.acquisition_engine import AcquisitionEngine
    return jsonify(AcquisitionEngine.estimate_budget(
        data.get("channels") or [], int(data.get("duration_days", 30)),
    ))


@app.route("/api/v7/content/plan", methods=["POST"])
@login_required
@role_required('admin', 'sales')
def api_v7_content_plan():
    """Content Factory: 生成内容计划"""
    data = request.json or {}
    from ai_engine.content_factory import ContentFactory
    return jsonify(ContentFactory.batch_generate(
        data.get("product", ""), data.get("target", ""),
        data.get("country", ""), int(data.get("count", 10)),
    ))


@app.route("/api/v7/distribute", methods=["POST"])
@login_required
@role_required('admin', 'sales')
def api_v7_distribute():
    """Channel Distributor: 创建分发计划"""
    data = request.json or {}
    from ai_engine.channel_distributor import ChannelDistributor
    return jsonify(ChannelDistributor.create_distribution_plan(
        data.get("product", ""), data.get("target", ""),
        data.get("country", ""), float(data.get("budget", 1000)),
    ))


@app.route("/api/v7/product/expand", methods=["POST"])
@login_required
@role_required('admin', 'sales')
def api_v7_product_expand():
    """Product Expander: 扩展产品矩阵"""
    data = request.json or {}
    from ai_engine.product_expander import ProductExpander
    return jsonify(ProductExpander.expand(
        data.get("product_id", ""), data.get("industry", ""),
    ))


@app.route("/api/v7/product/categories")
@login_required
@role_required('admin', 'sales')
def api_v7_product_categories():
    """获取所有产品类别"""
    from ai_engine.product_expander import ProductExpander
    return jsonify(ProductExpander.get_all_categories())


@app.route("/api/v7/market/analyze")
@login_required
@role_required('admin', 'sales')
def api_v7_market_analyze():
    """Market Explorer: 分析所有市场"""
    from ai_engine.market_explorer import MarketExplorer
    return jsonify(MarketExplorer.analyze_markets())


@app.route("/api/v7/market/discover")
@login_required
@role_required('admin', 'sales')
def api_v7_market_discover():
    """Market Explorer: 发现新市场"""
    from ai_engine.market_explorer import MarketExplorer
    return jsonify(MarketExplorer.discover_new_markets())


@app.route("/api/v7/market/summary")
@login_required
@role_required('admin', 'sales')
def api_v7_market_summary():
    """Market Explorer: 市场摘要"""
    from ai_engine.market_explorer import MarketExplorer
    return jsonify(MarketExplorer.get_market_summary())


@app.route("/api/v7/ad/optimize", methods=["POST"])
@login_required
@role_required('admin', 'sales')
def api_v7_ad_optimize():
    """Ad Optimizer: 生成优化方案"""
    data = request.json or {}
    from ai_engine.ad_optimizer import AdOptimizer
    return jsonify(AdOptimizer.optimize_campaign(
        data.get("product", ""), data.get("target", ""), data.get("country", ""),
    ))


@app.route("/api/v7/ad/ab-test", methods=["POST"])
@login_required
@role_required('admin', 'sales')
def api_v7_ad_ab_test():
    """Ad Optimizer: 创建A/B测试"""
    data = request.json or {}
    from ai_engine.ad_optimizer import AdOptimizer
    test = AdOptimizer.create_ab_test(
        data.get("platform", ""), data.get("variable", "headline"),
        data.get("base", ""), data.get("variants", []),
    )
    return jsonify(AdOptimizer.simulate_results(test))


@app.route("/api/v7/revenue/insights")
@login_required
@role_required('admin', 'sales')
def api_v7_revenue_insights():
    """Revenue Feedback Loop: 获取洞察"""
    days = request.args.get("days", 30, type=int)
    from ai_engine.revenue_feedback_loop import RevenueFeedbackLoop
    return jsonify(RevenueFeedbackLoop().get_insights(days))


@app.route("/api/v7/revenue/record", methods=["POST"])
@login_required
@role_required('admin', 'sales')
def api_v7_revenue_record():
    """Revenue Feedback Loop: 记录指标"""
    data = request.json or {}
    from ai_engine.revenue_feedback_loop import RevenueFeedbackLoop
    record_id = RevenueFeedbackLoop().record_metric(
        data.get("type", ""), data.get("source", ""),
        data.get("metric", ""), float(data.get("value", 0)),
        data.get("context"),
    )
    return jsonify({"record_id": record_id, "ok": record_id > 0})


@app.route("/api/v7/revenue/auto-optimize", methods=["POST"])
@login_required
@role_required('admin', 'sales')
def api_v7_revenue_auto_optimize():
    """Revenue Feedback Loop: 自动优化"""
    from ai_engine.revenue_feedback_loop import RevenueFeedbackLoop
    return jsonify(RevenueFeedbackLoop().auto_optimize())


@app.route("/api/v7/revenue/learning-curve")
@login_required
@role_required('admin', 'sales')
def api_v7_revenue_learning_curve():
    """Revenue Feedback Loop: 学习曲线"""
    days = request.args.get("days", 90, type=int)
    from ai_engine.revenue_feedback_loop import RevenueFeedbackLoop
    return jsonify(RevenueFeedbackLoop().get_learning_curve(days))


# =================== V8 — Business Universe OS ===================

@app.route("/api/v8/company/generate", methods=["POST"])
@login_required
@role_required('admin', 'sales')
def api_v8_company_generate():
    """Company Factory: 从核心能力生成多公司"""
    data = request.json or {}
    from ai_universe.company_factory import CompanyFactory
    return jsonify(CompanyFactory.generate_companies(
        data.get("base_capability", ""),
        data.get("markets"),  # None=auto-select all markets
    ))


@app.route("/api/v8/company/templates")
@login_required
@role_required('admin', 'sales')
def api_v8_company_templates():
    """Company Factory: 获取行业模板"""
    from ai_universe.company_factory import CompanyFactory
    return jsonify(list(CompanyFactory.COMPANY_TEMPLATES.keys()))


@app.route("/api/v8/clone/opportunities", methods=["POST"])
@login_required
@role_required('admin', 'sales')
def api_v8_clone_opportunities():
    """Business Clone Engine: 发现可复制机会"""
    data = request.json or {}
    from ai_universe.business_clone_engine import find_opportunities
    return jsonify(find_opportunities(
        data.get("source_market", ""),
    ))


@app.route("/api/v8/clone/plan", methods=["POST"])
@login_required
@role_required('admin', 'sales')
def api_v8_clone_plan():
    """Business Clone Engine: 创建复制计划"""
    data = request.json or {}
    from ai_universe.business_clone_engine import create_plan
    return jsonify(create_plan(
        data.get("model_id", ""),
        data.get("target_market", ""),
    ))


@app.route("/api/v8/clone/models")
@login_required
@role_required('admin', 'sales')
def api_v8_clone_models():
    """Business Clone Engine: 获取成功模型"""
    from ai_universe.business_clone_engine import BusinessCloneEngine
    return jsonify(BusinessCloneEngine.get_all_models())


@app.route("/api/v8/capital/allocate", methods=["POST"])
@login_required
@role_required('admin', 'sales')
def api_v8_capital_allocate():
    """Capital Allocator: 资金分配"""
    data = request.json or {}
    from ai_universe.capital_allocator import allocate
    return jsonify(allocate(
        float(data.get("total_capital", 10000)),
        data.get("portfolio"),
    ))


@app.route("/api/v8/capital/rebalance", methods=["POST"])
@login_required
@role_required('admin', 'sales')
def api_v8_capital_rebalance():
    """Capital Allocator: 重新平衡"""
    data = request.json or {}
    from ai_universe.capital_allocator import rebalance
    return jsonify(rebalance(
        data.get("current_allocation", []),
        data.get("performance", []),
    ))


@app.route("/api/v8/brand/generate", methods=["POST"])
@login_required
@role_required('admin', 'sales')
def api_v8_brand_generate():
    """Brand Generator: 生成品牌"""
    data = request.json or {}
    from ai_universe.brand_generator import create_brand
    return jsonify(create_brand(
        data.get("industry", ""),
        data.get("market", ""),
        data.get("style", ""),
    ))


@app.route("/api/v8/market/plan", methods=["POST"])
@login_required
@role_required('admin', 'sales')
def api_v8_market_plan():
    """Market Spinup: 创建市场启动计划"""
    data = request.json or {}
    from ai_universe.market_spinup import MarketSpinup
    return jsonify(MarketSpinup.create_launch_plan(
        data.get("product", ""),
        data.get("target_market", ""),
        data.get("industry", ""),
        float(data.get("budget", 2000)),
    ))


@app.route("/api/v8/market/success-probability", methods=["POST"])
@login_required
@role_required('admin', 'sales')
def api_v8_market_success():
    """Market Spinup: 估算成功率"""
    data = request.json or {}
    from ai_universe.market_spinup import MarketSpinup
    return jsonify(MarketSpinup.estimate_success_probability(
        data.get("product", ""),
        data.get("market", ""),
    ))


@app.route("/api/v8/portfolio/register", methods=["POST"])
@login_required
@role_required('admin', 'sales')
def api_v8_portfolio_register():
    """Portfolio Manager: 注册公司到投资组合"""
    data = request.json or {}
    from ai_universe.portfolio_manager import register
    return jsonify(register(
        data.get("name", ""),
        data.get("market", ""),
        data.get("industry", ""),
        float(data.get("invested", 0)),
    ))


@app.route("/api/v8/portfolio/summary")
@login_required
@role_required('admin', 'sales')
def api_v8_portfolio_summary():
    """Portfolio Manager: 投资组合摘要"""
    from ai_universe.portfolio_manager import portfolio
    return jsonify(portfolio.get_portfolio_summary())


@app.route("/api/v8/portfolio/analyze")
@login_required
@role_required('admin', 'sales')
def api_v8_portfolio_analyze():
    """Portfolio Manager: 分析投资组合健康度"""
    from ai_universe.portfolio_manager import analyze
    return jsonify(analyze())


@app.route("/api/v8/portfolio/growth")
@login_required
@role_required('admin', 'sales')
def api_v8_portfolio_growth():
    """Portfolio Manager: 增长建议"""
    from ai_universe.portfolio_manager import portfolio
    return jsonify(portfolio.get_growth_recommendations())


@app.route("/api/v8/risk/assess", methods=["POST"])
@login_required
@role_required('admin', 'sales')
def api_v8_risk_assess():
    """Risk Balancer: 风险评估"""
    data = request.json or {}
    from ai_universe.risk_balancer import assess
    return jsonify(assess(data.get("portfolio", [])))


@app.route("/api/v8/risk/balance", methods=["POST"])
@login_required
@role_required('admin', 'sales')
def api_v8_risk_balance():
    """Risk Balancer: 自动平衡"""
    data = request.json or {}
    from ai_universe.risk_balancer import auto_balance
    return jsonify(auto_balance(data.get("portfolio", [])))


@app.route("/api/v8/risk/hedge", methods=["POST"])
@login_required
@role_required('admin', 'sales')
def api_v8_risk_hedge():
    """Risk Balancer: 建议对冲策略"""
    data = request.json or {}
    from ai_universe.risk_balancer import RiskBalancer
    return jsonify(RiskBalancer.suggest_hedge(
        data.get("risk_assessment", {}),
    ))


if __name__ == "__main__":
    import webbrowser
    port = 5789

    # 启动前清理
    _kill_orphan_playwright()
    _ensure_single_instance(port)

    # 注册退出清理
    atexit.register(_cleanup)
    if sys.platform != "win32":
        signal.signal(signal.SIGTERM, lambda *a: _cleanup() or sys.exit(0))

    print(f"GLOWFORGE CRM 启动中: http://localhost:{port}")
    webbrowser.open(f"http://localhost:{port}")

    # 后台启动WhatsApp 24×7监控
    def _start_bg():
        time.sleep(3)
        # 先尝试连接独立 WhatsApp 服务（127.0.0.1:15789）
        import urllib.request, json as _json
        remote_url = "http://127.0.0.1:15789"
        try:
            with urllib.request.urlopen(f"{remote_url}/health", timeout=3) as r:
                data = _json.loads(r.read().decode())
                if data.get("ok"):
                    set_remote_server(remote_url)
                    print("[启动] ✓ 已连接到独立 WhatsApp 服务")
                    return
        except Exception:
            print("[启动] 独立 WhatsApp 服务未运行，启动本地引擎...")
        # 回退：本地启动引擎
        try:
            start_monitor(_whatsapp_message_handler)
        except Exception as e:
            print(f"[启动] WhatsApp引擎启动失败: {e}")
    threading.Thread(target=_start_bg, daemon=True).start()

    # ===== 系统监控日志 =====
    _start_time = datetime.now()

    def _log_monitor(msg, logpath):
        try:
            with open(logpath, "a", encoding="utf-8") as f:
                f.write(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] {msg}\n")
        except Exception:
            pass

    def _monitor_loop():
        MONITOR_LOG = os.path.join(BASE_DIR, ".whatsapp_session", "monitor.log")
        os.makedirs(os.path.dirname(MONITOR_LOG), exist_ok=True)
        _log_monitor("=== 系统启动 ===", MONITOR_LOG)

        while True:
            try:
                # 检查 WhatsApp 连通性
                wa_ok = False
                try:
                    import urllib.request
                    with urllib.request.urlopen("http://127.0.0.1:15789/health", timeout=5) as r:
                        wa_ok = r.status == 200
                except Exception:
                    wa_ok = False
                # 检查 AI 引擎
                ai_ok = True
                try:
                    from ai_engine import ask_ali
                    r = ask_ali("Say OK", "", max_tokens=10, timeout=10)
                    if not r:
                        ai_ok = False
                except Exception:
                    ai_ok = False

                stats = get_stats() if 'get_stats' in dir() else {}
                uptime = round((datetime.now() - _start_time).total_seconds() / 3600, 1)
                status = {
                    "time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "uptime_hours": uptime,
                    "customers": stats.get("customers", "?"),
                    "products": stats.get("products", "?"),
                    "quotes": stats.get("quotes", "?"),
                    "whatsapp": "OK" if wa_ok else "OFFLINE",
                    "ai_engine": "OK" if ai_ok else "FAIL",
                }
                _log_monitor(json.dumps(status), MONITOR_LOG)
                print(f"[Monitor] {json.dumps(status, ensure_ascii=False)}")
            except Exception as e:
                _log_monitor(f"[Monitor] 记录异常: {e}", MONITOR_LOG)
            threading.Event().wait(3600)  # 每小时

    threading.Thread(target=_monitor_loop, daemon=True).start()

    # 后台线程：每30分钟回收过期未跟进的潜客
    def _reclaim_pool_loop():
        while True:
            try:
                reclaimed = reclaim_expired_leads(days=14)
                if reclaimed:
                    print(f"[Pool] 自动回收 {len(reclaimed)} 个过期潜客: {[r['name'] for r in reclaimed]}")
            except Exception as e:
                print(f"[Pool] 回收检查异常: {e}")
            threading.Event().wait(1800)

    threading.Thread(target=_reclaim_pool_loop, daemon=True).start()

    # 后台线程：每15分钟检查沉默客户 → COLD状态
    def _lead_state_timeout_check():
        from lead_state_engine import batch_check_timeout
        while True:
            try:
                results = batch_check_timeout()
                if results:
                    for r in results:
                        print(f"[LeadState] 超时沉默: {r['name']} → COLD")
            except Exception as e:
                print(f"[LeadState] 超时检查异常: {e}")
            threading.Event().wait(900)

    threading.Thread(target=_lead_state_timeout_check, daemon=True).start()

    try:
        app.run(host="127.0.0.1", port=port, debug=True, use_reloader=False)
    finally:
        _cleanup()
