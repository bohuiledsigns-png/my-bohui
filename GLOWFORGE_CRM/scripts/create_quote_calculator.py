"""生成博汇发光字报价计算器 — 真正可填数据、可下拉选择、自动计算的 Excel"""
import openpyxl, os, shutil
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

wb = openpyxl.Workbook()

# ==================== 样式 ====================
HDR_FONT = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
HDR_FILL = PatternFill(start_color="1a237e", end_color="1a237e", fill_type="solid")
TITLE_FONT = Font(name="微软雅黑", bold=True, size=14, color="1a237e")
LABEL_FONT = Font(name="微软雅黑", bold=True, size=11)
LABEL_FONT2 = Font(name="微软雅黑", bold=True, size=10)
VALUE_FONT = Font(name="微软雅黑", size=11)
NOTE_FONT = Font(name="微软雅黑", size=9, color="616161")
RESULT_FONT = Font(name="微软雅黑", bold=True, size=11, color="c62828")
BIG_RESULT = Font(name="微软雅黑", bold=True, size=14, color="c62828")
MONEY_FMT = '#,##0.00'
NUM_FMT = '#,##0'
INPUT_FILL = PatternFill(start_color="fff9c4", end_color="fff9c4", fill_type="solid")
RESULT_FILL = PatternFill(start_color="ffebee", end_color="ffebee", fill_type="solid")
CALC_FILL = PatternFill(start_color="e8f5e9", end_color="e8f5e9", fill_type="solid")
SUB_FILL = PatternFill(start_color="e8eaf6", end_color="e8eaf6", fill_type="solid")
GOLD_FILL = PatternFill(start_color="d4af37", end_color="d4af37", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def cell(ws, row, col, value=None, font=VALUE_FONT, fill=None, fmt=None, align_center=True):
    c = ws.cell(row=row, column=col, value=value)
    c.font = font
    c.border = THIN_BORDER
    if fill: c.fill = fill
    if fmt: c.number_format = fmt
    if align_center: c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    return c

# ============================================================
# Sheet 1: 数据字典（隐藏，供 VLOOKUP 引用）
# ============================================================
wsd = wb.active
wsd.title = "数据字典"

# --- 产品类型表 (A-G) ---
hdrs = ["编码","中文名","英文名","定价系数","含亚克力","需电源","RGB"]
for ci, h in enumerate(hdrs, 1):
    cell(wsd, 1, ci, h, HDR_FONT, HDR_FILL)

products = [
    ["front","正面发光字","Front Lit",1.0,"是","是","否"],
    ["back","背面发光字","Halo Backlit",1.1,"是","是","否"],
    ["double","双面发光字","Dual Lit",1.2,"是","是","否"],
    ["bottom","底部发光字","Bottom Lit",1.15,"否","是","否"],
    ["korean","韩式发光字","Korean Style",1.15,"是","是","否"],
    ["rgb","RGB绚彩动态字","RGB Dynamic",1.3,"是","是","是"],
    ["neon","LED霓虹灯字","LED Neon",0.9,"否","是","否"],
    ["mini","迷你发光字","Mini LED",0.8,"否","是","否"],
    ["metal","金属字(无灯)","Solid Metal",0.6,"否","否","否"],
    ["acrylic","亚克力展示品","Acrylic Display",0.7,"否","否","否"],
    ["flat","平雕字(无灯)","Flat Letter",0.55,"否","否","否"],
]
for ri, row in enumerate(products, 2):
    for ci, v in enumerate(row, 1):
        cell(wsd, ri, ci, v, fill=SUB_FILL if ri%2==0 else None)

prod_last = 1 + len(products)

# --- 材质单价表 (starts after gap) ---
ms = prod_last + 3
cell(wsd, ms, 1, "材质单价表", TITLE_FONT)
mh = ms + 1
mhds = ["编码","中文名","英文名","类型","30-59cm\n(元/cm)","60-99cm\n(元/cm)","≥100cm\n(元/㎡)","可电镀颜色"]
for ci, h in enumerate(mhds, 1):
    cell(wsd, mh, ci, h, HDR_FONT, HDR_FILL)

materials = [
    ["ss201","201不锈钢","201 Stainless Steel","不锈钢",5.0,6.0,650,"钛金/玫瑰金/古铜/枪黑/青古铜"],
    ["ss304","304不锈钢","304 Stainless Steel","不锈钢",6.0,7.0,750,"钛金/玫瑰金/古铜/枪黑/青古铜"],
    ["ss316","316不锈钢","316 Stainless Steel","不锈钢",7.0,8.0,850,"无"],
    ["zn","镀锌板","Galvanized Steel","铁类",4.5,5.5,600,"无"],
    ["zn_paint","镀锌板烤漆","Painted Galvanized","铁类",5.0,6.0,650,"无"],
    ["ti_gold","钛金板","Titanium Gold","电镀类",9.0,10.0,1000,"无(电镀色)"],
    ["rose_gold","玫瑰金板","Rose Gold","电镀类",9.0,10.0,1000,"无(电镀色)"],
    ["antique_bz","古铜板","Antique Bronze","电镀类",8.5,9.5,950,"无(电镀色)"],
    ["gun_black","枪黑板","Gun Black","电镀类",8.5,9.5,950,"无(电镀色)"],
    ["green_bz","青古铜板","Green Bronze","电镀类",9.0,10.0,1000,"无(电镀色)"],
]
for ri, row in enumerate(materials):
    r = mh + 1 + ri
    for ci, v in enumerate(row, 1):
        fmt = MONEY_FMT if ci in (5,6,7) else None
        cell(wsd, r, ci, v, fill=SUB_FILL if ri%2==0 else None, fmt=fmt)

mat_last = mh + len(materials)

# --- 色温表 ---
cs = mat_last + 3
cell(wsd, cs, 1, "LED色温表", TITLE_FONT)
ch = cs + 1
ctds = ["序号","色温","色温(K)","适合场景"]
for ci, h in enumerate(ctds, 1):
    cell(wsd, ch, ci, h, HDR_FONT, HDR_FILL)
temps = [
    [1,"暖白","3000K","室内/餐厅/温馨氛围"],
    [2,"自然白","4000K","通用/办公室/商业空间"],
    [3,"冷白","6000K","户外/工厂/明亮效果"],
    [4,"正白光","8000K","高亮招牌/户外广告"],
    [5,"RGB","全彩","炫彩动态/节日装饰"],
]
for ri, row in enumerate(temps):
    r = ch + 1 + ri
    for ci, v in enumerate(row, 1):
        cell(wsd, r, ci, v, fill=SUB_FILL if ri%2==0 else None)

# --- 汇率 ---
exs = ch + 1 + len(temps) + 2
cell(wsd, exs, 1, "汇率参数", TITLE_FONT)
for i, (lbl, val, note) in enumerate([
    ("内部汇率",6.8,"1 USD = 6.8 CNY"),
    ("客户汇率(默认)",6.6,"1 USD = 6.6 CNY"),
]):
    r = exs + 1 + i
    cell(wsd, r, 1, lbl, LABEL_FONT2)
    cell(wsd, r, 2, val, fmt='0.00')
    cell(wsd, r, 3, note, NOTE_FONT)

# Column widths
for c in range(1, 9):
    wsd.column_dimensions[get_column_letter(c)].width = [16,20,26,12,16,16,16,22][c-1]

# ============================================================
# Sheet 2: 报价计算器
# ============================================================
ws = wb.create_sheet("报价计算器")

# Named ranges for VLOOKUP
# PRODUCT table range starts at COLUMN B (Chinese names match dropdown)
# B(1)=中文名 C(2)=英文名 D(3)=定价系数 E(4)=含亚克力 F(5)=需电源 G(6)=RGB
prod_tbl = f"'数据字典'!$B$1:$G${prod_last}"
# MATERIAL table range starts at COLUMN B (Chinese names match dropdown)
# B(1)=中文名 C(2)=英文名 D(3)=类型 E(4)=cm30 F(5)=cm60 G(6)=sqm H(7)=颜色
mat_tbl = f"'数据字典'!$B${mh}:$H${mat_last}"
prod_names = f"'数据字典'!$B$2:$B${prod_last}"
mat_names = f"'数据字典'!$B${mh+1}:$B${mat_last}"

# === TITLE ===
cell(ws, 1, 1, "博汇 GLOWFORGE 发光字报价计算器", TITLE_FONT, GOLD_FILL)
for c in range(2, 9):
    ws.cell(row=1, column=c).fill = GOLD_FILL
    ws.cell(row=1, column=c).border = THIN_BORDER

# === INPUT SECTION ===
row = 3
cell(ws, row, 1, "=== 输入参数 ===", Font(name="微软雅黑", bold=True, size=12, color="1a237e"))
for c in range(2, 9):
    ws.cell(row=row, column=c).border = THIN_BORDER

# Input definitions: (row, label, default, hint, is_dropdown, dv_formula)
inp = [
    (5,  "产品类型", "正面发光字",   "从下拉菜单选择", True,  prod_names),
    (6,  "材质",     "304不锈钢",    "从下拉菜单选择", True,  mat_names),
    (7,  "宽度(cm)", 50,             "输入数字",       False, None),
    (8,  "高度(cm)", 50,             "输入数字",       False, None),
    (9,  "厚度(cm)", 8,              "输入数字",       False, None),
    (10, "数量",     1,              "输入数字",       False, None),
    (11, "加价率(%)", 30,            "默认30",         False, None),
    (12, "USD汇率",  6.6,            "默认6.6",        False, None),
]

for r, lbl, default, hint, is_dd, dvf in inp:
    cell(ws, r, 1, lbl, LABEL_FONT2, INPUT_FILL)
    cell(ws, r, 2, default, VALUE_FONT, INPUT_FILL, fmt=NUM_FMT if isinstance(default, int) else ('0.00' if isinstance(default, float) else None))
    cell(ws, r, 3, hint, NOTE_FONT)
    for c in range(4, 9):
        ws.cell(row=r, column=c).fill = INPUT_FILL
        ws.cell(row=r, column=c).border = THIN_BORDER
    if is_dd:
        dv = DataValidation(type="list", formula1=dvf, allow_blank=True)
        dv.error = "请从下拉菜单选择"
        ws.add_data_validation(dv)
        dv.add(f'B{r}')

# === INTERMEDIATE CALC ===
row = 14
cell(ws, row, 1, "=== 中间计算 ===", Font(name="微软雅黑", bold=True, size=12, color="1a237e"))
for c in range(2, 9):
    ws.cell(row=row, column=c).border = THIN_BORDER

# Calc rows (15+)
# Use named references: B5=product, B6=material, B7=width, B8=height, B9=thick, B10=qty, B11=markup%, B12=usd
cr = 15
calc_items = [
    (cr,   "最大边长 maxLen (cm)", f'=MAX(B7,B8)', NUM_FMT),
    (cr+1, "面积 (㎡)",            f'=B7*B8/10000', '0.0000'),
    (cr+2, "定价系数",             f'=IFERROR(VLOOKUP(B5,{prod_tbl},3,0),1)', '0.000'),
    (cr+3, "cm档单价 (元/cm)",     f'=IF(B{cr}<60,VLOOKUP(B6,{mat_tbl},4,0),IF(B{cr}<100,VLOOKUP(B6,{mat_tbl},5,0),0))', MONEY_FMT),
    (cr+4, "㎡档单价 (元/㎡)",     f'=VLOOKUP(B6,{mat_tbl},6,0)', MONEY_FMT),
    (cr+5, "是否电镀材质",        f'=IF(OR(B6="ti_gold",B6="rose_gold",B6="antique_bz",B6="gun_black",B6="green_bz"),"是","否")', None),
    (cr+6, "是否需电源",          f'=IFERROR(VLOOKUP(B5,{prod_tbl},5,0),"否")', None),
    (cr+7, "是否RGB",             f'=IFERROR(VLOOKUP(B5,{prod_tbl},6,0),"否")', None),
]

for r, lbl, fml, fmt in calc_items:
    cell(ws, r, 1, lbl, LABEL_FONT2, CALC_FILL)
    cell(ws, r, 2, fml, VALUE_FONT, CALC_FILL, fmt=fmt)
    for c in range(3, 9):
        ws.cell(row=r, column=c).fill = CALC_FILL
        ws.cell(row=r, column=c).border = THIN_BORDER

# Reference: B15=maxLen, B16=area, B17=coeff, B18=cmPrice, B19=sqmPrice, B20=isElectro, B21=hasPower, B22=isRGB
ml = "B15"   # maxLen
ar = "B16"   # area
cf = "B17"   # coeff
cp = "B18"   # cm price
sp = "B19"   # sqm price
el = "B20"   # is electroplate
pw = "B21"   # has power
rg = "B22"   # is RGB

# === RESULTS ===
rr = 24
cell(ws, rr, 1, "=== 报价计算结果 ===", Font(name="微软雅黑", bold=True, size=12, color="c62828"))
for c in range(2, 9):
    ws.cell(row=rr, column=c).border = THIN_BORDER

# Row references for calculations
# base price = rr+1
# net price = rr+2
# large area fee = rr+3
# electro fee = rr+4
# crate fee = rr+5
# power fee = rr+6
# accessories = rr+7
# service fee = rr+8
# total = rr+9
# final RMB = rr+10
# final USD = rr+11

rp = rr + 1  # result pointer

# Base price
cell(ws, rp, 1, "基础价 (元)", LABEL_FONT2, RESULT_FILL)
cell(ws, rp, 2, f'=IF({ml}<100,{ml}*{cp}*B10,{ar}*{sp}*B10)', VALUE_FONT, RESULT_FILL, MONEY_FMT)
cell(ws, rp, 3, "maxLen<100? maxLen×cm单价×数量 : 面积×㎡单价×数量", NOTE_FONT)

# Net price
rp += 1
cell(ws, rp, 1, "净价 (元)", RESULT_FONT, RESULT_FILL)
cell(ws, rp, 2, f'=B{rr+1}*{cf}', VALUE_FONT, RESULT_FILL, MONEY_FMT)
cell(ws, rp, 3, "= 基础价 × 定价系数", NOTE_FONT)

# Large area surcharge
rp += 1
cell(ws, rp, 1, "大面积附加费 (元)", LABEL_FONT2, RESULT_FILL)
cell(ws, rp, 2, f'=IF(AND({ml}>=100,{el}="否"),{ar}*100*B10,0)', VALUE_FONT, RESULT_FILL, MONEY_FMT)
cell(ws, rp, 3, "非电镀且maxLen≥100: 面积×100×数量", NOTE_FONT)

# Electroplating fee
rp += 1
cell(ws, rp, 1, "电镀费 (元)", LABEL_FONT2, RESULT_FILL)
cell(ws, rp, 2, f'=IF({el}="是",IF({ml}<100,{ml}*2.5*B10,{ar}*250*B10),0)', VALUE_FONT, RESULT_FILL, MONEY_FMT)
cell(ws, rp, 3, "电镀材质: maxLen<100? maxLen×2.5×数量 : 面积×250×数量", NOTE_FONT)

# Crate fee (estimated)
rp += 1
cell(ws, rp, 1, "木箱费 (元)", LABEL_FONT2, RESULT_FILL)
cell(ws, rp, 2, f'={ar}*B9/100*350*B10', VALUE_FONT, RESULT_FILL, MONEY_FMT)
cell(ws, rp, 3, "面积×厚度/100×350×数量 (估算)", NOTE_FONT)

# Power supply fee
rp += 1
cell(ws, rp, 1, "电源费 (元)", LABEL_FONT2, RESULT_FILL)
cell(ws, rp, 2, f'=IF({pw}="是",IF(AND({rg}="是",(B7+B8)*B10/100<3),360,CEILING((B7+B8)/100*2*B10*12*1.15/400,1)*150),0)', VALUE_FONT, RESULT_FILL, MONEY_FMT)
cell(ws, rp, 3, "需电源? RGB小单360套餐/大单按功率: ceil(周长×12W/m×1.15/400)×150", NOTE_FONT)

# Accessories fee
rp += 1
cell(ws, rp, 1, "配件费 (元)", LABEL_FONT2, RESULT_FILL)
cell(ws, rp, 2, '=B10*3+10', VALUE_FONT, RESULT_FILL, MONEY_FMT)
cell(ws, rp, 3, "螺丝(数量×3)+安装模板(10元)", NOTE_FONT)

# Service fee (10% of sum of specific items: net + large_area + electro + crate)
# Items at: rr+2 (net), rr+3 (large), rr+4 (electro), rr+5 (crate)
rp += 1
svc_row = rp
cell(ws, rp, 1, "服务费 (元)", LABEL_FONT2, RESULT_FILL)
# Build formula referencing specific rows
net_r = rr + 2
la_r = rr + 3
el_r = rr + 4
cr_r = rr + 5
cell(ws, rp, 2, f'=(B{net_r}+B{la_r}+B{el_r}+B{cr_r})*0.1', VALUE_FONT, RESULT_FILL, MONEY_FMT)
cell(ws, rp, 3, "(净价+大面积+电镀+木箱)×10%", NOTE_FONT)

# Total internal cost
rp += 1
tot_row = rp
# Sum from base price (rr+1) through service fee (svc_row)
cell(ws, rp, 1, "内部成本总计 (元)", RESULT_FONT, RESULT_FILL)
cell(ws, rp, 2, f'=SUM(B{rr+1}:B{svc_row})', VALUE_FONT, RESULT_FILL, MONEY_FMT)
cell(ws, rp, 3, "以上各项之和", NOTE_FONT)

# Final RMB
rp += 1
fnl_rmb = rp
cell(ws, rp, 1, "★ 正式报价 (RMB)", RESULT_FONT, GOLD_FILL)
cell(ws, rp, 2, f'=B{tot_row}*(1+B11/100)', BIG_RESULT, GOLD_FILL, MONEY_FMT)
for c in range(3, 9):
    ws.cell(row=rp, column=c).fill = GOLD_FILL
    ws.cell(row=rp, column=c).border = THIN_BORDER
cell(ws, rp, 3, "= 内部成本 × (1+加价率)", NOTE_FONT)

# Final USD
rp += 1
fnl_usd = rp
cell(ws, rp, 1, "★ 正式报价 (USD)", RESULT_FONT, GOLD_FILL)
cell(ws, rp, 2, f'=B{fnl_rmb}/B12', BIG_RESULT, GOLD_FILL, MONEY_FMT)
for c in range(3, 9):
    ws.cell(row=rp, column=c).fill = GOLD_FILL
    ws.cell(row=rp, column=c).border = THIN_BORDER
cell(ws, rp, 3, "= RMB报价 ÷ USD汇率", NOTE_FONT)

# === INSTRUCTIONS ===
nr = rp + 2
cell(ws, nr, 1, "使用说明:", LABEL_FONT)
notes = [
    "1. 黄色格子 = 输入参数，可直接修改或从下拉菜单选择",
    "2. 绿色格子 = 中间计算，自动生成无需修改",
    "3. 红色/金色格子 = 最终报价结果",
    "4. 修改任意黄色格子，所有结果自动更新",
    "5. 不同产品类型/材质切换，公式自动适配",
]
for i, n in enumerate(notes):
    cell(ws, nr+1+i, 1, n, NOTE_FONT)

# === FORMULA REFERENCE ===
nr2 = nr + len(notes) + 2
cell(ws, nr2, 1, "完整公式参考:", LABEL_FONT)
formulas = [
    "净价 = maxLen<100 ? maxLen×cm单价×数量×系数 : 面积×㎡单价×数量×系数",
    "大面积附加费 = 非电镀材质且maxLen≥100cm ? 面积×100×数量 : 0",
    "电镀费 = 电镀材质 ? (maxLen<100 ? maxLen×2.5×数量 : 面积×250×数量) : 0",
    "木箱费(估) = 面积×厚度×350/100 × 数量",
    "电源费 = 需电源 ? (RGB小单360套餐 : 按功率算) : 0",
    "配件费 = 数量×3+10",
    "服务费 = (净价+大面积+电镀+木箱) × 10%",
    "正式报价(RMB) = (以上各项之和) × (1+加价率)",
    "正式报价(USD) = RMB报价 ÷ 汇率",
]
for i, f in enumerate(formulas):
    cell(ws, nr2+1+i, 1, f, Font(name="微软雅黑", size=9, color="616161"))

# Column widths
ws.column_dimensions['A'].width = 26
ws.column_dimensions['B'].width = 24
ws.column_dimensions['C'].width = 50
for c in range(4, 9):
    ws.column_dimensions[get_column_letter(c)].width = 14

# ============================================================
# Save
# ============================================================
output = r'D:\Bohui_Global_Push\GLOWFORGE_CRM\博汇发光字报价计算器.xlsx'
wb.save(output)
print(f"已生成: {output}  ({os.path.getsize(output)/1024:.1f} KB)")

desktop = r'I:\桌面\博汇发光字报价计算器.xlsx'
shutil.copy2(output, desktop)
print(f"已复制到桌面: {desktop}")
